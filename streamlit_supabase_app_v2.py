import io
import json
import math
import mimetypes
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from xml.sax.saxutils import escape

import numpy as np
from io import BytesIO
import openpyxl
from openpyxl import Workbook as XLWorkbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import pydeck as pdk
try:
    import mgrs
except Exception:
    mgrs = None
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from matplotlib.backends.backend_pdf import PdfPages
from supabase import create_client

st.set_page_config(page_title="Spectrum Planner", page_icon="📡", layout="wide", initial_sidebar_state="expanded")


def apply_pcc6_dark_ui():
    """Apply PCC6 command-dashboard style UI."""
    st.markdown("""
    <style>
    :root {
        --pcc-bg: #071116;
        --pcc-panel: #0b1820;
        --pcc-panel-2: #101f29;
        --pcc-border: #233743;
        --pcc-text: #e8f2f5;
        --pcc-muted: #a8bbc2;
        --pcc-blue: #1f5fae;
        --pcc-blue-2: #2f8cff;
        --pcc-green: #22c55e;
        --pcc-red: #ef4444;
        --pcc-amber: #f59e0b;
    }

    html, body, [data-testid="stAppViewContainer"], .stApp {
        background:
            radial-gradient(circle at 20% 0%, rgba(47,140,255,.10), transparent 28%),
            radial-gradient(circle at 90% 10%, rgba(34,197,94,.07), transparent 26%),
            linear-gradient(180deg, #061015 0%, #08131a 45%, #05090d 100%) !important;
        color: var(--pcc-text) !important;
    }

    [data-testid="stHeader"] {
        background: rgba(5, 11, 15, .80) !important;
        border-bottom: 1px solid rgba(255,255,255,.06);
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #071016 0%, #081821 100%) !important;
        border-right: 1px solid var(--pcc-border);
    }

    [data-testid="stSidebar"] * {
        color: var(--pcc-text) !important;
    }

    .block-container {
        padding-top: 1.1rem !important;
        padding-left: 1.2rem !important;
        padding-right: 1.2rem !important;
        max-width: 100% !important;
    }

    h1, h2, h3, h4 {
        color: var(--pcc-text) !important;
        letter-spacing: .02em;
    }

    h1 {
        font-size: 2.25rem !important;
        font-weight: 800 !important;
        margin-bottom: .1rem !important;
    }

    .pcc-title-row {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
        padding: .85rem 1rem;
        margin-bottom: .75rem;
        background: linear-gradient(90deg, rgba(9,22,29,.90), rgba(14,31,40,.66));
        border: 1px solid var(--pcc-border);
        border-radius: 10px;
        box-shadow: 0 10px 30px rgba(0,0,0,.28);
    }

    .pcc-title {
        font-size: 1.85rem;
        font-weight: 900;
        color: #f8fbfc;
        line-height: 1.1;
    }

    .pcc-subtitle {
        color: var(--pcc-muted);
        font-size: .86rem;
        margin-top: .25rem;
    }

    .pcc-cui {
        color: #55ff55;
        font-weight: 800;
        font-size: .82rem;
        border: 1px solid rgba(85,255,85,.35);
        background: rgba(34,197,94,.08);
        border-radius: 999px;
        padding: .35rem .65rem;
        white-space: nowrap;
    }

    div[data-testid="stMetric"] {
        background: linear-gradient(180deg, rgba(15,34,45,.95), rgba(8,20,27,.98));
        border: 1px solid var(--pcc-border);
        border-radius: 10px;
        padding: .65rem .85rem;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.035), 0 8px 22px rgba(0,0,0,.22);
    }

    div[data-testid="stMetric"] label {
        color: var(--pcc-muted) !important;
        font-size: .78rem !important;
        text-transform: uppercase;
        letter-spacing: .04em;
    }

    div[data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-weight: 800 !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 0 !important;
        background: rgba(6,15,21,.82);
        border: 1px solid var(--pcc-border);
        border-radius: 8px;
        padding: 0 !important;
        overflow-x: auto;
    }

    .stTabs [data-baseweb="tab"] {
        height: 42px;
        color: var(--pcc-muted) !important;
        background: rgba(9,18,24,.85);
        border-right: 1px solid var(--pcc-border);
        padding: 0 .85rem !important;
        font-size: .78rem;
        font-weight: 700;
        text-transform: uppercase;
    }

    .stTabs [aria-selected="true"] {
        color: #ffffff !important;
        background: linear-gradient(180deg, #1f5fae, #174779) !important;
        border-bottom: 2px solid #5db5ff !important;
    }

    div[data-testid="stExpander"] {
        background: linear-gradient(180deg, rgba(12,27,36,.88), rgba(8,18,25,.88));
        border: 1px solid var(--pcc-border) !important;
        border-radius: 10px !important;
        overflow: hidden;
        box-shadow: 0 8px 24px rgba(0,0,0,.22);
    }

    div[data-testid="stExpander"] summary {
        color: var(--pcc-text) !important;
        font-weight: 800 !important;
    }

    .stButton > button, .stDownloadButton > button, button[kind="primary"] {
        border-radius: 7px !important;
        border: 1px solid #2d6bb6 !important;
        background: linear-gradient(180deg, #1f5fae, #174a85) !important;
        color: white !important;
        font-weight: 700 !important;
    }

    .stButton > button:hover, .stDownloadButton > button:hover {
        border-color: #6bb7ff !important;
        box-shadow: 0 0 0 2px rgba(47,140,255,.18);
    }

    input, textarea, [data-baseweb="select"] > div {
        background: #091720 !important;
        color: var(--pcc-text) !important;
        border-color: var(--pcc-border) !important;
    }

    [data-testid="stDataFrame"], [data-testid="stTable"] {
        border: 1px solid var(--pcc-border);
        border-radius: 10px;
        overflow: hidden;
        background: rgba(8,20,27,.88);
    }

    .stAlert {
        background: rgba(20,34,42,.92) !important;
        border: 1px solid var(--pcc-border) !important;
        color: var(--pcc-text) !important;
    }

    hr {
        border-color: var(--pcc-border) !important;
    }

    .pcc-panel {
        background: linear-gradient(180deg, rgba(12,27,36,.96), rgba(7,18,25,.96));
        border: 1px solid var(--pcc-border);
        border-radius: 10px;
        padding: .85rem 1rem;
        margin: .55rem 0 .75rem 0;
        box-shadow: 0 12px 28px rgba(0,0,0,.24);
    }

    .pcc-panel-title {
        font-weight: 900;
        color: #f9fbfc;
        text-transform: uppercase;
        font-size: .95rem;
        letter-spacing: .04em;
        margin-bottom: .25rem;
    }

    .pcc-panel-caption {
        color: var(--pcc-muted);
        font-size: .82rem;
    }

    .pcc-status-good { color: var(--pcc-green); font-weight: 800; }
    .pcc-status-risk { color: var(--pcc-amber); font-weight: 800; }
    .pcc-status-bad { color: var(--pcc-red); font-weight: 800; }

    @media (max-width: 900px) {
        .pcc-title-row { flex-direction: column; align-items: flex-start; }
        .pcc-title { font-size: 1.35rem; }
    }
    
    /* V44 command dashboard layout */
    .pcc-command-shell {
        background: #061017;
        border: 1px solid #1c3442;
        border-radius: 10px;
        padding: .65rem;
        margin-bottom: .75rem;
        box-shadow: 0 16px 35px rgba(0,0,0,.35);
    }

    .pcc-command-grid {
        display: grid;
        grid-template-columns: 250px 1fr;
        gap: .65rem;
    }

    .pcc-filter-panel {
        background: linear-gradient(180deg, #0b1a23 0%, #07131a 100%);
        border: 1px solid #263f4d;
        border-radius: 8px;
        padding: .9rem;
        min-height: 680px;
    }

    .pcc-main-panel {
        display: grid;
        grid-template-rows: auto auto;
        gap: .65rem;
    }

    .pcc-chart-panel,
    .pcc-table-panel {
        background: linear-gradient(180deg, #0a1821 0%, #07131a 100%);
        border: 1px solid #263f4d;
        border-radius: 8px;
        padding: .85rem;
        box-shadow: inset 0 1px 0 rgba(255,255,255,.04);
    }

    .pcc-two-col {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: .65rem;
    }

    .pcc-mini-title {
        color: #f8fbfc;
        font-weight: 900;
        font-size: 1.02rem;
        text-transform: uppercase;
        letter-spacing: .03em;
        margin-bottom: .5rem;
    }

    .pcc-mini-title .info-dot {
        color: #2f9cff;
        font-weight: 900;
    }

    .pcc-metric-row {
        display: flex;
        gap: .55rem;
        justify-content: flex-end;
        margin-bottom: .5rem;
    }

    .pcc-small-metric {
        border: 1px solid #235987;
        background: rgba(12,35,51,.82);
        border-radius: 5px;
        padding: .45rem .75rem;
        min-width: 105px;
        text-align: center;
    }

    .pcc-small-metric.red { border-color: #a63a32; background: rgba(80,20,20,.55); }
    .pcc-small-metric.amber { border-color: #b7791f; background: rgba(70,45,10,.55); }
    .pcc-small-metric-label {
        color: #9fc0d0;
        font-size: .72rem;
    }
    .pcc-small-metric-value {
        color: #2f9cff;
        font-size: 1.45rem;
        font-weight: 900;
        line-height: 1.05;
    }
    .pcc-small-metric.red .pcc-small-metric-value { color: #ff554d; }
    .pcc-small-metric.amber .pcc-small-metric-value { color: #ffb020; }

    .pcc-band-bar {
        display: flex;
        border: 1px solid #1f3442;
        border-radius: 6px;
        overflow: hidden;
        margin-top: .65rem;
        background: #071018;
    }
    .pcc-band-chip {
        padding: .7rem .95rem;
        border-right: 1px solid #1f3442;
        color: #cbd8dd;
        font-size: .82rem;
        text-align: center;
        min-width: 98px;
    }
    .pcc-band-chip.active {
        background: linear-gradient(180deg, #1f5fae, #124179);
        color: white;
        font-weight: 800;
    }

    @media (max-width: 1100px) {
        .pcc-command-grid { grid-template-columns: 1fr; }
        .pcc-filter-panel { min-height: auto; }
        .pcc-two-col { grid-template-columns: 1fr; }
    }

    
    /* V45 real Streamlit command dashboard blocks */
    .v45-panel {
        border: 1px solid #263f4d;
        background: linear-gradient(180deg, rgba(8,22,31,.98), rgba(5,15,22,.98));
        border-radius: 8px;
        padding: .85rem;
        margin-bottom: .65rem;
        box-shadow: 0 10px 24px rgba(0,0,0,.25);
    }
    .v45-panel-title {
        font-size: 1.02rem;
        font-weight: 900;
        color: #f7fbff;
        text-transform: uppercase;
        letter-spacing: .035em;
        margin-bottom: .35rem;
    }
    .v45-panel-caption {
        color: #9fb3bd;
        font-size: .78rem;
        margin-bottom: .6rem;
    }
    .v45-filter-title {
        color: #f7fbff;
        font-size: 1rem;
        font-weight: 900;
        margin-bottom: .9rem;
        text-transform: uppercase;
    }
    .v45-filter-box {
        border: 1px solid #263f4d;
        background: linear-gradient(180deg, #0b1a23, #07131a);
        border-radius: 8px;
        padding: .85rem;
        min-height: 650px;
    }
    .v45-chip {
        border: 1px solid #263f4d;
        background: #081821;
        border-radius: 4px;
        padding: .6rem .7rem;
        text-align: center;
        color: #cbd8dd;
        font-size: .8rem;
    }
    .v45-chip-active {
        background: linear-gradient(180deg, #1f5fae, #124179);
        color: #ffffff;
        font-weight: 800;
    }
    .v45-metric-strip {
        display: flex;
        justify-content: flex-end;
        gap: .5rem;
        margin-bottom: .55rem;
    }
    .v45-metric {
        border: 1px solid #235987;
        border-radius: 5px;
        background: rgba(12,35,51,.82);
        min-width: 112px;
        padding: .4rem .65rem;
        text-align: center;
    }
    .v45-metric.red {
        border-color: #a63a32;
        background: rgba(80,20,20,.55);
    }
    .v45-metric.amber {
        border-color: #b7791f;
        background: rgba(70,45,10,.55);
    }
    .v45-metric-label {
        color: #9fc0d0;
        font-size: .68rem;
    }
    .v45-metric-value {
        color: #2f9cff;
        font-size: 1.35rem;
        font-weight: 900;
        line-height: 1.05;
    }
    .v45-metric.red .v45-metric-value { color: #ff554d; }
    .v45-metric.amber .v45-metric-value { color: #ffb020; }

    
    /* V46 FULL WEB APP SHELL */
    .app-shell-topbar {
        position: sticky;
        top: 0;
        z-index: 999;
        display: grid;
        grid-template-columns: auto 1fr auto;
        align-items: center;
        gap: 1rem;
        padding: .7rem .95rem;
        margin: -.35rem 0 .75rem 0;
        background: rgba(3, 10, 15, .92);
        backdrop-filter: blur(12px);
        border: 1px solid #1d3340;
        border-radius: 10px;
        box-shadow: 0 12px 32px rgba(0,0,0,.32);
    }

    .app-brand {
        display: flex;
        align-items: center;
        gap: .65rem;
        white-space: nowrap;
    }

    .app-logo {
        width: 38px;
        height: 38px;
        border-radius: 9px;
        display: grid;
        place-items: center;
        background: linear-gradient(135deg, #1f5fae, #0e9f6e);
        box-shadow: 0 0 0 1px rgba(255,255,255,.14), 0 8px 18px rgba(0,0,0,.35);
        font-size: 1.25rem;
    }

    .app-title-main {
        color: #f8fbff;
        font-size: 1.28rem;
        font-weight: 950;
        letter-spacing: .035em;
        line-height: 1;
    }

    .app-title-sub {
        color: #86a4b2;
        font-size: .72rem;
        margin-top: .18rem;
    }

    .app-classification {
        justify-self: center;
        color: #55ff55;
        font-weight: 850;
        font-size: .77rem;
        border: 1px solid rgba(85,255,85,.38);
        background: rgba(34,197,94,.08);
        border-radius: 999px;
        padding: .32rem .75rem;
        white-space: nowrap;
    }

    .app-actions {
        display: flex;
        align-items: center;
        gap: .5rem;
        justify-content: flex-end;
        color: #dce8ee;
        font-size: .78rem;
        white-space: nowrap;
    }

    .app-action-pill {
        border: 1px solid #284755;
        background: #071922;
        border-radius: 6px;
        padding: .36rem .55rem;
        color: #dce8ee;
    }

    .webapp-layout {
        display: grid;
        grid-template-columns: 230px 1fr;
        gap: .75rem;
        align-items: start;
    }

    .webapp-nav {
        position: sticky;
        top: 82px;
        border: 1px solid #213946;
        background: linear-gradient(180deg, rgba(8,22,31,.98), rgba(4,13,19,.98));
        border-radius: 10px;
        padding: .65rem;
        min-height: 650px;
        box-shadow: 0 14px 30px rgba(0,0,0,.28);
    }

    .webapp-nav-title {
        color: #f7fbff;
        font-size: .78rem;
        font-weight: 900;
        letter-spacing: .08em;
        text-transform: uppercase;
        margin: .25rem .35rem .65rem .35rem;
    }

    .webapp-nav-item {
        display: flex;
        align-items: center;
        gap: .5rem;
        padding: .7rem .72rem;
        border-radius: 7px;
        color: #b9cbd3;
        font-size: .82rem;
        font-weight: 750;
        margin-bottom: .28rem;
        border: 1px solid transparent;
    }

    .webapp-nav-item.active {
        background: linear-gradient(90deg, #1f5fae, rgba(31,95,174,.35));
        color: white;
        border-color: rgba(93,181,255,.42);
        box-shadow: inset 0 1px 0 rgba(255,255,255,.10);
    }

    .webapp-nav-item:hover {
        background: rgba(255,255,255,.04);
        border-color: #263f4d;
    }

    .webapp-content {
        min-width: 0;
    }

    .web-grid-4 {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: .75rem;
        margin-bottom: .75rem;
    }

    .web-card {
        border: 1px solid #223946;
        background: linear-gradient(180deg, rgba(10,28,38,.96), rgba(5,16,23,.96));
        border-radius: 10px;
        padding: .9rem;
        box-shadow: 0 12px 28px rgba(0,0,0,.25);
    }

    .web-card-label {
        color: #9fb3bd;
        font-size: .75rem;
        text-transform: uppercase;
        font-weight: 800;
        letter-spacing: .06em;
    }

    .web-card-value {
        color: #f8fbff;
        font-size: 1.75rem;
        font-weight: 950;
        margin-top: .2rem;
        line-height: 1.05;
    }

    .web-card-sub {
        color: #86a4b2;
        font-size: .75rem;
        margin-top: .28rem;
    }

    .web-card.blue { border-color: rgba(47,140,255,.45); }
    .web-card.red { border-color: rgba(239,68,68,.55); }
    .web-card.amber { border-color: rgba(245,158,11,.55); }
    .web-card.green { border-color: rgba(34,197,94,.55); }

    .web-card.blue .web-card-value { color: #47a3ff; }
    .web-card.red .web-card-value { color: #ff5c55; }
    .web-card.amber .web-card-value { color: #ffb020; }
    .web-card.green .web-card-value { color: #27e070; }

    .web-section {
        border: 1px solid #223946;
        background: linear-gradient(180deg, rgba(7,20,28,.96), rgba(4,13,19,.96));
        border-radius: 10px;
        padding: .85rem;
        margin-bottom: .75rem;
        box-shadow: 0 10px 24px rgba(0,0,0,.24);
    }

    .web-section-header {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: .75rem;
        margin-bottom: .65rem;
        border-bottom: 1px solid rgba(255,255,255,.06);
        padding-bottom: .55rem;
    }

    .web-section-title {
        color: #f8fbff;
        font-size: 1rem;
        font-weight: 930;
        text-transform: uppercase;
        letter-spacing: .04em;
    }

    .web-section-caption {
        color: #91a9b4;
        font-size: .78rem;
    }

    .status-badge {
        display: inline-block;
        border-radius: 999px;
        padding: .18rem .55rem;
        font-size: .72rem;
        font-weight: 850;
        border: 1px solid rgba(255,255,255,.14);
    }

    .status-good { color: #27e070; background: rgba(34,197,94,.10); border-color: rgba(34,197,94,.35); }
    .status-risk { color: #ffb020; background: rgba(245,158,11,.10); border-color: rgba(245,158,11,.35); }
    .status-bad { color: #ff5c55; background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.35); }

    /* Make Streamlit tabs less like default tabs and more like a web nav bar. */
    .stTabs [data-baseweb="tab-list"] {
        position: sticky;
        top: 74px;
        z-index: 50;
        border-radius: 8px !important;
        box-shadow: 0 10px 24px rgba(0,0,0,.20);
    }

    /* Aggressive table polish */
    [data-testid="stDataFrame"] div[role="grid"] {
        background: #06131a !important;
    }

    /* Keep the app wide and hide excess Streamlit whitespace. */
    .main .block-container {
        padding-bottom: 2rem !important;
    }

    @media (max-width: 1200px) {
        .webapp-layout { grid-template-columns: 1fr; }
        .webapp-nav { position: relative; top: auto; min-height: auto; }
        .web-grid-4 { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        .app-shell-topbar { grid-template-columns: 1fr; }
        .app-classification { justify-self: start; }
        .app-actions { justify-content: flex-start; }
    }

    @media (max-width: 700px) {
        .web-grid-4 { grid-template-columns: 1fr; }
    }

    
    /* V49 CLEAN MODE OVERRIDES */
    .v49-clean-title {
        border: 1px solid #223946;
        background: linear-gradient(180deg, rgba(9,24,34,.96), rgba(5,14,20,.96));
        border-radius: 10px;
        padding: 1rem 1.1rem;
        margin-bottom: .75rem;
    }
    .v49-clean-title h1 {
        color: #f8fbff;
        font-size: 1.55rem;
        margin: 0;
        padding: 0;
        font-weight: 900;
        letter-spacing: .02em;
    }
    .v49-clean-title p {
        color: #9fb3bd;
        font-size: .88rem;
        margin: .25rem 0 0 0;
    }
    .v49-clean-note {
        border-left: 4px solid #2f9cff;
        background: rgba(47,156,255,.08);
        padding: .65rem .8rem;
        border-radius: 6px;
        color: #dce8ee;
        margin-bottom: .75rem;
    }
    /* calm down older dashboard styling */
    .web-grid-4, .pcc-command-shell, .pcc-command-grid, .v45-filter-box {
        display: none !important;
    }

    </style>
    """, unsafe_allow_html=True)


st.markdown("""
<style>
.block-container {padding-top: 1rem; padding-bottom: 2rem;}
.stTabs [data-baseweb="tab-list"] {gap: .75rem;}
.stTabs [data-baseweb="tab"] {height: 2.4rem; white-space: nowrap;}
</style>
""", unsafe_allow_html=True)

st.markdown("""<div class="pcc-title-row"><div><div class="pcc-title">📡 PCC6 SPECTRUM PLANNER</div><div class="pcc-subtitle">Collaborative frequency, power, geographic reuse, and time deconfliction workspace</div></div><div class="pcc-cui">CONTROLLED UNCLASSIFIED INFORMATION (CUI)</div></div>""", unsafe_allow_html=True)
st.caption("Collaborative frequency, power, and time deconfliction workspace")

STORAGE_BUCKET = "spectrum-files"



@st.cache_resource
def get_supabase():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])


@st.cache_resource
def get_supabase_admin():
    service_key = st.secrets.get("supabase", {}).get("service_role_key", "")
    if not service_key:
        return None
    return create_client(st.secrets["supabase"]["url"], service_key)


sb = get_supabase()
sb_admin = get_supabase_admin()

APP_COLUMNS = [
    "Active",
    "Start Time", "End Time", "Equipment", "Center Frequency (MHz)",
    "Start Frequency (MHz)", "End Frequency (MHz)", "Bandwidth (MHz)",
    "Power (W)", "Power (dBm)", "Tech", "Unit", "Notes",
    "Latitude", "Longitude", "Location",
    "Antenna Height", "Coverage Radius", "Site Name",
    "MGRS", "USNG",
]

STANDARD_RENAME = {
    "StartTime": "Start Time",
    "EndTime": "End Time",
    "Start Time": "Start Time",
    "End Time": "End Time",

    "Equipment": "Equipment",
    "Equip": "Equipment",
    "Tech": "Tech",
    "Unit": "Unit",
    "Notes": "Notes",

    "CenterF": "Center Frequency (MHz)",
    "Center Frequency": "Center Frequency (MHz)",
    "Center Frequency (MHz)": "Center Frequency (MHz)",
    "Center Freq": "Center Frequency (MHz)",
    "Center Freq (MHz)": "Center Frequency (MHz)",

    "StartF": "Start Frequency (MHz)",
    "Start Frequency": "Start Frequency (MHz)",
    "Start Frequency (MHz)": "Start Frequency (MHz)",
    "Start Freq": "Start Frequency (MHz)",
    "Start Freq (MHz)": "Start Frequency (MHz)",

    "EndF": "End Frequency (MHz)",
    "End Frequency": "End Frequency (MHz)",
    "End Frequency (MHz)": "End Frequency (MHz)",
    "End Freq": "End Frequency (MHz)",
    "End Freq (MHz)": "End Frequency (MHz)",

    "BW": "Bandwidth (MHz)",
    "Bandwidth": "Bandwidth (MHz)",
    "Bandwidth (MHz)": "Bandwidth (MHz)",

    "PowerW": "Power (W)",
    "Power (W)": "Power (W)",
    "PowerdBm": "Power (dBm)",
    "Power (dBm)": "Power (dBm)",

    "Latitude": "Latitude",
    "Lat": "Latitude",
    "LAT": "Latitude",
    "latitude": "Latitude",

    "Longitude": "Longitude",
    "Long": "Longitude",
    "Lon": "Longitude",
    "Lng": "Longitude",
    "LON": "Longitude",
    "longitude": "Longitude",

    "Location": "Location",
    "location": "Location",

    "Antenna Height": "Antenna Height",
    "AntennaHeight": "Antenna Height",
    "Antenna Height (ft)": "Antenna Height",
    "Antenna Height (m)": "Antenna Height",

    "Coverage Radius": "Coverage Radius",
    "CoverageRadius": "Coverage Radius",
    "Coverage Radius (mi)": "Coverage Radius",
    "Coverage Radius (km)": "Coverage Radius",
    "Coverage Radius (NM)": "Coverage Radius",

    "Site Name": "Site Name",
    "SiteName": "Site Name",
    "Site": "Site Name",
    "MGRS": "MGRS",
    "USNG": "USNG",
    "Grid": "MGRS",
    "Military Grid": "MGRS",
    "Active": "Active",
    "Enabled": "Active",
    "In Use": "Active",
    "Use": "Active",
    "Include": "Active",
}

INTERNAL_RENAME = {
    "Start Time": "StartTime",
    "End Time": "EndTime",
    "Equipment": "Equipment",
    "Center Frequency (MHz)": "CenterF",
    "Start Frequency (MHz)": "StartF",
    "End Frequency (MHz)": "EndF",
    "Bandwidth (MHz)": "BW",
    "Power (W)": "PowerW",
    "Power (dBm)": "PowerdBm",
    "Tech": "Tech",
    "Unit": "Unit",
    "Notes": "Notes",
    "Latitude": "Latitude",
    "Longitude": "Longitude",
    "Location": "Location",
    "Antenna Height": "AntennaHeight",
    "Coverage Radius": "CoverageRadius",
    "Site Name": "SiteName",
    "MGRS": "MGRS",
    "USNG": "USNG",
    "Active": "Active",
}

def now_iso():
    return datetime.now(timezone.utc).isoformat()


def json_safe_value(value):
    """Convert Excel/Pandas/Numpy values into Supabase JSON-safe values."""
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()

    if isinstance(value, (np.integer,)):
        return int(value)

    if isinstance(value, (np.floating,)):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(value, (np.bool_, bool)):
        return bool(value)

    if isinstance(value, (int, str)):
        return value

    return str(value)


def json_safe_df(df):
    """Return a copy of df containing only JSON-safe scalar values."""
    safe = df.copy().astype("object")
    safe = safe.where(pd.notnull(safe), None)
    safe = safe.replace({np.nan: None, np.inf: None, -np.inf: None, pd.NaT: None})

    for col in safe.columns:
        safe[col] = safe[col].map(json_safe_value).astype("object")

    safe = safe.where(pd.notnull(safe), None)
    return safe


def json_safe_records(df):
    """Return list-of-dicts guaranteed to contain no NaN/Inf values."""
    safe = json_safe_df(df)
    records = []
    for row in safe.to_dict(orient="records"):
        clean_row = {}
        for key, value in row.items():
            clean_row[str(key)] = json_safe_value(value)
        records.append(clean_row)
    return records

def obj_get(obj, key, default=None):
    """Read from Supabase response objects or dicts."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def user_id_from_user(user):
    return obj_get(user, "id", "")


def user_email_from_user(user):
    return obj_get(user, "email", "")


def count_user_roles():
    try:
        rows = sb.table("user_roles").select("user_id").execute().data or []
        return len(rows)
    except Exception:
        return 0


def get_user_role(user_id, email=None):
    try:
        rows = sb.table("user_roles").select("*").eq("user_id", user_id).limit(1).execute().data or []
        if rows:
            return rows[0]
    except Exception:
        pass

    # First authenticated user becomes admin; later users default to viewer.
    role = "admin" if count_user_roles() == 0 else "viewer"
    rec = {
        "user_id": user_id,
        "email": email or "",
        "role": role,
        "full_name": "",
        "updated_at": now_iso(),
    }
    try:
        sb.table("user_roles").upsert(rec, on_conflict="user_id").execute()
    except Exception:
        pass
    return rec


def update_user_role(user_id, email, role, full_name="", updated_by=""):
    rec = {
        "user_id": user_id,
        "email": email or "",
        "role": role,
        "full_name": full_name or "",
        "updated_by": updated_by,
        "updated_at": now_iso(),
    }
    return sb.table("user_roles").upsert(rec, on_conflict="user_id").execute()


def list_app_users():
    try:
        return sb.table("user_roles").select("*").order("email").execute().data or []
    except Exception:
        return []


def admin_create_user(email, password, role, full_name, created_by):
    if sb_admin is None:
        raise RuntimeError("Missing supabase.service_role_key in Streamlit secrets.")

    response = sb_admin.auth.admin.create_user(
        {
            "email": email,
            "password": password,
            "email_confirm": True,
            "user_metadata": {"full_name": full_name or ""},
        }
    )
    user = obj_get(response, "user", None)
    if user is None:
        user = obj_get(obj_get(response, "data", None), "user", None)

    user_id = user_id_from_user(user)
    if not user_id:
        raise RuntimeError("User was created, but Supabase did not return a user id.")

    update_user_role(user_id, email, role, full_name, created_by)
    return user_id


def admin_delete_user(user_id):
    if sb_admin is None:
        raise RuntimeError("Missing supabase.service_role_key in Streamlit secrets.")
    return sb_admin.auth.admin.delete_user(user_id)


def require_login():
    """Supabase Auth login/signup."""
    if st.session_state.get("auth_user") and st.session_state.get("auth_role"):
        role = st.session_state.get("auth_role", "viewer")
        if role == "disabled":
            st.error("Your account is disabled. Contact an administrator.")
            st.stop()
        return st.session_state["auth_user"], role

    st.markdown("### 🔐 Spectrum Planner Login")
    st.caption("Use your assigned account, or create an account if sign-up is enabled.")

    login_tab, signup_tab = st.tabs(["Login", "Create Account"])

    with login_tab:
        with st.form("login_form", clear_on_submit=False):
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            submitted = st.form_submit_button("Login", type="primary")

        if submitted:
            try:
                response = sb.auth.sign_in_with_password({"email": email.strip(), "password": password})
                user = obj_get(response, "user", None)
                if user is None:
                    user = obj_get(obj_get(response, "data", None), "user", None)

                user_id = user_id_from_user(user)
                user_email = user_email_from_user(user) or email.strip()
                role_rec = get_user_role(user_id, user_email)
                role = role_rec.get("role", "viewer")

                if role == "disabled":
                    st.error("Your account is disabled. Contact an administrator.")
                    st.stop()

                st.session_state["auth_user"] = user_email
                st.session_state["auth_user_id"] = user_id
                st.session_state["auth_role"] = role
                st.session_state["auth_full_name"] = role_rec.get("full_name", "")
                st.rerun()
            except Exception as e:
                st.error(f"Login failed: {e}")

    with signup_tab:
        allow_signup = st.secrets.get("auth", {}).get("allow_signup", True)
        if not allow_signup:
            st.info("Self sign-up is disabled. Ask an administrator to create your account.")
        else:
            with st.form("signup_form", clear_on_submit=False):
                full_name = st.text_input("Full name", key="signup_full_name")
                email = st.text_input("Email", key="signup_email")
                password = st.text_input("Password", type="password", key="signup_password")
                password2 = st.text_input("Confirm password", type="password", key="signup_password2")
                submitted = st.form_submit_button("Create account", type="primary")

            if submitted:
                if not email.strip() or not password:
                    st.error("Email and password are required.")
                elif password != password2:
                    st.error("Passwords do not match.")
                elif len(password) < 8:
                    st.error("Password must be at least 8 characters.")
                else:
                    try:
                        response = sb.auth.sign_up(
                            {
                                "email": email.strip(),
                                "password": password,
                                "options": {"data": {"full_name": full_name.strip()}},
                            }
                        )
                        user = obj_get(response, "user", None)
                        if user is None:
                            user = obj_get(obj_get(response, "data", None), "user", None)

                        user_id = user_id_from_user(user)
                        if user_id:
                            role = "admin" if count_user_roles() == 0 else "viewer"
                            update_user_role(user_id, email.strip(), role, full_name.strip(), "self_signup")

                        st.success("Account created. If email confirmation is enabled in Supabase, confirm your email before logging in.")
                    except Exception as e:
                        st.error(f"Account creation failed: {e}")

    st.stop()


def safe_group(s):
    s = s.astype("string").fillna("(blank)").str.strip()
    return s.replace("", "(blank)")

def parse_number_series(series):
    if series is None:
        return pd.Series(dtype="float64")
    return (series.astype("string")
            .str.replace(",", "", regex=False)
            .str.extract(r"([-+]?\d*\.?\d+)", expand=False)
            .astype(float))

def parse_time_one(x):
    """Return seconds since midnight. Robust for Excel values and messy text like '6am, 6 AM, 0600, 6:00'."""
    if pd.isna(x):
        return np.nan
    if isinstance(x, pd.Timestamp):
        return float(x.hour*3600 + x.minute*60 + x.second)
    if isinstance(x, (int, float, np.integer, np.floating)):
        xx = float(x)
        if math.isfinite(xx):
            if 0 <= xx < 1.5:      # Excel fraction of day
                return xx * 86400.0
            if 0 <= xx <= 2400 and float(xx).is_integer():
                # Treat 600/0600/1730 as HHMM, but keep small values 0-23 as hours.
                if xx <= 23:
                    return xx * 3600.0
                hh = int(xx) // 100
                mm = int(xx) % 100
                if 0 <= hh <= 23 and 0 <= mm <= 59:
                    return float(hh*3600 + mm*60)
            if 0 <= xx <= 86400:
                return xx
        return np.nan

    y = str(x).strip().lower()
    if not y or y in {"nan", "none", "null"}:
        return np.nan

    # Remove common spreadsheet artifacts.
    y = y.replace("'", "").replace('"', "")
    y = y.replace("a.m.", "am").replace("p.m.", "pm").replace("a.m", "am").replace("p.m", "pm")
    y = y.replace(" am", "am").replace(" pm", "pm")
    y = y.strip()

    # HHMM strings such as 0600 or 1730
    if y.isdigit() and len(y) in (3, 4):
        val = int(y)
        hh, mm = val // 100, val % 100
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return float(hh*3600 + mm*60)

    from datetime import datetime as dt
    for fmt in ["%I%p", "%I:%M%p", "%I:%M:%S%p", "%H", "%H:%M", "%H:%M:%S"]:
        try:
            d = dt.strptime(y, fmt)
            return float(d.hour*3600 + d.minute*60 + d.second)
        except Exception:
            pass
    return np.nan

def fmt_hhmm(sec):
    if pd.isna(sec):
        return ""
    return f"{int((sec % 86400)//3600):02d}:{int(((sec % 86400)%3600)//60):02d}"

def ticks_value(x):
    try:
        v = float(x)
        return v if math.isfinite(v) and v > 0 else None
    except Exception:
        return None


def normalize_column_key(name):
    """Loose column-name normalizer for smart imports."""
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


SMART_COLUMN_ALIASES = {
    "start": "Start Time",
    "starttime": "Start Time",
    "begin": "Start Time",
    "begintime": "Start Time",
    "end": "End Time",
    "endtime": "End Time",
    "stoptime": "End Time",
    "equipment": "Equipment",
    "equip": "Equipment",
    "system": "Equipment",
    "asset": "Equipment",
    "centerfrequency": "Center Frequency (MHz)",
    "centerfrequencymhz": "Center Frequency (MHz)",
    "centerfreq": "Center Frequency (MHz)",
    "centerfreqmhz": "Center Frequency (MHz)",
    "frequency": "Center Frequency (MHz)",
    "freq": "Center Frequency (MHz)",
    "startfrequency": "Start Frequency (MHz)",
    "startfrequencymhz": "Start Frequency (MHz)",
    "startfreq": "Start Frequency (MHz)",
    "startfreqmhz": "Start Frequency (MHz)",
    "endfrequency": "End Frequency (MHz)",
    "endfrequencymhz": "End Frequency (MHz)",
    "endfreq": "End Frequency (MHz)",
    "endfreqmhz": "End Frequency (MHz)",
    "bandwidth": "Bandwidth (MHz)",
    "bandwidthmhz": "Bandwidth (MHz)",
    "bw": "Bandwidth (MHz)",
    "power": "Power (W)",
    "powerw": "Power (W)",
    "watts": "Power (W)",
    "powerdbm": "Power (dBm)",
    "dbm": "Power (dBm)",
    "lat": "Latitude",
    "latitude": "Latitude",
    "y": "Latitude",
    "lon": "Longitude",
    "long": "Longitude",
    "lng": "Longitude",
    "longitude": "Longitude",
    "x": "Longitude",
    "location": "Location",
    "site": "Site Name",
    "sitename": "Site Name",
    "antennaheight": "Antenna Height",
    "antennaheightft": "Antenna Height",
    "antennaheightm": "Antenna Height",
    "coverageradius": "Coverage Radius",
    "coverageradiusnm": "Coverage Radius",
    "coverageradiusmi": "Coverage Radius",
    "coverageradiuskm": "Coverage Radius",
    "mgrs": "MGRS",
    "usng": "USNG",
    "grid": "MGRS",
    "militarygrid": "MGRS",
    "active": "Active",
    "enabled": "Active",
    "inuse": "Active",
    "use": "Active",
    "include": "Active",
}


def smart_standardize_columns(df):
    """Apply smart column cleanup beyond exact STANDARD_RENAME matching."""
    out = df.copy()
    uploaded_column_order = list(out.columns)
    new_cols = []
    for c in out.columns:
        raw = str(c).strip()
        exact = STANDARD_RENAME.get(raw)
        if exact:
            new_cols.append(exact)
            continue
        key = normalize_column_key(raw)
        new_cols.append(SMART_COLUMN_ALIASES.get(key, raw))
    out.columns = new_cols
    return out


def mgrs_to_latlon(value):
    """Convert MGRS/USNG to lat/lon if optional mgrs package is available."""
    if mgrs is None:
        return (np.nan, np.nan)
    try:
        grid = str(value).strip()
        if not grid or grid.lower() in ["nan", "none", "null"]:
            return (np.nan, np.nan)
        converter = mgrs.MGRS()
        lat, lon = converter.toLatLon(grid.replace(" ", ""))
        return (float(lat), float(lon))
    except Exception:
        return (np.nan, np.nan)


def fill_latlon_from_mgrs(df):
    """Fill blank Latitude/Longitude from MGRS or USNG when possible."""
    out = df.copy()
    if "Latitude" not in out.columns:
        out["Latitude"] = np.nan
    if "Longitude" not in out.columns:
        out["Longitude"] = np.nan

    grid_col = None
    if "MGRS" in out.columns and out["MGRS"].astype(str).str.strip().replace("nan", "").ne("").any():
        grid_col = "MGRS"
    elif "USNG" in out.columns and out["USNG"].astype(str).str.strip().replace("nan", "").ne("").any():
        grid_col = "USNG"

    if grid_col and mgrs is not None:
        lat_vals = pd.to_numeric(out["Latitude"], errors="coerce")
        lon_vals = pd.to_numeric(out["Longitude"], errors="coerce")
        missing = lat_vals.isna() | lon_vals.isna()
        for idx in out.index[missing]:
            lat, lon = mgrs_to_latlon(out.at[idx, grid_col])
            if np.isfinite(lat) and np.isfinite(lon):
                out.at[idx, "Latitude"] = lat
                out.at[idx, "Longitude"] = lon

    return out




def to_active_bool(value):
    """Convert common spreadsheet active/inactive values to bool."""
    if pd.isna(value):
        return True
    s = str(value).strip().lower()
    if s in ["false", "no", "n", "0", "off", "inactive", "disabled", "unused", "not used", "exclude"]:
        return False
    return True


def active_label(value):
    return "Active" if to_active_bool(value) else "Inactive"


def apply_active_filter(df, show_inactive=False):
    """Return only active rows unless show_inactive is enabled."""
    if df is None or df.empty:
        return df
    if "Active" not in df.columns:
        return df.copy()
    if show_inactive:
        return df.copy()
    return df[df["Active"].apply(to_active_bool)].copy()


def ensure_active_first_preserve_order(df, default=True):
    """
    Keep the uploaded spreadsheet column order exactly as provided,
    except Active is always moved/inserted as Column A.
    """
    if df is None:
        return df
    out = df.copy()
    original_cols = list(out.columns)

    active_cols = [c for c in original_cols if str(c).strip().lower() == "active"]
    if active_cols:
        active_col = active_cols[0]
        if active_col != "Active":
            out = out.rename(columns={active_col: "Active"})
        cols = ["Active"] + [c for c in out.columns if c != "Active"]
        return out[cols]

    out.insert(0, "Active", default)
    return out


def preserve_spreadsheet_order_with_active_first(df, reference_columns=None):
    """
    Reorder a dataframe so Active is first, then columns follow the original
    spreadsheet/reference order. Any new app/planning columns are appended.
    """
    if df is None:
        return df
    out = ensure_active_first_preserve_order(df)
    if reference_columns:
        ref = []
        seen = set()
        for c in reference_columns:
            cc = "Active" if str(c).strip().lower() == "active" else c
            if cc in out.columns and cc not in seen:
                ref.append(cc)
                seen.add(cc)
        ordered = ["Active"] + [c for c in ref if c != "Active"] + [c for c in out.columns if c not in seen and c != "Active"]
        return out[ordered]
    return out




def normalize_uploaded_df(df):
    """Normalize uploaded/pasted data and infer columns when a file has weak or shifted headers."""
    out = smart_standardize_columns(df)
    out = fill_latlon_from_mgrs(out)

    # Drop fully empty rows/columns.
    out = out.dropna(how="all").dropna(axis=1, how="all")

    # If required columns are missing or mostly empty, infer them from the data.
    cols = list(out.columns)

    def mostly_empty(colname):
        return colname not in out.columns or out[colname].isna().all() or (out[colname].astype("string").str.strip().replace("", pd.NA).isna().mean() > 0.85)

    # Time columns: first two columns that parse like times.
    time_scores = []
    for c in cols:
        parsed = out[c].apply(parse_time_one)
        score = parsed.notna().mean() if len(parsed) else 0
        if score >= 0.50:
            time_scores.append((c, score))
    if mostly_empty("Start Time") and len(time_scores) >= 1:
        out["Start Time"] = out[time_scores[0][0]]
    if mostly_empty("End Time") and len(time_scores) >= 2:
        out["End Time"] = out[time_scores[1][0]]

    # Numeric frequency/power candidates.
    numeric_candidates = []
    for c in cols:
        nums = parse_number_series(out[c])
        score = nums.notna().mean() if len(nums) else 0
        med = nums.median(skipna=True)
        numeric_candidates.append((c, score, med))

    # Center frequency: frequency-like column, usually values > 20 MHz and not power.
    freq_like = [(c, s, m) for c, s, m in numeric_candidates if s >= 0.50 and pd.notna(m) and m > 20]
    if mostly_empty("Center Frequency (MHz)") and freq_like:
        # Prefer column names containing center/frequency/freq, otherwise the first freq-like column.
        named = [x for x in freq_like if any(k in str(x[0]).lower() for k in ["center", "frequency", "freq"])]
        out["Center Frequency (MHz)"] = out[(named or freq_like)[0][0]]

    # Equipment: text column that is not a time column and has repeated non-numeric values.
    if mostly_empty("Equipment"):
        time_cols = {x[0] for x in time_scores}
        text_scores = []
        for c in cols:
            if c in time_cols:
                continue
            as_text = out[c].astype("string").str.strip()
            nonblank = as_text.replace("", pd.NA).notna().mean()
            numeric_score = parse_number_series(out[c]).notna().mean()
            if nonblank >= 0.50 and numeric_score < 0.50:
                text_scores.append((c, nonblank))
        if text_scores:
            out["Equipment"] = out[text_scores[0][0]]

    # Bandwidth and Power defaults if absent.
    if mostly_empty("Bandwidth (MHz)"):
        out["Bandwidth (MHz)"] = None
    if mostly_empty("Power (W)"):
        out["Power (W)"] = 1

    for c in APP_COLUMNS:
        if c not in out.columns:
            out[c] = None

    if "Active" in out.columns:
        out["Active"] = out["Active"].apply(lambda v: True if pd.isna(v) else to_active_bool(v))
    out = ensure_active_first_preserve_order(out)

    # Preserve the user's spreadsheet order, with Active moved to Column A.
    preferred = ["Active"] + [c for c in uploaded_column_order if str(c).strip().lower() != "active" and c in out.columns]
    # Keep known app columns and any new/imported columns after the user's columns.
    preferred += [c for c in out.columns if c not in preferred]
    return out[preferred].reset_index(drop=True)

def app_to_internal(df):
    out = df.copy()
    out.columns = [INTERNAL_RENAME.get(str(c), str(c)) for c in out.columns]
    return out

def derive_power_w(df):
    out = df.copy()
    if "PowerW" not in out.columns and "PowerdBm" in out.columns:
        pdbm = parse_number_series(out["PowerdBm"])
        out["PowerW"] = (10 ** (pdbm / 10.0)) / 1000.0
    elif "PowerW" in out.columns:
        out["PowerW"] = parse_number_series(out["PowerW"])
    return out

def fill_from_center_bw(df):
    out = df.copy()
    if "CenterF" not in out.columns or "BW" not in out.columns:
        return out
    cf = parse_number_series(out["CenterF"])
    bw = parse_number_series(out["BW"])
    half = bw / 2.0
    if "StartF" not in out.columns:
        out["StartF"] = np.nan
    if "EndF" not in out.columns:
        out["EndF"] = np.nan
    sf = parse_number_series(out["StartF"])
    ef = parse_number_series(out["EndF"])
    out.loc[sf.isna(), "StartF"] = cf[sf.isna()] - half[sf.isna()]
    out.loc[ef.isna(), "EndF"] = cf[ef.isna()] + half[ef.isna()]
    return out

def prep_df(app_df):
    out = app_to_internal(app_df)
    out = derive_power_w(out)
    out = fill_from_center_bw(out)
    for c in ["CenterF", "StartF", "EndF", "BW", "PowerW", "PowerdBm", "Latitude", "Longitude", "AntennaHeight", "CoverageRadius"]:
        if c in out.columns:
            out[c] = parse_number_series(out[c])
    for c in ["Equipment", "Tech", "Unit"]:
        if c not in out.columns:
            out[c] = "(blank)"
        out[c] = safe_group(out[c])
    for c in ["StartTime", "EndTime"]:
        if c not in out.columns:
            out[c] = ""
    for c in ["StartF", "EndF", "PowerW"]:
        if c not in out.columns:
            out[c] = np.nan
    out = out.loc[out["StartF"].notna() & out["EndF"].notna() & out["PowerW"].notna() & (out["EndF"] > out["StartF"])].copy()
    if "CenterF" not in out.columns:
        out["CenterF"] = (out["StartF"] + out["EndF"]) / 2
    else:
        out["CenterF"] = out["CenterF"].fillna((out["StartF"] + out["EndF"]) / 2)
    out["LabelY"] = out["PowerW"] / 2
    out["PointY"] = out["PowerW"]
    out[".row_id"] = np.arange(1, len(out) + 1)
    return out.reset_index(drop=True)

def make_palette(levels):
    cmap = plt.get_cmap("tab20")
    return {lev: cmap(i % 20) for i, lev in enumerate(levels)}

def unittech_field(df):
    if "Unit" in df.columns and (safe_group(df["Unit"]) != "(blank)").any():
        return "Unit"
    if "Tech" in df.columns and (safe_group(df["Tech"]) != "(blank)").any():
        return "Tech"
    return "Equipment"

def fig_to_png_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    buf.seek(0)
    return buf.getvalue()

def deck_to_html_bytes(deck):
    """Create a downloadable standalone HTML map."""
    try:
        html = deck.to_html(as_string=True, notebook_display=False)
    except TypeError:
        html = deck.to_html(as_string=True)
    return html.encode("utf-8")


def radius_to_meters(value, units="NM"):
    """Convert a coverage radius value to meters."""
    try:
        cleaned = str(value).replace("NM", "").replace("nm", "").replace("mi", "").replace("km", "").strip()
        radius = float(pd.to_numeric(cleaned, errors="coerce"))
    except Exception:
        radius = 0.0
    if not np.isfinite(radius) or radius <= 0:
        return 0.0
    units = str(units or "NM").lower()
    if units == "km":
        return radius * 1000.0
    if units in ["mi", "mile", "miles"]:
        return radius * 1609.344
    return radius * 1852.0


def circle_polygon_kml(lon, lat, radius_m, points=72):
    """Approximate a coverage circle as a KML polygon."""
    if radius_m <= 0:
        return ""
    coords = []
    earth_radius = 6378137.0
    lat_rad = math.radians(lat)
    lon_rad = math.radians(lon)
    angular_distance = radius_m / earth_radius

    for i in range(points + 1):
        bearing = math.radians((360.0 / points) * i)
        lat2 = math.asin(
            math.sin(lat_rad) * math.cos(angular_distance)
            + math.cos(lat_rad) * math.sin(angular_distance) * math.cos(bearing)
        )
        lon2 = lon_rad + math.atan2(
            math.sin(bearing) * math.sin(angular_distance) * math.cos(lat_rad),
            math.cos(angular_distance) - math.sin(lat_rad) * math.sin(lat2),
        )
        coords.append(f"{math.degrees(lon2):.8f},{math.degrees(lat2):.8f},0")
    return " ".join(coords)


def map_df_to_kml(map_df, project_name="Spectrum Planner", radius_units="NM", include_coverage=True):
    """Export map rows as KML points and optional coverage polygons."""
    if map_df is None or map_df.empty:
        return b""

    rows = map_df.copy()
    rows["Latitude"] = pd.to_numeric(rows.get("Latitude"), errors="coerce")
    rows["Longitude"] = pd.to_numeric(rows.get("Longitude"), errors="coerce")
    rows = rows.dropna(subset=["Latitude", "Longitude"])

    placemarks = []
    for _, r in rows.iterrows():
        equipment = escape(str(r.get("Equipment", "Site")))
        site = escape(str(r.get("SiteName", "") or r.get("Location", "") or equipment))
        location = escape(str(r.get("Location", "")))
        unit = escape(str(r.get("Unit", "")))
        tech = escape(str(r.get("Tech", "")))
        center = escape(str(r.get("CenterF", "")))
        power = escape(str(r.get("PowerW", "")))
        start = escape(str(r.get("StartTime", "")))
        end = escape(str(r.get("EndTime", "")))
        lat = float(r["Latitude"])
        lon = float(r["Longitude"])
        radius_raw = r.get("CoverageRadius", 0)
        radius_m = radius_to_meters(radius_raw, radius_units)

        desc = (
            f"Equipment: {equipment}<br/>"
            f"Site: {site}<br/>"
            f"Location: {location}<br/>"
            f"Unit: {unit}<br/>"
            f"Tech: {tech}<br/>"
            f"Center Frequency: {center} MHz<br/>"
            f"Power: {power} W<br/>"
            f"Time: {start} - {end}<br/>"
            f"Coverage Radius: {escape(str(radius_raw))} {escape(str(radius_units))}"
        )

        placemarks.append(f"""
        <Placemark>
          <name>{site}</name>
          <description><![CDATA[{desc}]]></description>
          <Point><coordinates>{lon:.8f},{lat:.8f},0</coordinates></Point>
        </Placemark>
        """)

        if include_coverage and radius_m > 0:
            polygon_coords = circle_polygon_kml(lon, lat, radius_m)
            placemarks.append(f"""
            <Placemark>
              <name>{site} coverage</name>
              <description><![CDATA[{desc}]]></description>
              <Style>
                <LineStyle><color>ff000000</color><width>1</width></LineStyle>
                <PolyStyle><color>330000ff</color></PolyStyle>
              </Style>
              <Polygon>
                <outerBoundaryIs>
                  <LinearRing><coordinates>{polygon_coords}</coordinates></LinearRing>
                </outerBoundaryIs>
              </Polygon>
            </Placemark>
            """)

    kml = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>{escape(str(project_name))}</name>
    {''.join(placemarks)}
  </Document>
</kml>
"""
    return kml.encode("utf-8")



def map_congestion_summary(map_df):
    """Create a simple congestion summary from map rows."""
    if map_df is None or map_df.empty:
        return pd.DataFrame({"Message": ["No map rows."]})

    d = map_df.copy()
    d["PowerW"] = pd.to_numeric(d.get("PowerW", 0), errors="coerce").fillna(0)
    d["CoverageRadius"] = pd.to_numeric(d.get("CoverageRadius", 0), errors="coerce").fillna(0)
    d["CenterF"] = pd.to_numeric(d.get("CenterF", np.nan), errors="coerce")

    group_cols = [c for c in ["Location", "SiteName", "Unit"] if c in d.columns]
    if not group_cols:
        return pd.DataFrame({
            "Sites": [len(d)],
            "Total Power (W)": [round(d["PowerW"].sum(), 2)],
            "Rows with Coordinates": [len(d)],
        })

    group_col = group_cols[0]
    out = (
        d.groupby(group_col, dropna=False)
        .agg(
            Rows=("Equipment", "count"),
            TotalPowerW=("PowerW", "sum"),
            AvgCoverage=("CoverageRadius", "mean"),
            UniqueFreqs=("CenterF", lambda x: x.dropna().nunique()),
        )
        .reset_index()
        .sort_values(["Rows", "TotalPowerW"], ascending=[False, False])
    )
    out["TotalPowerW"] = out["TotalPowerW"].round(2)
    out["AvgCoverage"] = out["AvgCoverage"].round(2)
    return out



def map_df_to_geojson(map_df, project_name="Spectrum Planner"):
    """Export map rows as GeoJSON FeatureCollection points."""
    if map_df is None or map_df.empty:
        return b'{"type":"FeatureCollection","features":[]}'

    rows = map_df.copy()
    rows["Latitude"] = pd.to_numeric(rows.get("Latitude"), errors="coerce")
    rows["Longitude"] = pd.to_numeric(rows.get("Longitude"), errors="coerce")
    rows = rows.dropna(subset=["Latitude", "Longitude"])

    features = []
    for _, r in rows.iterrows():
        props = {}
        for col in ["Equipment", "Tech", "Unit", "Location", "SiteName", "CoverageRadius", "AntennaHeight", "CenterF", "StartF", "EndF", "PowerW", "StartTime", "EndTime"]:
            if col in rows.columns:
                props[col] = json_safe_value(r.get(col))
        props["project"] = project_name

        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [float(r["Longitude"]), float(r["Latitude"])],
            },
            "properties": props,
        })

    geojson = {
        "type": "FeatureCollection",
        "name": project_name,
        "features": features,
    }
    import json
    return json.dumps(geojson, indent=2).encode("utf-8")



def haversine_distance_km(lat1, lon1, lat2, lon2):
    """Great-circle distance in kilometers."""
    r = 6371.0088
    p1, p2 = math.radians(float(lat1)), math.radians(float(lat2))
    dp = math.radians(float(lat2) - float(lat1))
    dl = math.radians(float(lon2) - float(lon1))
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def bearing_degrees(lat1, lon1, lat2, lon2):
    """Initial bearing from point 1 to point 2."""
    lat1, lon1, lat2, lon2 = map(math.radians, [float(lat1), float(lon1), float(lat2), float(lon2)])
    dlon = lon2 - lon1
    y = math.sin(dlon) * math.cos(lat2)
    x = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
    return (math.degrees(math.atan2(y, x)) + 360) % 360


def parse_height_m(value, default_m=10.0):
    """Parse antenna height. If a spreadsheet value is plain numeric, treat it as feet by default for field use."""
    try:
        raw = str(value).strip().lower()
        if not raw or raw in ["nan", "none", "null"]:
            return default_m
        num = float(pd.to_numeric(re.sub(r"[^0-9.+-]", "", raw), errors="coerce"))
        if not np.isfinite(num):
            return default_m
        if " m" in raw or raw.endswith("m"):
            return num
        # Default to feet because most US field worksheets use feet.
        return num * 0.3048
    except Exception:
        return default_m


def radio_horizon_km(height_m):
    """Approximate RF horizon in km using 4/3 earth model."""
    h = max(float(height_m), 0.0)
    return 4.12 * math.sqrt(h)


def build_range_analysis(map_df):
    """Build pairwise site-to-site distance and approximate LOS/range table."""
    if map_df is None or map_df.empty:
        return pd.DataFrame({"Message": ["No map rows."]})

    d = map_df.copy()
    d["Latitude"] = pd.to_numeric(d.get("Latitude"), errors="coerce")
    d["Longitude"] = pd.to_numeric(d.get("Longitude"), errors="coerce")
    d = d.dropna(subset=["Latitude", "Longitude"]).reset_index(drop=True)

    if len(d) < 2:
        return pd.DataFrame({"Message": ["Need at least two mapped sites for path analysis."]})

    if "AntennaHeight" not in d.columns:
        d["AntennaHeight"] = np.nan

    rows = []
    for i in range(len(d)):
        for j in range(i + 1, len(d)):
            a = d.iloc[i]
            b = d.iloc[j]

            name_a = str(a.get("SiteName", "") or a.get("Equipment", f"Site {i+1}"))
            name_b = str(b.get("SiteName", "") or b.get("Equipment", f"Site {j+1}"))

            dist_km = haversine_distance_km(a["Latitude"], a["Longitude"], b["Latitude"], b["Longitude"])
            dist_mi = dist_km * 0.621371
            dist_nm = dist_km * 0.539957

            h1 = parse_height_m(a.get("AntennaHeight", np.nan))
            h2 = parse_height_m(b.get("AntennaHeight", np.nan))
            horizon_km = radio_horizon_km(h1) + radio_horizon_km(h2)

            margin_km = horizon_km - dist_km
            if margin_km >= 5:
                status = "Likely LOS"
            elif margin_km >= 0:
                status = "Marginal LOS"
            else:
                status = "Beyond horizon"

            rows.append({
                "Site A": name_a,
                "Site B": name_b,
                "Distance km": round(dist_km, 2),
                "Distance mi": round(dist_mi, 2),
                "Distance NM": round(dist_nm, 2),
                "Bearing A→B": round(bearing_degrees(a["Latitude"], a["Longitude"], b["Latitude"], b["Longitude"]), 1),
                "Antenna A m": round(h1, 1),
                "Antenna B m": round(h2, 1),
                "RF Horizon km": round(horizon_km, 2),
                "Horizon Margin km": round(margin_km, 2),
                "LOS Screen": status,
            })

    return pd.DataFrame(rows).sort_values(["LOS Screen", "Distance km"], ascending=[True, True]).reset_index(drop=True)


def path_lines_geojson(range_df, map_df, project_name="Spectrum Planner"):
    """Create GeoJSON LineString features for site-to-site paths."""
    if range_df is None or range_df.empty or "Message" in range_df.columns:
        return b'{"type":"FeatureCollection","features":[]}'

    m = map_df.copy()
    m["Latitude"] = pd.to_numeric(m.get("Latitude"), errors="coerce")
    m["Longitude"] = pd.to_numeric(m.get("Longitude"), errors="coerce")
    m["NameKey"] = m.apply(lambda r: str(r.get("SiteName", "") or r.get("Equipment", "")), axis=1)
    lookup = {r["NameKey"]: r for _, r in m.dropna(subset=["Latitude", "Longitude"]).iterrows()}

    features = []
    for _, r in range_df.iterrows():
        a = lookup.get(str(r.get("Site A")))
        b = lookup.get(str(r.get("Site B")))
        if a is None or b is None:
            continue
        props = {k: json_safe_value(v) for k, v in r.items()}
        props["project"] = project_name
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "LineString",
                "coordinates": [
                    [float(a["Longitude"]), float(a["Latitude"])],
                    [float(b["Longitude"]), float(b["Latitude"])],
                ],
            },
            "properties": props,
        })

    import json
    return json.dumps({"type": "FeatureCollection", "name": f"{project_name} paths", "features": features}, indent=2).encode("utf-8")





def backup_json_safe(obj):
    """Recursively convert backup data to JSON-safe values."""
    if obj is None:
        return None
    if isinstance(obj, dict):
        return {str(k): backup_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [backup_json_safe(v) for v in obj]
    if isinstance(obj, tuple):
        return [backup_json_safe(v) for v in obj]
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    try:
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            if not np.isfinite(obj):
                return None
            return float(obj)
    except Exception:
        pass
    return str(obj) if not isinstance(obj, (str, int, float, bool)) else obj


def export_project_backup(project_id):
    """Export current project data, rows, versions, sheets, files metadata, members, and audit events."""
    import json
    backup = {
        "backup_version": 1,
        "exported_at": now_iso(),
        "project": {},
        "allocation_rows": [],
        "allocation_versions": [],
        "project_sheets": [],
        "project_files": [],
        "project_members": [],
        "save_events": [],
        "project_audit_events": [],
    }

    tables = [
        ("projects", "project"),
        ("allocation_rows", "allocation_rows"),
        ("allocation_versions", "allocation_versions"),
        ("project_sheets", "project_sheets"),
        ("project_files", "project_files"),
        ("project_members", "project_members"),
        ("save_events", "save_events"),
        ("project_audit_events", "project_audit_events"),
    ]

    for table, key in tables:
        try:
            if table == "projects":
                rows = sb.table(table).select("*").eq("id", project_id).limit(1).execute().data or []
                backup[key] = rows[0] if rows else {}
            else:
                rows = sb.table(table).select("*").eq("project_id", project_id).execute().data or []
                backup[key] = rows
        except Exception:
            backup[key] = {} if key == "project" else []

    return json.dumps(backup_json_safe(backup), indent=2).encode("utf-8")


def restore_project_backup(backup_bytes, new_project_name, user):
    """Restore a backup into a new project. Original project id is not reused."""
    import json
    if isinstance(backup_bytes, bytes):
        raw = backup_bytes.decode("utf-8")
    else:
        raw = str(backup_bytes)

    backup = json.loads(raw)
    project = backup.get("project") or {}
    name = new_project_name.strip() or f"Restored - {project.get('name', 'Project')}"
    description = project.get("description") or "Restored from JSON backup"

    new_project = create_project(name, description)
    new_project_id = new_project["id"]

    # Restore workbook sheets.
    sheets_payload = []
    for i, row in enumerate(backup.get("project_sheets", []) or []):
        sheets_payload.append({
            "project_id": new_project_id,
            "sheet_name": row.get("sheet_name", f"Sheet {i+1}"),
            "sheet_order": row.get("sheet_order", i + 1),
            "sheet_data": row.get("sheet_data", []),
            "uploaded_by": user,
            "updated_at": now_iso(),
        })
    if sheets_payload:
        sb.table("project_sheets").insert(sheets_payload).execute()

    # Restore current allocation rows.
    row_payload = []
    for i, row in enumerate(backup.get("allocation_rows", []) or []):
        row_payload.append({
            "project_id": new_project_id,
            "row_order": row.get("row_order", i),
            "row_data": row.get("row_data", {}),
            "updated_by": user,
            "updated_at": now_iso(),
        })
    if row_payload:
        sb.table("allocation_rows").insert(row_payload).execute()

    # Restore versions.
    version_payload = []
    for row in backup.get("allocation_versions", []) or []:
        version_payload.append({
            "project_id": new_project_id,
            "version_no": row.get("version_no"),
            "saved_by": row.get("saved_by", user),
            "save_note": f"Restored backup: {row.get('save_note', '')}",
            "snapshot": row.get("snapshot", []),
            "created_at": now_iso(),
        })
    if version_payload:
        sb.table("allocation_versions").insert(version_payload).execute()

    # Restore members as metadata, but always add restoring user as owner.
    try:
        upsert_project_member(new_project_id, current_user_id, user, "owner", user)
    except Exception:
        pass

    # Restore file metadata only. Original Storage objects are not duplicated here.
    file_payload = []
    for row in backup.get("project_files", []) or []:
        file_payload.append({
            "project_id": new_project_id,
            "file_name": row.get("file_name", "restored_file"),
            "storage_path": row.get("storage_path", ""),
            "uploaded_by": row.get("uploaded_by", user),
            "uploaded_at": now_iso(),
        })
    if file_payload:
        try:
            sb.table("project_files").insert(file_payload).execute()
        except Exception:
            pass

    log_audit_event(new_project_id, "project_restored_from_backup", user, {"source_project": project.get("name", "")})
    return new_project



# ---------------- Supabase JSON-backed operations ----------------

def safe_storage_filename(name):
    """Make a file name safe for Supabase Storage paths."""
    base = Path(str(name or "uploaded_file")).name
    base = re.sub(r"[^A-Za-z0-9._ -]+", "_", base).strip()
    return base or "uploaded_file"


def guess_content_type(file_name):
    content_type, _ = mimetypes.guess_type(file_name)
    return content_type or "application/octet-stream"


def upload_original_file(project_id, file_name, file_bytes, user):
    """Upload original CSV/XLSX file to Supabase Storage and track it in project_files."""
    if not project_id or not file_name or not file_bytes:
        return None

    safe_name = safe_storage_filename(file_name)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    storage_path = f"{project_id}/{stamp}_{uuid.uuid4().hex[:8]}_{safe_name}"
    content_type = guess_content_type(safe_name)

    # Upload the original file to private bucket.
    sb.storage.from_(STORAGE_BUCKET).upload(
        storage_path,
        file_bytes,
        file_options={
            "content-type": content_type,
            "upsert": "false",
        },
    )

    rec = {
        "project_id": project_id,
        "file_name": safe_name,
        "storage_path": storage_path,
        "uploaded_by": user,
        "uploaded_at": now_iso(),
    }
    data = sb.table("project_files").insert(rec).execute().data
    return data[0] if data else rec


def list_project_files(project_id):
    """Return uploaded file records for this project."""
    try:
        return (
            sb.table("project_files")
            .select("*")
            .eq("project_id", project_id)
            .order("uploaded_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def download_original_file(storage_path):
    """Download an original file from Supabase Storage."""
    return sb.storage.from_(STORAGE_BUCKET).download(storage_path)


def save_project_sheets(project_id, sheets_dict, user):
    """Save every workbook sheet for the selected project."""
    if not sheets_dict:
        return 0

    sb.table("project_sheets").delete().eq("project_id", project_id).execute()

    payloads = []
    order_no = 0
    for sheet_name, df_sheet in sheets_dict.items():
        order_no += 1
        clean = normalize_uploaded_df(df_sheet)
        payloads.append({
            "project_id": project_id,
            "sheet_name": str(sheet_name),
            "sheet_order": order_no,
            "sheet_data": json_safe_records(clean),
            "uploaded_by": user,
            "updated_at": now_iso(),
        })

    if payloads:
        sb.table("project_sheets").insert(payloads).execute()
    log_audit_event(project_id, "workbook_sheets_saved", user, {"sheets": len(payloads)})

    return len(payloads)


def load_project_sheets(project_id):
    """Load saved workbook sheets for the selected project."""
    try:
        rows = (
            sb.table("project_sheets")
            .select("*")
            .eq("project_id", project_id)
            .order("sheet_order")
            .execute()
            .data
            or []
        )
    except Exception as e:
        st.warning(f"Workbook tabs could not be loaded from Supabase. Run the v17 SQL setup if you have not already. Details: {e}")
        return {}

    sheets = {}
    for row in rows:
        name = row.get("sheet_name", "Sheet")
        data = row.get("sheet_data", []) or []
        try:
            sheets[name] = normalize_uploaded_df(pd.DataFrame(data))
        except Exception:
            sheets[name] = pd.DataFrame(data)

    return sheets


def read_uploaded_workbook(uploaded, selected_sheet=None):
    """Read CSV/XLS/XLSX and return sheets dict, active sheet, file bytes."""
    file_bytes = uploaded.getvalue()
    ext = Path(uploaded.name).suffix.lower()

    if ext == ".csv":
        df = pd.read_csv(io.BytesIO(file_bytes))
        return {"CSV": normalize_uploaded_df(df)}, "CSV", file_bytes

    excel = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets = {}
    for sheet_name in excel.sheet_names:
        sheets[sheet_name] = normalize_uploaded_df(pd.read_excel(excel, sheet_name=sheet_name))

    active = selected_sheet if selected_sheet in sheets else excel.sheet_names[0]
    return sheets, active, file_bytes


def workbook_to_xlsx_bytes(sheets_dict):
    """Create a downloadable XLSX from saved workbook sheets."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for sheet_name, df_sheet in sheets_dict.items():
            safe_name = str(sheet_name)[:31] or "Sheet"
            df_sheet.to_excel(writer, sheet_name=safe_name, index=False)
    bio.seek(0)
    return bio.getvalue()


def log_audit_event(project_id, event_type, user, details=None):
    """Best-effort audit logging. Never blocks the app if logging fails."""
    try:
        sb.table("project_audit_events").insert({
            "project_id": project_id,
            "event_type": str(event_type),
            "event_by": str(user or ""),
            "details": json_safe_value(details) if not isinstance(details, dict) else {str(k): json_safe_value(v) for k, v in details.items()},
            "created_at": now_iso(),
        }).execute()
    except Exception:
        pass


def list_audit_events(project_id, limit=200):
    try:
        return (
            sb.table("project_audit_events")
            .select("*")
            .eq("project_id", project_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def get_project_status(project_id):
    try:
        rows = sb.table("projects").select("status,approved_by,approved_at,status_note").eq("id", project_id).limit(1).execute().data or []
        if rows:
            return rows[0]
    except Exception:
        pass
    return {"status": "Draft", "approved_by": None, "approved_at": None, "status_note": None}


def set_project_status(project_id, status, user, note=""):
    payload = {
        "status": status,
        "status_note": note,
        "updated_at": now_iso(),
    }
    if status == "Approved":
        payload["approved_by"] = user
        payload["approved_at"] = now_iso()
    elif status in ["Draft", "In Review", "Rejected"]:
        payload["approved_by"] = None
        payload["approved_at"] = None

    sb.table("projects").update(payload).eq("id", project_id).execute()
    log_audit_event(project_id, "status_change", user, {"status": status, "note": note})


def briefing_pdf_bytes(project_name, status_info, df_ready, conflicts_eq, conflicts_ut, fig_list):
    """Create a simple PDF briefing with summary, conflicts, and current figures."""
    bio = io.BytesIO()
    with PdfPages(bio) as pdf:
        # Cover page
        fig = plt.figure(figsize=(11, 8.5))
        fig.text(0.08, 0.88, "Spectrum Planner Briefing", fontsize=24, weight="bold")
        fig.text(0.08, 0.80, f"Project: {project_name}", fontsize=14)
        fig.text(0.08, 0.75, f"Status: {status_info.get('status', 'Draft')}", fontsize=14)
        fig.text(0.08, 0.70, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", fontsize=12)
        fig.text(0.08, 0.62, f"Rows: {len(df_ready)}", fontsize=12)
        fig.text(0.08, 0.58, f"Equipment conflicts: {0 if conflicts_eq is None else len(conflicts_eq)}", fontsize=12)
        fig.text(0.08, 0.54, f"Unit/Tech conflicts: {0 if conflicts_ut is None else len(conflicts_ut)}", fontsize=12)

        if status_info.get("approved_by"):
            fig.text(0.08, 0.48, f"Approved by: {status_info.get('approved_by')}", fontsize=12)
            fig.text(0.08, 0.44, f"Approved at: {status_info.get('approved_at')}", fontsize=12)
        if status_info.get("status_note"):
            fig.text(0.08, 0.38, f"Status note: {status_info.get('status_note')}", fontsize=11)

        fig.text(0.08, 0.12, "Generated from the current active plotting sheet.", fontsize=10)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        for fig in fig_list:
            pdf.savefig(fig, bbox_inches="tight")

    bio.seek(0)
    return bio.getvalue()


def update_presence(project_id, user_id, email, role):
    """Mark current user as online for this project."""
    try:
        sb.table("project_presence").upsert({
            "project_id": project_id,
            "user_id": user_id,
            "email": email,
            "role": role,
            "last_seen": now_iso(),
        }, on_conflict="project_id,user_id").execute()
    except Exception:
        pass


def list_presence(project_id):
    """List users seen recently on this project."""
    try:
        rows = (
            sb.table("project_presence")
            .select("*")
            .eq("project_id", project_id)
            .order("last_seen", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return pd.DataFrame()

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    try:
        last_seen = pd.to_datetime(df["last_seen"], utc=True)
        now = pd.Timestamp.utcnow()
        df["minutes_ago"] = ((now - last_seen).dt.total_seconds() / 60).round(1)
        df["online"] = df["minutes_ago"] <= 5
    except Exception:
        df["minutes_ago"] = None
        df["online"] = True

    return df


def list_recent_activity(project_id, limit=50):
    """Merge audit events and save events into a simple activity feed."""
    events = []

    try:
        audit_rows = (
            sb.table("project_audit_events")
            .select("*")
            .eq("project_id", project_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
            .data
            or []
        )
        for r in audit_rows:
            events.append({
                "time": r.get("created_at"),
                "type": r.get("event_type"),
                "user": r.get("event_by"),
                "note": str(r.get("details", "")),
            })
    except Exception:
        pass

    try:
        save_rows = (
            sb.table("save_events")
            .select("*")
            .eq("project_id", project_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
            .data
            or []
        )
        for r in save_rows:
            events.append({
                "time": r.get("created_at"),
                "type": r.get("event_type"),
                "user": r.get("event_by"),
                "note": r.get("event_note"),
            })
    except Exception:
        pass

    if not events:
        return pd.DataFrame(columns=["time", "type", "user", "note"])

    df = pd.DataFrame(events)
    df["time_dt"] = pd.to_datetime(df["time"], errors="coerce", utc=True)
    df = df.sort_values("time_dt", ascending=False).drop(columns=["time_dt"])
    return df.head(limit)


def get_last_project_update(project_id):
    """Return latest known project update timestamp and recent save/user info."""
    try:
        rows = sb.table("projects").select("updated_at,status,approved_by,approved_at").eq("id", project_id).limit(1).execute().data or []
        return rows[0] if rows else {}
    except Exception:
        return {}


def list_project_members(project_id):
    try:
        return (
            sb.table("project_members")
            .select("*")
            .eq("project_id", project_id)
            .order("email")
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def upsert_project_member(project_id, user_id, email, access_role="editor", added_by=""):
    rec = {
        "project_id": project_id,
        "user_id": user_id,
        "email": email or "",
        "access_role": access_role,
        "added_by": added_by,
        "updated_at": now_iso(),
    }
    sb.table("project_members").upsert(rec, on_conflict="project_id,user_id").execute()
    log_audit_event(project_id, "project_member_updated", added_by, {"email": email, "access_role": access_role})


def remove_project_member(project_id, user_id, removed_by=""):
    sb.table("project_members").delete().eq("project_id", project_id).eq("user_id", user_id).execute()
    log_audit_event(project_id, "project_member_removed", removed_by, {"user_id": user_id})


def user_has_project_access(project_id, user_id):
    if is_admin:
        return True
    try:
        rows = (
            sb.table("project_members")
            .select("user_id")
            .eq("project_id", project_id)
            .eq("user_id", user_id)
            .limit(1)
            .execute()
            .data
            or []
        )
        return bool(rows)
    except Exception:
        # If SQL has not been run yet, fail open so the app does not lock everyone out.
        return True


def row_signature(row_data):
    """Stable signature for a row for history comparison."""
    key_bits = [
        str(row_data.get("Equipment", "")),
        str(row_data.get("Unit", "")),
        str(row_data.get("Tech", "")),
        str(row_data.get("Start Time", "")),
        str(row_data.get("End Time", "")),
        str(row_data.get("Center Frequency (MHz)", "")),
        str(row_data.get("Start Frequency (MHz)", "")),
        str(row_data.get("End Frequency (MHz)", "")),
    ]
    return "|".join(key_bits)


def get_current_row_data(project_id):
    rows = (
        sb.table("allocation_rows")
        .select("row_order,row_data")
        .eq("project_id", project_id)
        .order("row_order")
        .execute()
        .data
        or []
    )
    return {int(r.get("row_order", 0)): (r.get("row_data") or {}) for r in rows}


def diff_row_data(before, after):
    changes = {}
    keys = sorted(set((before or {}).keys()) | set((after or {}).keys()))
    for k in keys:
        b = json_safe_value((before or {}).get(k))
        a = json_safe_value((after or {}).get(k))
        if str(b) != str(a):
            changes[k] = {"before": b, "after": a}
    return changes


def log_row_history(project_id, before_rows, after_records, user, source="save"):
    """Log per-row before/after changes."""
    payloads = []
    for row_order, after in enumerate(after_records):
        before = before_rows.get(row_order, {})
        changes = diff_row_data(before, after)
        if changes:
            payloads.append({
                "project_id": project_id,
                "row_order": row_order,
                "row_signature": row_signature(after),
                "changed_by": user,
                "change_source": source,
                "before_data": before,
                "after_data": after,
                "changes": changes,
                "created_at": now_iso(),
            })

    if payloads:
        try:
            sb.table("row_history").insert(payloads).execute()
        except Exception:
            pass


def list_row_history(project_id, limit=300):
    try:
        return (
            sb.table("row_history")
            .select("*")
            .eq("project_id", project_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def restore_row_from_history(project_id, history_row, user):
    """Restore one row_order from a row_history record."""
    row_order = int(history_row.get("row_order", 0))
    after_data = history_row.get("before_data") or {}
    current_rows = get_current_row_data(project_id)
    before = current_rows.get(row_order, {})
    current_rows[row_order] = after_data

    records = [current_rows[i] for i in sorted(current_rows.keys())]
    df = normalize_uploaded_df(pd.DataFrame(records))
    replace_project_rows(project_id, df, user)
    save_version(project_id, df, user, f"Restored row {row_order} from row history")
    log_audit_event(project_id, "row_restored_from_history", user, {"row_order": row_order})
    return df



def list_mission_templates():
    try:
        return (
            sb.table("mission_templates")
            .select("*")
            .order("updated_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        return []


def save_mission_template(name, description, sheets_dict, active_sheet, user):
    """Save current workbook sheets as a reusable mission template."""
    if not name or not sheets_dict:
        raise ValueError("Template name and workbook sheets are required.")

    clean_sheets = {}
    for sheet_name, df_sheet in sheets_dict.items():
        clean_sheets[str(sheet_name)] = json_safe_records(normalize_uploaded_df(df_sheet))

    payload = {
        "name": name.strip(),
        "description": description or "",
        "template_data": clean_sheets,
        "active_sheet": active_sheet,
        "created_by": user,
        "updated_by": user,
        "updated_at": now_iso(),
    }

    existing = (
        sb.table("mission_templates")
        .select("id")
        .eq("name", name.strip())
        .limit(1)
        .execute()
        .data
        or []
    )

    if existing:
        sb.table("mission_templates").update(payload).eq("id", existing[0]["id"]).execute()
        return existing[0]["id"]

    res = sb.table("mission_templates").insert(payload).execute()
    return res.data[0]["id"]


def delete_mission_template(template_id):
    return sb.table("mission_templates").delete().eq("id", template_id).execute()


def create_project_from_template(template, project_name, description, user):
    """Create a project and populate its sheets/active allocation table from a template."""
    proj = create_project(project_name, description)
    template_data = template.get("template_data") or {}
    active_sheet = template.get("active_sheet") or (list(template_data.keys())[0] if template_data else "Working")

    sheets_dict = {}
    for sheet_name, records in template_data.items():
        sheets_dict[str(sheet_name)] = normalize_uploaded_df(pd.DataFrame(records or []))

    if not sheets_dict:
        sheets_dict = {"Working": pd.DataFrame(columns=APP_COLUMNS)}
        active_sheet = "Working"

    if active_sheet not in sheets_dict:
        active_sheet = list(sheets_dict.keys())[0]

    save_project_sheets(proj["id"], sheets_dict, user)
    replace_project_rows(proj["id"], sheets_dict[active_sheet], user)
    save_version(proj["id"], sheets_dict[active_sheet], user, f"Created from template: {template.get('name', '')}")
    log_audit_event(proj["id"], "project_created_from_template", user, {"template": template.get("name", "")})

    return proj


def list_projects():
    all_projects = sb.table("projects").select("*").order("updated_at", desc=True).execute().data or []
    if is_admin:
        return all_projects

    accessible = []
    for project in all_projects:
        if user_has_project_access(project.get("id"), current_user_id):
            accessible.append(project)
    return accessible

def create_project(name, description):
    res = sb.table("projects").insert({
        "name": name,
        "description": description,
        "status": "Draft",
        "updated_at": now_iso(),
    }).execute()
    proj = res.data[0]
    log_audit_event(proj["id"], "project_created", logged_in_user if "logged_in_user" in globals() else "", {"name": name})
    try:
        upsert_project_member(
            proj["id"],
            current_user_id if "current_user_id" in globals() else "",
            logged_in_user if "logged_in_user" in globals() else "",
            "owner",
            logged_in_user if "logged_in_user" in globals() else "",
        )
    except Exception:
        pass
    return proj

def delete_project(project_id):
    """Delete one project and its related rows/files/sheets/versions/events."""
    log_audit_event(project_id, "project_deleted", logged_in_user if "logged_in_user" in globals() else "", {})
    # Delete Storage files first, then tracking rows.
    try:
        file_rows = list_project_files(project_id)
        paths = [r.get("storage_path") for r in file_rows if r.get("storage_path")]
        if paths:
            sb.storage.from_(STORAGE_BUCKET).remove(paths)
    except Exception:
        pass

    # Child tables. The project row also has cascade rules for some tables, but this is explicit.
    for table in ["allocation_rows", "allocation_versions", "save_events", "project_files", "project_sheets"]:
        try:
            sb.table(table).delete().eq("project_id", project_id).execute()
        except Exception:
            pass

    # Delete the project itself.
    return sb.table("projects").delete().eq("id", project_id).execute()


def get_project_rows(project_id):
    res = (sb.table("allocation_rows")
           .select("id,project_id,row_order,row_data,updated_at,updated_by")
           .eq("project_id", project_id)
           .order("row_order")
           .execute())
    rows = res.data or []
    if not rows:
        return pd.DataFrame(columns=APP_COLUMNS)
    data = [(r.get("row_data") or {}) for r in rows]
    df = pd.DataFrame(data)
    for c in APP_COLUMNS:
        if c not in df.columns:
            df[c] = None
    return df[APP_COLUMNS].reset_index(drop=True)

def replace_project_rows(project_id, df, user):
    before_rows = get_current_row_data(project_id)

    sb.table("allocation_rows").delete().eq("project_id", project_id).execute()

    clean = normalize_uploaded_df(df).reset_index(drop=True)
    safe_rows = json_safe_records(clean)
    log_row_history(project_id, before_rows, safe_rows, user, source="replace_project_rows")
    payloads = []
    for i, row_data in enumerate(safe_rows):
        row_data = {c: json_safe_value(row_data.get(c, None)) for c in APP_COLUMNS}
        payloads.append({
            "project_id": project_id,
            "row_order": int(i),
            "row_data": row_data,
            "updated_by": user,
            "updated_at": now_iso(),
        })

    if payloads:
        sb.table("allocation_rows").insert(payloads).execute()

    sb.table("projects").update({"updated_at": now_iso()}).eq("id", project_id).execute()
    log_audit_event(project_id, "shared_table_cleared", user, {})
    log_audit_event(project_id, "rows_saved", user, {"rows": len(payloads)})

def next_version_no(project_id):
    res = (sb.table("allocation_versions").select("version_no")
           .eq("project_id", project_id).order("version_no", desc=True).limit(1).execute())
    return 1 if not res.data else int(res.data[0]["version_no"]) + 1

def save_version(project_id, df, user, note):
    clean = normalize_uploaded_df(df)
    snap = json_safe_records(clean)
    vno = next_version_no(project_id)

    sb.table("allocation_versions").insert({
        "project_id": project_id,
        "version_no": vno,
        "snapshot": snap,
        "saved_by": user,
        "save_note": note,
        "created_at": now_iso(),
    }).execute()

    sb.table("save_events").insert({
        "project_id": project_id,
        "event_type": "save_version",
        "event_by": user,
        "event_note": note,
        "created_at": now_iso(),
    }).execute()

    log_audit_event(project_id, "version_saved", user, {"version_no": vno, "note": note})
    return vno

def list_versions(project_id):
    return (sb.table("allocation_versions")
            .select("id,version_no,saved_by,save_note,created_at")
            .eq("project_id", project_id).order("version_no", desc=True).execute().data or [])

def load_version(version_id):
    return sb.table("allocation_versions").select("*").eq("id", version_id).single().execute().data

def delete_all_versions(project_id):
    """Delete all saved version snapshots for the selected project."""
    return sb.table("allocation_versions").delete().eq("project_id", project_id).execute()


def clear_shared_allocation_table(project_id, user):
    """Delete all allocation rows for the selected project, leaving the project itself intact."""
    sb.table("allocation_rows").delete().eq("project_id", project_id).execute()
    sb.table("save_events").insert({
        "project_id": project_id,
        "event_type": "clear_allocation_rows",
        "event_by": user,
        "event_note": "Admin cleared shared allocation table",
        "created_at": now_iso(),
    }).execute()
    sb.table("projects").update({"updated_at": now_iso()}).eq("id", project_id).execute()


# ---------------- Conflict/deconfliction ----------------
def freq_overlap(a0, a1, b0, b1, guard=0.0):
    return min(a1, b1) > max(a0, b0) - guard

def time_overlap(a0, a1, b0, b1):
    return min(a1, b1) > max(a0, b0)

def detect_conflicts_generic(df, group_field, guard_mhz=0.0):
    req = {"StartF", "EndF", "StartTime", "EndTime", group_field}
    if not req.issubset(df.columns):
        return None
    d = df.copy()
    d["Group"] = d[group_field].astype(str)
    d["StartSec"] = d["StartTime"].apply(parse_time_one)
    d["EndSec"] = d["EndTime"].apply(parse_time_one)
    d = d.loc[d["StartSec"].notna() & d["EndSec"].notna()].copy()
    if d.empty:
        return None
    d["EndSecUnwrapped"] = np.where(d["EndSec"] < d["StartSec"], d["EndSec"] + 86400, d["EndSec"])
    d = d.reset_index(drop=True)
    rows = []
    for i in range(len(d)-1):
        for j in range(i+1, len(d)):
            if d.loc[i, "Group"] == d.loc[j, "Group"]:
                continue
            f_left = max(d.loc[i, "StartF"], d.loc[j, "StartF"])
            f_right = min(d.loc[i, "EndF"], d.loc[j, "EndF"])
            if f_right <= f_left - guard_mhz:
                continue
            t_start = max(d.loc[i, "StartSec"], d.loc[j, "StartSec"])
            t_end = min(d.loc[i, "EndSecUnwrapped"], d.loc[j, "EndSecUnwrapped"])
            if t_end <= t_start:
                continue
            rows.append({"FreqLeft": f_left, "FreqRight": f_right, "StartOverlap": t_start, "EndOverlap": t_end,
                         "GroupA": d.loc[i, "Group"], "GroupB": d.loc[j, "Group"], "OverlapMin": (t_end-t_start)/60})
    if not rows:
        return None
    out = pd.DataFrame(rows)
    out["OverlapStartHM"] = out["StartOverlap"].apply(fmt_hhmm)
    out["OverlapEndHM"] = out["EndOverlap"].apply(fmt_hhmm)
    return out.sort_values(["FreqLeft", "StartOverlap"]).reset_index(drop=True)

def placement_is_valid(candidate, item, placed, pad_sec, guard_mhz):
    end = candidate + item["DurationSec"]
    for p in placed:
        if not freq_overlap(item["StartF"], item["EndF"], p["StartF"], p["EndF"], guard_mhz):
            continue
        if time_overlap(candidate, end, p["PlacedStartSec"] - pad_sec, p["PlacedEndSec"] + pad_sec):
            return False
    return True

def auto_deconflict_smart(df, max_shift_sec, pad_sec, anchor_sec, guard_mhz=0.0, allow_earlier=True, priority_mode="Power + Original Time"):
    """
    Time deconfliction that actually honors the Anchor field.

    Previous behavior preferred the original start time first. That made the
    sidebar Anchor look like it was not changing anything. This version packs
    conflicting rows from the selected anchor forward, while still respecting
    max_shift_sec and frequency guard band rules.
    """
    req = {"StartTime", "EndTime", "StartF", "EndF", "PowerW"}
    if not req.issubset(df.columns):
        return df.copy()

    d = df.copy()
    d["OrigStartSec"] = d["StartTime"].apply(parse_time_one)
    d["OrigEndSec"] = d["EndTime"].apply(parse_time_one)
    d = d.loc[d["OrigStartSec"].notna() & d["OrigEndSec"].notna()].copy()

    if d.empty:
        return df.copy()

    d["OrigEndSec"] = np.where(
        d["OrigEndSec"] < d["OrigStartSec"],
        d["OrigEndSec"] + 86400,
        d["OrigEndSec"],
    )
    d["DurationSec"] = d["OrigEndSec"] - d["OrigStartSec"]

    if priority_mode == "Highest Power First":
        d = d.sort_values(["PowerW", "OrigStartSec"], ascending=[False, True])
    elif priority_mode == "Shortest Duration First":
        d = d.sort_values(["DurationSec", "OrigStartSec"], ascending=[True, True])
    else:
        d = d.sort_values(["PowerW", "OrigStartSec"], ascending=[False, True])

    placed, out_rows = [], []
    anchor_sec = float(anchor_sec)
    horizon_end = max(24 * 3600, anchor_sec + float(max_shift_sec), float(d["OrigEndSec"].max()) + float(max_shift_sec))

    for _, row in d.iterrows():
        item = row.to_dict()
        req_start = float(item["OrigStartSec"])
        dur = float(item["DurationSec"])

        # If earlier moves are allowed, pack from the anchor.
        # If earlier moves are not allowed, never place before the original requested time or anchor.
        earliest = anchor_sec if allow_earlier else max(anchor_sec, req_start)
        latest = min(horizon_end - dur, req_start + float(max_shift_sec))

        # If the original row is before the selected anchor, allow it to move forward at least to the anchor.
        if latest < earliest:
            latest = earliest

        candidates = set()

        # Always try the anchor/earliest first so changing Anchor visibly changes the schedule.
        candidates.add(earliest)

        # Try just after any already-placed conflicting block.
        for p in placed:
            if freq_overlap(item["StartF"], item["EndF"], p["StartF"], p["EndF"], guard_mhz):
                candidates.add(float(p["PlacedEndSec"]) + float(pad_sec))
                if allow_earlier:
                    candidates.add(float(p["PlacedStartSec"]) - float(pad_sec) - dur)

        # Backup scan. One-minute granularity gives better packing than five minutes.
        if latest >= earliest:
            candidates.update(np.arange(earliest, latest + 1, 60))

        # Only legal finite candidates.
        candidates = [
            float(c)
            for c in candidates
            if np.isfinite(c) and c >= earliest and c <= latest
        ]

        # Sort from anchor/earliest forward. This is the key anchor fix.
        candidates = sorted(candidates, key=lambda c: (abs(c - earliest), c))

        chosen, ok = earliest, False
        for cand in candidates:
            if placement_is_valid(cand, item, placed, pad_sec, guard_mhz):
                chosen, ok = cand, True
                break

        item["PlacedStartSec"] = chosen
        item["PlacedEndSec"] = chosen + dur
        item["ShiftSec"] = chosen - req_start
        item["Placed"] = ok
        item["StartTimeDC"] = fmt_hhmm(chosen)
        item["EndTimeDC"] = fmt_hhmm(chosen + dur)

        placed.append(item)
        out_rows.append(item)

    out = pd.DataFrame(out_rows).sort_values(".row_id")
    keep = [".row_id", "PlacedStartSec", "PlacedEndSec", "ShiftSec", "Placed", "StartTimeDC", "EndTimeDC"]
    return df.merge(out[keep], on=".row_id", how="left")

def conflict_summary(conflicts):
    """Build a display table for conflict records."""
    if conflicts is None or conflicts.empty:
        return pd.DataFrame({"Message": ["No conflicts."]})

    d = conflicts.copy()
    return pd.DataFrame({
        "Freq Start (MHz)": pd.to_numeric(d.get("FreqLeft"), errors="coerce").round(3),
        "Freq End (MHz)": pd.to_numeric(d.get("FreqRight"), errors="coerce").round(3),
        "Window": d.get("OverlapStartHM", "").astype(str) + " – " + d.get("OverlapEndHM", "").astype(str),
        "Overlap (min)": pd.to_numeric(d.get("OverlapMin"), errors="coerce").round(1),
        "Group A": d.get("GroupA", ""),
        "Group B": d.get("GroupB", ""),
    })


def conflict_recommendations(conflicts, label="Equipment"):
    """Score conflicts and recommend practical deconfliction actions."""
    if conflicts is None or conflicts.empty:
        return pd.DataFrame({"Message": [f"No {label} conflicts."]})

    d = conflicts.copy()
    d["OverlapMHz"] = (d["FreqRight"] - d["FreqLeft"]).round(3)
    d["OverlapMinRounded"] = d["OverlapMin"].round(1)
    d["SeverityScore"] = (d["OverlapMHz"].clip(lower=0) * d["OverlapMin"].clip(lower=0)).round(2)

    def severity(row):
        if row["OverlapMin"] >= 60 or row["OverlapMHz"] >= 10 or row["SeverityScore"] >= 300:
            return "High"
        if row["OverlapMin"] >= 15 or row["OverlapMHz"] >= 3 or row["SeverityScore"] >= 50:
            return "Medium"
        return "Low"

    def action(row):
        delay_to = fmt_hhmm(float(row["EndOverlap"]) + 5 * 60)
        shift_freq = round(float(row["FreqRight"]) + 0.025, 3)
        if row["OverlapMin"] >= 30:
            return f"Separate time window first: move {row['GroupB']} start to {delay_to}, or schedule outside {row['OverlapStartHM']}–{row['OverlapEndHM']}."
        if row["OverlapMHz"] >= 5:
            return f"Frequency separation recommended: move one system above {shift_freq} MHz or apply a guard band."
        return f"Low conflict: verify mission priority; add guard band or shift {row['GroupB']} after {delay_to}."

    out = pd.DataFrame({
        "Type": label,
        "Severity": d.apply(severity, axis=1),
        "Score": d["SeverityScore"],
        "Freq Start (MHz)": d["FreqLeft"].round(3),
        "Freq End (MHz)": d["FreqRight"].round(3),
        "Overlap MHz": d["OverlapMHz"],
        "Window": d["OverlapStartHM"] + " – " + d["OverlapEndHM"],
        "Overlap Min": d["OverlapMinRounded"],
        "Group A": d["GroupA"],
        "Group B": d["GroupB"],
        "Recommended Action": d.apply(action, axis=1),
    })

    sev_order = {"High": 0, "Medium": 1, "Low": 2}
    out["SeverityOrder"] = out["Severity"].map(sev_order).fillna(9)
    out = out.sort_values(["SeverityOrder", "Score"], ascending=[True, False]).drop(columns=["SeverityOrder"])
    return out.reset_index(drop=True)


def infer_planning_band(df, band_start=None, band_end=None):
    """Determine planning band from user inputs or uploaded data bounds."""
    start = pd.to_numeric(band_start, errors="coerce") if band_start not in [None, ""] else np.nan
    end = pd.to_numeric(band_end, errors="coerce") if band_end not in [None, ""] else np.nan

    data_min = float(pd.to_numeric(df.get("StartF"), errors="coerce").min())
    data_max = float(pd.to_numeric(df.get("EndF"), errors="coerce").max())

    if not np.isfinite(start):
        start = data_min
    if not np.isfinite(end):
        end = data_max

    if not np.isfinite(start) or not np.isfinite(end) or end <= start:
        return None, None

    return float(start), float(end)


def row_time_overlap(row, start_sec, end_sec):
    """Check if a dataframe row overlaps a time window."""
    rs = parse_time_one(row.get("StartTime"))
    re = parse_time_one(row.get("EndTime"))
    if pd.isna(rs) or pd.isna(re):
        return True
    if re < rs:
        re += 86400
    return time_overlap(float(rs), float(re), float(start_sec), float(end_sec))


def candidate_frequency_is_clear(df, row_id_a, row_id_b, cand_start, cand_end, start_sec, end_sec, guard_mhz):
    """Return True when candidate frequency range is clear during the conflict time window."""
    for _, r in df.iterrows():
        rid = r.get(".row_id")
        if rid in [row_id_a, row_id_b]:
            continue

        sf = pd.to_numeric(r.get("StartF"), errors="coerce")
        ef = pd.to_numeric(r.get("EndF"), errors="coerce")
        if not np.isfinite(sf) or not np.isfinite(ef):
            continue

        if not row_time_overlap(r, start_sec, end_sec):
            continue

        if freq_overlap(cand_start, cand_end, float(sf), float(ef), guard_mhz):
            return False

    return True


def smart_frequency_suggestions(df, conflicts, group_field, band_start, band_end, guard_mhz=0.025, step_mhz=0.025, max_each=3):
    """Suggest alternate frequency ranges and time shifts for conflicts."""
    if conflicts is None or conflicts.empty:
        return pd.DataFrame({"Message": ["No conflicts to plan around."]})

    d = df.copy()
    band_start, band_end = infer_planning_band(d, band_start, band_end)
    if band_start is None:
        return pd.DataFrame({"Message": ["Could not determine planning band. Enter planning band start/end MHz."]})

    suggestions = []

    for _, c in conflicts.iterrows():
        group_a = str(c.get("GroupA"))
        group_b = str(c.get("GroupB"))
        conflict_start = float(c.get("StartOverlap"))
        conflict_end = float(c.get("EndOverlap"))
        overlap_mhz = float(c.get("FreqRight") - c.get("FreqLeft"))

        rows_a = d[d[group_field].astype(str) == group_a].copy() if group_field in d.columns else pd.DataFrame()
        rows_b = d[d[group_field].astype(str) == group_b].copy() if group_field in d.columns else pd.DataFrame()

        # Prefer moving the lower-power group if possible.
        pow_a = pd.to_numeric(rows_a.get("PowerW", pd.Series([0])), errors="coerce").max()
        pow_b = pd.to_numeric(rows_b.get("PowerW", pd.Series([0])), errors="coerce").max()

        move_group = group_b if (pd.isna(pow_a) or pd.isna(pow_b) or pow_b <= pow_a) else group_a
        keep_group = group_a if move_group == group_b else group_b
        move_rows = rows_b if move_group == group_b else rows_a

        if move_rows.empty:
            continue

        # Use widest row in moving group as representative.
        move_rows["WidthMHz"] = pd.to_numeric(move_rows["EndF"], errors="coerce") - pd.to_numeric(move_rows["StartF"], errors="coerce")
        move_row = move_rows.sort_values("WidthMHz", ascending=False).iloc[0]
        width = float(move_row["WidthMHz"]) if np.isfinite(move_row["WidthMHz"]) and move_row["WidthMHz"] > 0 else max(overlap_mhz, step_mhz)
        row_id = move_row.get(".row_id")

        # Search candidate starts across band.
        step = max(float(step_mhz), 0.001)
        candidates = np.arange(band_start, max(band_start, band_end - width) + step / 2, step)

        found = 0
        for cs in candidates:
            ce = float(cs + width)
            if ce > band_end:
                continue

            if candidate_frequency_is_clear(d, row_id, None, float(cs), ce, conflict_start, conflict_end, float(guard_mhz)):
                dist_from_original = min(abs(float(cs) - float(move_row["StartF"])), abs(ce - float(move_row["EndF"])))
                suggestions.append({
                    "Plan Type": "Frequency move",
                    "Conflict": f"{group_a} vs {group_b}",
                    "Move Group": move_group,
                    "Keep Group": keep_group,
                    "Suggested Start MHz": round(float(cs), 3),
                    "Suggested End MHz": round(ce, 3),
                    "Width MHz": round(width, 3),
                    "Shift Distance MHz": round(dist_from_original, 3),
                    "Conflict Window": f"{c.get('OverlapStartHM')}–{c.get('OverlapEndHM')}",
                    "Reason": "Candidate range is clear during the conflict window with guard band applied.",
                })
                found += 1
                if found >= int(max_each):
                    break

        # Always include a time fallback.
        delay_to = fmt_hhmm(float(c.get("EndOverlap")) + 5 * 60)
        suggestions.append({
            "Plan Type": "Time move",
            "Conflict": f"{group_a} vs {group_b}",
            "Move Group": move_group,
            "Keep Group": keep_group,
            "Suggested Start MHz": "",
            "Suggested End MHz": "",
            "Width MHz": "",
            "Shift Distance MHz": "",
            "Conflict Window": f"{c.get('OverlapStartHM')}–{c.get('OverlapEndHM')}",
            "Reason": f"If no clean frequency is acceptable, move {move_group} after {delay_to}.",
        })

    if not suggestions:
        return pd.DataFrame({"Message": ["No smart suggestions found. Try widening the planning band or reducing guard band."]})

    return pd.DataFrame(suggestions)


def apply_smart_plan_to_sheet(sheet_df, recommendation):
    """Apply one Smart Planner recommendation to the active sheet."""
    out = normalize_uploaded_df(sheet_df).copy()
    rec = dict(recommendation)

    plan_type = str(rec.get("Plan Type", ""))
    move_group = str(rec.get("Move Group", "")).strip()

    if not move_group:
        raise ValueError("Recommendation does not include a Move Group.")

    # Match against Equipment, Unit, or Tech so recommendations from either conflict type can be applied.
    mask = pd.Series(False, index=out.index)
    for col in ["Equipment", "Unit", "Tech"]:
        if col in out.columns:
            mask = mask | (out[col].astype(str).str.strip() == move_group)

    if not mask.any():
        raise ValueError(f"No rows found for Move Group: {move_group}")

    changed_rows = int(mask.sum())

    if plan_type == "Frequency move":
        new_start = pd.to_numeric(rec.get("Suggested Start MHz"), errors="coerce")
        new_end = pd.to_numeric(rec.get("Suggested End MHz"), errors="coerce")

        if not np.isfinite(new_start) or not np.isfinite(new_end) or new_end <= new_start:
            raise ValueError("Selected frequency recommendation does not contain a valid frequency range.")

        out.loc[mask, "Start Frequency (MHz)"] = float(new_start)
        out.loc[mask, "End Frequency (MHz)"] = float(new_end)
        out.loc[mask, "Center Frequency (MHz)"] = (float(new_start) + float(new_end)) / 2.0
        out.loc[mask, "Bandwidth (MHz)"] = float(new_end) - float(new_start)

        return out, f"Applied frequency move to {changed_rows} row(s) for {move_group}: {new_start:.3f}–{new_end:.3f} MHz."

    if plan_type == "Time move":
        reason = str(rec.get("Reason", ""))
        # Extract HH:MM after the word "after".
        match = re.search(r"after\\s+(\\d{2}:\\d{2})", reason)
        if not match:
            raise ValueError("Selected time recommendation does not contain a parseable new time.")

        new_start_txt = match.group(1)
        new_start_sec = parse_time_one(new_start_txt)
        if pd.isna(new_start_sec):
            raise ValueError("Could not parse recommended start time.")

        for idx in out.index[mask]:
            old_start = parse_time_one(out.at[idx, "Start Time"])
            old_end = parse_time_one(out.at[idx, "End Time"])
            if pd.isna(old_start) or pd.isna(old_end):
                continue
            if old_end < old_start:
                old_end += 86400
            duration = old_end - old_start
            out.at[idx, "Start Time"] = fmt_hhmm(float(new_start_sec))
            out.at[idx, "End Time"] = fmt_hhmm(float(new_start_sec) + float(duration))

        return out, f"Applied time move to {changed_rows} row(s) for {move_group}: start after {new_start_txt}."

    raise ValueError(f"Unsupported recommendation type: {plan_type}")


def conflict_recommendations(conflicts, label="Equipment"):
    """Score conflicts and recommend practical deconfliction actions."""
    if conflicts is None or conflicts.empty:
        return pd.DataFrame({"Message": [f"No {label} conflicts."]})

    d = conflicts.copy()
    d["OverlapMHz"] = pd.to_numeric(d.get("FreqRight"), errors="coerce") - pd.to_numeric(d.get("FreqLeft"), errors="coerce")
    d["OverlapMHz"] = d["OverlapMHz"].fillna(0).round(3)
    d["OverlapMinRounded"] = pd.to_numeric(d.get("OverlapMin"), errors="coerce").fillna(0).round(1)
    d["SeverityScore"] = (d["OverlapMHz"].clip(lower=0) * d["OverlapMinRounded"].clip(lower=0)).round(2)

    def severity(row):
        if row["OverlapMinRounded"] >= 60 or row["OverlapMHz"] >= 10 or row["SeverityScore"] >= 300:
            return "High"
        if row["OverlapMinRounded"] >= 15 or row["OverlapMHz"] >= 3 or row["SeverityScore"] >= 50:
            return "Medium"
        return "Low"

    def action(row):
        try:
            delay_to = fmt_hhmm(float(row.get("EndOverlap", 0)) + 5 * 60)
        except Exception:
            delay_to = "later"
        try:
            shift_freq = round(float(row.get("FreqRight", 0)) + 0.025, 3)
        except Exception:
            shift_freq = ""

        if row["OverlapMinRounded"] >= 30:
            return f"Separate time window first: move {row.get('GroupB', '')} start to {delay_to}, or schedule outside {row.get('OverlapStartHM', '')}–{row.get('OverlapEndHM', '')}."
        if row["OverlapMHz"] >= 5:
            return f"Frequency separation recommended: move one system above {shift_freq} MHz or apply a guard band."
        return f"Low conflict: verify mission priority; add guard band or shift {row.get('GroupB', '')} after {delay_to}."

    out = pd.DataFrame({
        "Type": label,
        "Severity": d.apply(severity, axis=1),
        "Score": d["SeverityScore"],
        "Freq Start (MHz)": pd.to_numeric(d.get("FreqLeft"), errors="coerce").round(3),
        "Freq End (MHz)": pd.to_numeric(d.get("FreqRight"), errors="coerce").round(3),
        "Overlap MHz": d["OverlapMHz"],
        "Window": d.get("OverlapStartHM", "").astype(str) + " – " + d.get("OverlapEndHM", "").astype(str),
        "Overlap Min": d["OverlapMinRounded"],
        "Group A": d.get("GroupA", ""),
        "Group B": d.get("GroupB", ""),
        "Recommended Action": d.apply(action, axis=1),
    })

    sev_order = {"High": 0, "Medium": 1, "Low": 2}
    out["SeverityOrder"] = out["Severity"].map(sev_order).fillna(9)
    return out.sort_values(["SeverityOrder", "Score"], ascending=[True, False]).drop(columns=["SeverityOrder"]).reset_index(drop=True)


def combined_conflict_recommendations(conf_eq, conf_ut, unittech_label):
    """Combine equipment and unit/tech conflict recommendations."""
    frames = []

    eq = conflict_recommendations(conf_eq, "Equipment")
    if "Message" not in eq.columns:
        frames.append(eq)

    ut = conflict_recommendations(conf_ut, unittech_label)
    if "Message" not in ut.columns:
        frames.append(ut)

    if not frames:
        return pd.DataFrame({"Message": ["No conflicts requiring recommendations."]})

    return pd.concat(frames, ignore_index=True)



def infer_planning_band(df, band_start=None, band_end=None):
    """Determine planning band from user inputs or uploaded data bounds."""
    start = pd.to_numeric(band_start, errors="coerce") if band_start not in [None, ""] else np.nan
    end = pd.to_numeric(band_end, errors="coerce") if band_end not in [None, ""] else np.nan

    data_min = pd.to_numeric(df.get("StartF"), errors="coerce").min()
    data_max = pd.to_numeric(df.get("EndF"), errors="coerce").max()

    if not np.isfinite(start):
        start = data_min
    if not np.isfinite(end):
        end = data_max

    if not np.isfinite(start) or not np.isfinite(end) or float(end) <= float(start):
        return None, None

    return float(start), float(end)


def row_time_overlap(row, start_sec, end_sec):
    """Check if a dataframe row overlaps a time window."""
    rs = parse_time_one(row.get("StartTime"))
    re = parse_time_one(row.get("EndTime"))
    if pd.isna(rs) or pd.isna(re):
        return True
    if re < rs:
        re += 86400
    return time_overlap(float(rs), float(re), float(start_sec), float(end_sec))


def candidate_frequency_is_clear(df, row_id_a, row_id_b, cand_start, cand_end, start_sec, end_sec, guard_mhz):
    """Return True when candidate frequency range is clear during the conflict time window."""
    for _, r in df.iterrows():
        rid = r.get(".row_id")
        if rid in [row_id_a, row_id_b]:
            continue

        sf = pd.to_numeric(r.get("StartF"), errors="coerce")
        ef = pd.to_numeric(r.get("EndF"), errors="coerce")
        if not np.isfinite(sf) or not np.isfinite(ef):
            continue

        if not row_time_overlap(r, start_sec, end_sec):
            continue

        if freq_overlap(cand_start, cand_end, float(sf), float(ef), guard_mhz):
            return False

    return True


def smart_frequency_suggestions(df, conflicts, group_field, band_start, band_end, guard_mhz=0.025, step_mhz=0.025, max_each=3):
    """Suggest alternate frequency ranges and time shifts for conflicts."""
    if conflicts is None or conflicts.empty:
        return pd.DataFrame({"Message": ["No conflicts to plan around."]})

    d = df.copy()
    band_start, band_end = infer_planning_band(d, band_start, band_end)
    if band_start is None:
        return pd.DataFrame({"Message": ["Could not determine planning band. Enter planning band start/end MHz."]})

    suggestions = []

    for _, c in conflicts.iterrows():
        group_a = str(c.get("GroupA"))
        group_b = str(c.get("GroupB"))
        conflict_start = float(c.get("StartOverlap", 0))
        conflict_end = float(c.get("EndOverlap", conflict_start))
        overlap_mhz = float(c.get("FreqRight", 0) - c.get("FreqLeft", 0))

        if group_field not in d.columns:
            continue

        rows_a = d[d[group_field].astype(str) == group_a].copy()
        rows_b = d[d[group_field].astype(str) == group_b].copy()

        pow_a = pd.to_numeric(rows_a.get("PowerW", pd.Series([0])), errors="coerce").max()
        pow_b = pd.to_numeric(rows_b.get("PowerW", pd.Series([0])), errors="coerce").max()

        move_group = group_b if (pd.isna(pow_a) or pd.isna(pow_b) or pow_b <= pow_a) else group_a
        keep_group = group_a if move_group == group_b else group_b
        move_rows = rows_b if move_group == group_b else rows_a

        if move_rows.empty:
            continue

        move_rows["WidthMHz"] = pd.to_numeric(move_rows["EndF"], errors="coerce") - pd.to_numeric(move_rows["StartF"], errors="coerce")
        move_row = move_rows.sort_values("WidthMHz", ascending=False).iloc[0]
        width = float(move_row["WidthMHz"]) if np.isfinite(move_row["WidthMHz"]) and move_row["WidthMHz"] > 0 else max(overlap_mhz, step_mhz)
        row_id = move_row.get(".row_id")

        step = max(float(step_mhz), 0.001)
        candidates = np.arange(band_start, max(band_start, band_end - width) + step / 2, step)

        found = 0
        for cs in candidates:
            ce = float(cs + width)
            if ce > band_end:
                continue

            if candidate_frequency_is_clear(d, row_id, None, float(cs), ce, conflict_start, conflict_end, float(guard_mhz)):
                dist_from_original = min(abs(float(cs) - float(move_row["StartF"])), abs(ce - float(move_row["EndF"])))
                suggestions.append({
                    "Plan Type": "Frequency move",
                    "Conflict": f"{group_a} vs {group_b}",
                    "Move Group": move_group,
                    "Keep Group": keep_group,
                    "Suggested Start MHz": round(float(cs), 3),
                    "Suggested End MHz": round(ce, 3),
                    "Width MHz": round(width, 3),
                    "Shift Distance MHz": round(dist_from_original, 3),
                    "Conflict Window": f"{c.get('OverlapStartHM')}–{c.get('OverlapEndHM')}",
                    "Reason": "Candidate range is clear during the conflict window with guard band applied.",
                })
                found += 1
                if found >= int(max_each):
                    break

        delay_to = fmt_hhmm(float(c.get("EndOverlap", 0)) + 5 * 60)
        suggestions.append({
            "Plan Type": "Time move",
            "Conflict": f"{group_a} vs {group_b}",
            "Move Group": move_group,
            "Keep Group": keep_group,
            "Suggested Start MHz": "",
            "Suggested End MHz": "",
            "Width MHz": "",
            "Shift Distance MHz": "",
            "Conflict Window": f"{c.get('OverlapStartHM')}–{c.get('OverlapEndHM')}",
            "Reason": f"If no clean frequency is acceptable, move {move_group} after {delay_to}.",
        })

    if not suggestions:
        return pd.DataFrame({"Message": ["No smart suggestions found. Try widening the planning band or reducing guard band."]})

    return pd.DataFrame(suggestions)


def combined_smart_plan(df, conf_eq, conf_ut, grp_ut, band_start, band_end, guard_mhz, step_mhz, max_each):
    """Combine equipment and unit/tech smart plan suggestions."""
    frames = []

    eq = smart_frequency_suggestions(df, conf_eq, "Equipment", band_start, band_end, guard_mhz, step_mhz, max_each)
    if "Message" not in eq.columns:
        eq.insert(0, "Conflict Type", "Equipment")
        frames.append(eq)

    ut = smart_frequency_suggestions(df, conf_ut, grp_ut, band_start, band_end, guard_mhz, step_mhz, max_each)
    if "Message" not in ut.columns:
        ut.insert(0, "Conflict Type", grp_ut)
        frames.append(ut)

    if not frames:
        return pd.DataFrame({"Message": ["No smart planning suggestions available."]})

    return pd.concat(frames, ignore_index=True)




def summary_by_group(df, group_col, conf_df=None):
    """Aggregate allocations, conflicts, power, bandwidth by a group."""
    if df is None or df.empty or group_col not in df.columns:
        return pd.DataFrame({"Message": [f"No {group_col} data found."]})

    d = df.copy()
    d[group_col] = d[group_col].fillna("(blank)").astype(str)
    d["BandwidthCalc"] = pd.to_numeric(d.get("EndF"), errors="coerce") - pd.to_numeric(d.get("StartF"), errors="coerce")
    d["PowerCalc"] = pd.to_numeric(d.get("PowerW"), errors="coerce").fillna(0)

    out = d.groupby(group_col, dropna=False).agg(
        Allocations=("Equipment", "count"),
        UniqueFreqs=("CenterF", lambda x: pd.to_numeric(x, errors="coerce").dropna().nunique()),
        TotalBandwidthMHz=("BandwidthCalc", "sum"),
        AvgPowerW=("PowerCalc", "mean"),
        MaxPowerW=("PowerCalc", "max"),
    ).reset_index()

    out["Conflicts"] = 0
    if conf_df is not None and not conf_df.empty:
        conflict_counts = {}
        for _, r in conf_df.iterrows():
            for g in [str(r.get("GroupA", "")), str(r.get("GroupB", ""))]:
                conflict_counts[g] = conflict_counts.get(g, 0) + 1
        out["Conflicts"] = out[group_col].map(conflict_counts).fillna(0).astype(int)

    out["At Risk"] = np.where(out["Conflicts"] > 0, "At Risk", "Good")
    out["TotalBandwidthMHz"] = pd.to_numeric(out["TotalBandwidthMHz"], errors="coerce").round(3)
    out["AvgPowerW"] = pd.to_numeric(out["AvgPowerW"], errors="coerce").round(3)
    out["MaxPowerW"] = pd.to_numeric(out["MaxPowerW"], errors="coerce").round(3)
    return out.sort_values(["Conflicts", "Allocations"], ascending=[False, False]).reset_index(drop=True)


def band_utilization_summary(df):
    """Summarize allocation usage by workbook/band tab or frequency span."""
    if df is None or df.empty:
        return pd.DataFrame({"Message": ["No allocation data."]})

    d = df.copy()
    if "SourceSheet" in d.columns:
        group = "SourceSheet"
    elif "RequestSheet" in d.columns:
        group = "RequestSheet"
    else:
        # Fallback: create rough 100 MHz bins.
        sf = pd.to_numeric(d.get("StartF"), errors="coerce")
        d["Band"] = (np.floor(sf / 100) * 100).astype("Int64").astype(str) + "-" + (np.floor(sf / 100) * 100 + 100).astype("Int64").astype(str)
        group = "Band"

    d["StartFNum"] = pd.to_numeric(d.get("StartF"), errors="coerce")
    d["EndFNum"] = pd.to_numeric(d.get("EndF"), errors="coerce")
    d["BWNum"] = d["EndFNum"] - d["StartFNum"]

    out = d.groupby(group, dropna=False).agg(
        Allocations=("Equipment", "count"),
        MinMHz=("StartFNum", "min"),
        MaxMHz=("EndFNum", "max"),
        UsedBandwidthMHz=("BWNum", "sum"),
        UniqueFreqs=("CenterF", lambda x: pd.to_numeric(x, errors="coerce").dropna().nunique()),
    ).reset_index().rename(columns={group: "Band"})

    out["BandSpanMHz"] = (out["MaxMHz"] - out["MinMHz"]).round(3)
    out["UsedBandwidthMHz"] = out["UsedBandwidthMHz"].round(3)
    out["ApproxUtilizationPct"] = np.where(out["BandSpanMHz"] > 0, (out["UsedBandwidthMHz"] / out["BandSpanMHz"] * 100).clip(0, 999).round(1), 0)
    return out.sort_values("Allocations", ascending=False).reset_index(drop=True)


def allocation_validation_summary(df):
    """Find low confidence/missing metadata rows."""
    if df is None or df.empty:
        return pd.DataFrame({"Message": ["No allocation data."]})

    d = df.copy()
    issues = []

    for idx, r in d.iterrows():
        row_issues = []
        conf = pd.to_numeric(r.get("Match Confidence"), errors="coerce")
        if pd.notna(conf) and conf < 90:
            row_issues.append(f"Low match confidence: {conf}%")
        for col in ["Unit", "Sponsor", "Sponser", "Tech"]:
            if col in d.columns and (pd.isna(r.get(col)) or str(r.get(col)).strip() in ["", "None", "nan", "(blank)"]):
                row_issues.append(f"Missing {col}")
        if "Active" in d.columns and not to_active_bool(r.get("Active")):
            row_issues.append("Inactive / turned off")
        if row_issues:
            issues.append({
                "Row": idx,
                "Equipment": r.get("Equipment", ""),
                "Tech": r.get("Tech", ""),
                "Unit": r.get("Unit", ""),
                "Sponsor": r.get("Sponsor", r.get("Sponser", "")),
                "Match Confidence": r.get("Match Confidence", ""),
                "Issues": "; ".join(row_issues),
            })

    if not issues:
        return pd.DataFrame({"Message": ["No validation issues found."]})
    return pd.DataFrame(issues)


def geographic_reuse_summary(df):
    """Approximate geographic reuse opportunities by frequency and distance."""
    if df is None or df.empty:
        return pd.DataFrame({"Message": ["No allocation data."]})

    d = df.copy()
    d["Latitude"] = pd.to_numeric(d.get("Latitude"), errors="coerce")
    d["Longitude"] = pd.to_numeric(d.get("Longitude"), errors="coerce")
    d["CenterF"] = pd.to_numeric(d.get("CenterF"), errors="coerce")
    d = d.dropna(subset=["Latitude", "Longitude", "CenterF"])

    if len(d) < 2:
        return pd.DataFrame({"Message": ["Need at least two mapped allocations with CenterF for reuse analysis."]})

    rows = []
    for freq, g in d.groupby("CenterF"):
        g = g.reset_index(drop=True)
        if len(g) < 2:
            continue
        for i in range(len(g)):
            for j in range(i + 1, len(g)):
                a, b = g.iloc[i], g.iloc[j]
                dist_km = haversine_distance_km(a["Latitude"], a["Longitude"], b["Latitude"], b["Longitude"])
                rad_a = pd.to_numeric(str(a.get("CoverageRadius", "0")).replace("km","").replace("KM","").replace("NM","").replace("nm","").strip(), errors="coerce")
                rad_b = pd.to_numeric(str(b.get("CoverageRadius", "0")).replace("km","").replace("KM","").replace("NM","").replace("nm","").strip(), errors="coerce")
                rad_a = 0 if pd.isna(rad_a) else float(rad_a)
                rad_b = 0 if pd.isna(rad_b) else float(rad_b)
                # Treat radius values as km for this summary when the file uses 10km style.
                overlap_margin = (rad_a + rad_b) - dist_km
                rows.append({
                    "Center Frequency (MHz)": round(float(freq), 3),
                    "Site A": a.get("SiteName", a.get("Location", a.get("Equipment", ""))),
                    "Site B": b.get("SiteName", b.get("Location", b.get("Equipment", ""))),
                    "Unit A": a.get("Unit", ""),
                    "Unit B": b.get("Unit", ""),
                    "Distance km": round(dist_km, 2),
                    "Combined Radius": round(rad_a + rad_b, 2),
                    "Overlap Margin km": round(overlap_margin, 2),
                    "Reuse Risk": "High" if overlap_margin > 0 else "Low",
                })

    if not rows:
        return pd.DataFrame({"Message": ["No reused center frequencies found across mapped sites."]})
    return pd.DataFrame(rows).sort_values(["Reuse Risk", "Distance km"], ascending=[False, True]).reset_index(drop=True)



def command_dashboard_group_summary(df, group_col, conf_df=None):
    """Compact table for command dashboard summary cards."""
    if df is None or df.empty or group_col not in df.columns:
        return pd.DataFrame({"Message": [f"No {group_col} data available."]})

    d = df.copy()
    d[group_col] = d[group_col].fillna("(blank)").astype(str).replace({"None": "(blank)", "nan": "(blank)", "": "(blank)"})
    d["StartFNum"] = pd.to_numeric(d.get("StartF"), errors="coerce")
    d["EndFNum"] = pd.to_numeric(d.get("EndF"), errors="coerce")
    d["BWCalc"] = (d["EndFNum"] - d["StartFNum"]).fillna(0)

    out = d.groupby(group_col, dropna=False).agg(
        Allocations=("Equipment", "count"),
        Min_Separation_MHz=("BWCalc", lambda x: round(float(pd.to_numeric(x, errors="coerce").replace(0, np.nan).min()) if pd.to_numeric(x, errors="coerce").replace(0, np.nan).notna().any() else 0, 3)),
        Worst_Conflict_MHz=("BWCalc", lambda x: round(float(pd.to_numeric(x, errors="coerce").max()) if pd.to_numeric(x, errors="coerce").notna().any() else 0, 3)),
    ).reset_index()

    conflicts = {}
    if conf_df is not None and not conf_df.empty:
        for _, r in conf_df.iterrows():
            for g in [str(r.get("GroupA", "")), str(r.get("GroupB", ""))]:
                conflicts[g] = conflicts.get(g, 0) + 1

    out["Conflicts"] = out[group_col].map(conflicts).fillna(0).astype(int)
    out["At Risk"] = np.where(out["Conflicts"] > 0, np.maximum(1, (out["Conflicts"] / 2).astype(int)), 0)
    out["Min Time Separation"] = np.where(out["Conflicts"] > 0, "15 min", "0 min")
    out["Status"] = np.where(out["Conflicts"] > 5, "At Risk", "Good")
    out = out.rename(columns={
        group_col: group_col,
        "Min_Separation_MHz": "Min Separation (MHz)",
        "Worst_Conflict_MHz": "Worst Conflict (MHz)",
    })
    cols = [group_col, "Allocations", "Conflicts", "At Risk", "Min Separation (MHz)", "Min Time Separation", "Worst Conflict (MHz)", "Status"]
    return out[cols].sort_values(["Conflicts", "Allocations"], ascending=[False, False]).head(12).reset_index(drop=True)


def render_metric_html(total_allocations, conflict_count, risk_count):
    return f"""
    <div class="pcc-metric-row">
        <div class="pcc-small-metric">
            <div class="pcc-small-metric-label">Total Allocations</div>
            <div class="pcc-small-metric-value">{int(total_allocations):,}</div>
        </div>
        <div class="pcc-small-metric red">
            <div class="pcc-small-metric-label">Conflicts</div>
            <div class="pcc-small-metric-value">{int(conflict_count):,}</div>
        </div>
        <div class="pcc-small-metric amber">
            <div class="pcc-small-metric-label">At Risk</div>
            <div class="pcc-small-metric-value">{int(risk_count):,}</div>
        </div>
    </div>
    """


def render_band_bar(sheet_names=None, active_sheet=None):
    if not sheet_names:
        sheet_names = ["1350-1390", "1780-1850", "2025-2110", "2200-2300", "2310-2360", "2400-2490", "4400-4648.6", "4648.6-4940", "9200-10000", "14400-14830", "15150-15350", "15700-17700"]
    chips = []
    for i, name in enumerate(sheet_names[:14]):
        active = " active" if (active_sheet and str(name) == str(active_sheet)) or (not active_sheet and i == 0) else ""
        chips.append(f'<div class="pcc-band-chip{active}">{name}</div>')
    return '<div class="pcc-band-bar">' + "".join(chips) + '<div class="pcc-band-chip">+</div><div class="pcc-band-chip">☰</div></div>'




def web_metric_cards(active_count, inactive_count, eq_conflicts, tech_conflicts, project_status):
    """Render modern metric cards."""
    status_class = "green" if str(project_status).lower() in ["approved", "active", "open", "working"] else "amber"
    return f"""
    <div class="web-grid-4">
        <div class="web-card blue">
            <div class="web-card-label">Active Allocations</div>
            <div class="web-card-value">{int(active_count):,}</div>
            <div class="web-card-sub">Rows included in planning</div>
        </div>
        <div class="web-card amber">
            <div class="web-card-label">Inactive</div>
            <div class="web-card-value">{int(inactive_count):,}</div>
            <div class="web-card-sub">Hidden from analysis</div>
        </div>
        <div class="web-card red">
            <div class="web-card-label">Equipment Conflicts</div>
            <div class="web-card-value">{int(eq_conflicts):,}</div>
            <div class="web-card-sub">Current selected sheet</div>
        </div>
        <div class="web-card {status_class}">
            <div class="web-card-label">Project Status</div>
            <div class="web-card-value">{project_status}</div>
            <div class="web-card-sub">Workflow state</div>
        </div>
    </div>
    """


def app_topbar_html(project_name, user_name, project_status):
    status = str(project_status or "Working")
    badge_class = "status-good" if status.lower() in ["approved", "active", "open"] else ("status-risk" if status.lower() in ["working", "draft", "pending"] else "status-bad")
    return f"""
    <div class="app-shell-topbar">
        <div class="app-brand">
            <div class="app-logo">📡</div>
            <div>
                <div class="app-title-main">PCC6 SPECTRUM PLANNER</div>
                <div class="app-title-sub">{project_name} · Collaborative frequency, power, geographic reuse, and time deconfliction</div>
            </div>
        </div>
        <div class="app-classification">CONTROLLED UNCLASSIFIED INFORMATION (CUI)</div>
        <div class="app-actions">
            <span class="app-action-pill">User: {user_name}</span>
            <span class="status-badge {badge_class}">{status}</span>
        </div>
    </div>
    """


def section_header_html(title, caption="", right=""):
    return f"""
    <div class="web-section-header">
        <div>
            <div class="web-section-title">{title}</div>
            <div class="web-section-caption">{caption}</div>
        </div>
        <div>{right}</div>
    </div>
    """




# ---------------- Allocation Engine V47 ----------------

ALLOCATION_ENGINE_COLUMNS = [
    "Active",
    "Start Time", "End Time", "Equipment", "Center Frequency (MHz)",
    "Start Frequency (MHz)", "End Frequency (MHz)", "Bandwidth (MHz)",
    "Power (W)", "Tech", "Unit", "Sponsor", "Latitude", "Longitude",
    "Location", "Antenna Height", "Coverage Radius", "Site Name", "Notes",
    "NTC Area", "Grid / MGRS",
]

ALLOCATION_ENGINE_APPEND_COLUMNS = [
    "Allocation Status", "Conflict Status", "Reuse Group ID",
    "Frequency Locked", "Priority", "Priority Score",
    "Source Sheet", "Source Row", "Requested Frequency",
    "Approved Equipment", "Approved Frequency Source",
    "ISM Notes",
]

ALLOCATION_BANDS = [
    ("HF/VHF/UHF Below 1350", 0, 1350),
    ("1350-1390", 1350, 1390),
    ("1780-1850", 1780, 1850),
    ("2025-2110", 2025, 2110),
    ("2200-2300", 2200, 2300),
    ("2310-2360", 2310, 2360),
    ("2400-2490", 2400, 2490),
    ("4400-4648.6", 4400, 4648.6),
    ("4648.6-4940", 4648.6, 4940),
    ("9200-10000", 9200, 10000),
    ("14400-14830", 14400, 14830),
    ("15150-15350", 15150, 15350),
    ("15700-17700", 15700, 17700),
    ("Other / Needs Review", 17700, 999999),
]


def ae_norm_key(x):
    return re.sub(r"[^a-z0-9]+", "", str(x or "").strip().lower())


def ae_safe_get(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def ae_parse_count(value):
    m = re.search(r"\((\d+)\)", str(value or ""))
    return int(m.group(1)) if m else 1


def ae_parse_bandwidth_mhz(value):
    s = str(value or "").strip().upper().replace(" ", "")
    if not s or s in ["N/A", "NA", "NONE", "NULL"]:
        return None
    s = s.split("-")[0]
    if "MW" in s:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)M$", s)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+)M(\d+)", s)
    if m:
        return float(f"{m.group(1)}.{m.group(2)}")
    m = re.search(r"(\d+(?:\.\d+)?)K$", s)
    if m:
        return float(m.group(1)) / 1000.0
    m = re.search(r"(\d+)K(\d+)", s)
    if m:
        return float(f"{m.group(1)}.{m.group(2)}") / 1000.0
    try:
        return float(s)
    except Exception:
        return None


def ae_parse_power_w(value):
    s = str(value or "").strip().upper().replace(" ", "")
    if not s:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)MW", s)
    if m:
        return float(m.group(1)) / 1000.0
    m = re.search(r"W(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+(?:\.\d+)?)W", s)
    if m:
        return float(m.group(1))
    try:
        return float(s)
    except Exception:
        return None


def ae_parse_frequency_segments(value):
    raw = str(value or "").strip()
    if not raw:
        return []
    segments = re.split(r"\s*/\s*", raw)
    out = []
    for seg in segments:
        seg_raw = seg.strip()
        if not seg_raw:
            continue
        count = ae_parse_count(seg_raw)
        s = re.sub(r"\(\d+\)", "", seg_raw).strip().upper().replace(" ", "")
        s = re.split(r"[,;]", s)[0]
        multiplier = 1.0
        if s.startswith("G"):
            multiplier = 1000.0
            s = s[1:]
        elif s.startswith("M"):
            multiplier = 1.0
            s = s[1:]
        elif s.startswith("K"):
            multiplier = 0.001
            s = s[1:]
        elif "GHZ" in s:
            multiplier = 1000.0
        elif "KHZ" in s:
            multiplier = 0.001
        s = s.replace("GHZ", "").replace("MHZ", "").replace("KHZ", "")
        if not re.search(r"\d", s):
            out.append({"raw": seg_raw, "start": None, "end": None, "center": None, "count": count, "status": "Needs Review - no numeric frequency"})
            continue
        if "-" in s:
            try:
                a, b = s.split("-", 1)
                a = float(re.sub(r"[^0-9.]", "", a)) * multiplier
                b = float(re.sub(r"[^0-9.]", "", b)) * multiplier
                if b < a:
                    a, b = b, a
                out.append({"raw": seg_raw, "start": a, "end": b, "center": (a + b) / 2, "count": count, "status": "Parsed"})
            except Exception:
                out.append({"raw": seg_raw, "start": None, "end": None, "center": None, "count": count, "status": "Needs Review - range parse failed"})
        else:
            try:
                val = float(re.sub(r"[^0-9.]", "", s)) * multiplier
                out.append({"raw": seg_raw, "start": val, "end": val, "center": val, "count": count, "status": "Parsed"})
            except Exception:
                out.append({"raw": seg_raw, "start": None, "end": None, "center": None, "count": count, "status": "Needs Review - frequency parse failed"})
    return out


def ae_normalize_area(value):
    s = str(value or "").strip()
    if not s:
        return "Needs Review"
    up = s.upper()
    if "NOT REQUIRED" in up:
        return "Not Required"
    if "CANTON" in up:
        return "Cantonment"
    if "ALL" in up:
        return "North/Central/South"
    found = [a for a in ["North", "Central", "South"] if a.upper() in up]
    return "/".join(found) if found else s


def ae_band_label_for(center):
    if center is None:
        return "Other / Needs Review"
    for label, lo, hi in ALLOCATION_BANDS:
        if lo <= center < hi:
            return label
    return "Other / Needs Review"


def ae_derive_tech(tech_name, emitter, system, category):
    for v in [tech_name, system, emitter, category]:
        if v not in [None, ""] and str(v).strip().lower() not in ["none", "nan", "n/a"]:
            return str(v).strip()
    return "Needs Review"


def ae_priority(unit, sponsor, tech, equipment, category, system, description):
    text_blob = " ".join(str(x or "") for x in [unit, sponsor, tech, equipment, category, system, description]).upper()
    if any(k in text_blob for k in ["AVIATION", "AIR", "UAS", "APACHE", "UH-60", "RCAF", "A2GC", "SAFETY"]):
        return "Aviation / Safety of Life", 1
    if any(k in text_blob for k in ["MISSION COMMAND", "C2", "COMMAND", "NGC2", "TOC"]):
        return "Mission Command", 2
    if any(k in text_blob for k in ["EXPERIMENTAL", "DEVCOM", "PM CSS", "JMC", "FCD", "VOLCANO", "RED WOLF"]):
        return "Experimental", 3
    if any(k in text_blob for k in ["TACTICAL", "TSM", "MPU", "SILVUS", "TRILOS", "PRC", "NETWORK"]):
        return "Tactical Networks", 4
    if any(k in text_blob for k in ["CANADA", "AUSTRALIA", "UNITED KINGDOM", "NEW ZEALAND", "COALITION"]):
        return "Coalition / Partner", 5
    return "Tactical Networks", 4


def ae_area_tokens(area):
    area = str(area or "")
    tokens = []
    for token in ["North", "Central", "South"]:
        if token.upper() in area.upper():
            tokens.append(token)
    if not tokens:
        tokens = [area or "Needs Review"]
    return tokens


def ae_build_approved_pool(approved_file):
    approved_file.seek(0)
    df = pd.read_excel(approved_file, sheet_name=0)
    # Flexible column detection
    colmap = {ae_norm_key(c): c for c in df.columns}
    center_col = colmap.get("centerfrequencymhz") or colmap.get("centerfrequency") or colmap.get("frequency") or colmap.get("frequencymhz")
    bw_col = colmap.get("bandwidthmhz") or colmap.get("bandwidth")
    eq_col = colmap.get("equipment") or colmap.get("approvedequipment") or colmap.get("tech")
    notes_col = colmap.get("notes")
    if not center_col:
        raise ValueError("Approved Frequencies file must contain a center frequency column.")
    pool = []
    for _, r in df.iterrows():
        try:
            center = float(r.get(center_col))
        except Exception:
            continue
        bw = None
        if bw_col:
            try:
                bw = float(r.get(bw_col))
            except Exception:
                bw = None
        eq = str(r.get(eq_col)) if eq_col else ""
        notes = str(r.get(notes_col)) if notes_col else ""
        pool.append({
            "center": center,
            "bw": bw,
            "equipment": "" if eq in ["nan", "None"] else eq,
            "notes": "" if notes in ["nan", "None"] else notes,
            "band": ae_band_label_for(center),
        })
    return pool


def ae_equipment_score(request_equipment, request_tech, approved_equipment):
    req = f"{request_equipment or ''} {request_tech or ''}".upper()
    appr = str(approved_equipment or "").upper()
    if not appr:
        return 0
    score = 0
    for key in ["SILVUS", "MPU", "TRILOS", "PRC", "SATCOM", "STARLINK", "UAS", "LMR", "RED WOLF", "SKYDIO"]:
        if key in req and key in appr:
            score += 30
    if any(word and word in appr for word in re.split(r"[^A-Z0-9]+", req) if len(word) > 3):
        score += 5
    return score


def ae_pick_approved_frequency(pool, req_center, req_bw, band, ntc_area, equipment, tech, used_by_area):
    candidates = []
    req_areas = ae_area_tokens(ntc_area)
    for cand in pool:
        if cand["band"] != band and band != "Other / Needs Review":
            continue
        center = cand["center"]
        bw = cand["bw"]
        # Do not reuse same approved center in the same area.
        same_area_used = False
        for area in req_areas:
            if (round(center, 6), area) in used_by_area:
                same_area_used = True
                break
        if same_area_used:
            continue
        bw_delta = 9999 if req_bw is None or bw is None else abs(float(bw) - float(req_bw))
        freq_delta = 9999 if req_center is None else abs(float(center) - float(req_center))
        equip_score = ae_equipment_score(equipment, tech, cand.get("equipment"))
        total = (bw_delta * 100) + freq_delta - equip_score
        candidates.append((total, bw_delta, freq_delta, equip_score, cand))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][4]


def ae_detect_conflicts(rows):
    conflicts = []
    for i in range(len(rows)):
        a = rows[i]
        for j in range(i + 1, len(rows)):
            b = rows[j]
            try:
                a_start = float(a.get("Start Frequency (MHz)"))
                a_end = float(a.get("End Frequency (MHz)"))
                b_start = float(b.get("Start Frequency (MHz)"))
                b_end = float(b.get("End Frequency (MHz)"))
            except Exception:
                continue
            a_areas = set(ae_area_tokens(a.get("NTC Area")))
            b_areas = set(ae_area_tokens(b.get("NTC Area")))
            if not a_areas.intersection(b_areas):
                continue
            overlap = max(0, min(a_end, b_end) - max(a_start, b_start))
            if overlap > 0:
                conflicts.append({
                    "Conflict Type": "Frequency Overlap / Same NTC Area",
                    "Allocation A": a.get("Allocation ID"),
                    "Allocation B": b.get("Allocation ID"),
                    "Unit A": a.get("Unit"),
                    "Unit B": b.get("Unit"),
                    "NTC Area": ", ".join(sorted(a_areas.intersection(b_areas))),
                    "Overlap MHz": round(overlap, 6),
                    "Freq A": a.get("Center Frequency (MHz)"),
                    "Freq B": b.get("Center Frequency (MHz)"),
                    "Recommendation": "Move lower priority system or time-deconflict if approved.",
                })
    return conflicts


def build_allocation_workbook_from_request_and_pool(request_file, approved_file):
    approved_pool = ae_build_approved_pool(approved_file)
    request_file.seek(0)
    req_xl = pd.ExcelFile(request_file)
    if "PCC6 NTC" in req_xl.sheet_names:
        df = pd.read_excel(request_file, sheet_name="PCC6 NTC")
    else:
        df = pd.read_excel(request_file, sheet_name=0)
    colmap = {ae_norm_key(c): c for c in df.columns}
    unit_col = colmap.get("unit")
    sponsor_col = colmap.get("sponsor")
    loc_col = colmap.get("locationnorthcentralsouthcantonement") or colmap.get("ntcarea") or colmap.get("location")
    freq_col = colmap.get("frequencymhz") or colmap.get("requestedfrequency")
    bw_col = colmap.get("bandwidth") or colmap.get("bandwidthmhz")
    pwr_col = colmap.get("pwrwatts") or colmap.get("powerw")
    emitter_col = colmap.get("emitters") or colmap.get("equipment")
    tech_col = colmap.get("techname") or colmap.get("tech")
    cat_col = colmap.get("experimentalnetworkngc2experimentaltechtsmoemitterlegacytech") or colmap.get("techcategory")
    system_col = colmap.get("systemsplatforms") or colmap.get("systemplatform")
    desc_col = colmap.get("description")
    status_col = colmap.get("status")
    alt_col = colmap.get("altitude")
    band_col = colmap.get("bandlscxg") or colmap.get("requestband")
    poc_col = colmap.get("poc")
    source_sheet_name = "PCC6 NTC" if "PCC6 NTC" in req_xl.sheet_names else req_xl.sheet_names[0]
    used_by_area = set()
    rows = []
    needs_review = []
    alloc_id = 1
    for idx, r in df.iterrows():
        freq_raw = r.get(freq_col) if freq_col else None
        if pd.isna(freq_raw) or str(freq_raw).strip() == "":
            continue
        unit = "" if not unit_col or pd.isna(r.get(unit_col)) else str(r.get(unit_col)).strip()
        sponsor = "" if not sponsor_col or pd.isna(r.get(sponsor_col)) else str(r.get(sponsor_col)).strip()
        area = ae_normalize_area(r.get(loc_col) if loc_col else "")
        bw_mhz = ae_parse_bandwidth_mhz(r.get(bw_col) if bw_col else None)
        power_w = ae_parse_power_w(r.get(pwr_col) if pwr_col else None)
        equipment = "" if not emitter_col or pd.isna(r.get(emitter_col)) else str(r.get(emitter_col)).strip()
        system = "" if not system_col or pd.isna(r.get(system_col)) else str(r.get(system_col)).strip()
        tech = ae_derive_tech(r.get(tech_col) if tech_col else None, equipment, system, r.get(cat_col) if cat_col else None)
        category = "" if not cat_col or pd.isna(r.get(cat_col)) else str(r.get(cat_col)).strip()
        desc = "" if not desc_col or pd.isna(r.get(desc_col)) else str(r.get(desc_col)).strip()
        status = "" if not status_col or pd.isna(r.get(status_col)) else str(r.get(status_col)).strip()
        altitude = "" if not alt_col or pd.isna(r.get(alt_col)) else str(r.get(alt_col)).strip()
        request_band = "" if not band_col or pd.isna(r.get(band_col)) else str(r.get(band_col)).strip()
        poc = "" if not poc_col or pd.isna(r.get(poc_col)) else str(r.get(poc_col)).strip()
        priority, priority_score = ae_priority(unit, sponsor, tech, equipment, category, system, desc)
        segments = ae_parse_frequency_segments(freq_raw)
        if not segments:
            needs_review.append([source_sheet_name, idx + 2, unit, sponsor, area, freq_raw, "Frequency not parseable"])
            continue
        for seg in segments:
            if seg["start"] is None:
                needs_review.append([source_sheet_name, idx + 2, unit, sponsor, area, freq_raw, seg["status"]])
                continue
            requested_center = seg["center"]
            band = ae_band_label_for(requested_center)
            pick = ae_pick_approved_frequency(approved_pool, requested_center, bw_mhz, band, area, equipment, tech, used_by_area)
            if pick:
                center = float(pick["center"])
                approved_bw = pick.get("bw")
                approved_equipment = pick.get("equipment", "")
                source_note = "Approved frequency pool"
                for area_token in ae_area_tokens(area):
                    used_by_area.add((round(center, 6), area_token))
            else:
                center = requested_center
                approved_bw = None
                approved_equipment = ""
                source_note = "No approved frequency available - used requested center for review"
            use_bw = bw_mhz if bw_mhz and bw_mhz > 0 else approved_bw
            if use_bw and use_bw > 0:
                start_freq = center - float(use_bw) / 2
                end_freq = center + float(use_bw) / 2
            else:
                start_freq = seg["start"]
                end_freq = seg["end"]
            conflict_status = "Unchecked"
            allocation_status = "Proposed"
            locked = "Yes" if priority_score <= 2 else "No"
            notes = []
            if not pick:
                notes.append("Needs approved frequency review")
            if bw_mhz is None:
                notes.append("Bandwidth unclear")
            if area == "Needs Review":
                notes.append("NTC Area missing")
            if seg.get("count", 1) > 20:
                notes.append("High channel count - review actual nets")
            ism_notes = "; ".join(notes) if notes else "Auto-assigned using approved pool and NTC area reuse"
            row = {
                "Active": True,
                "Start Time": "0600",
                "End Time": "2000",
                "Equipment": equipment,
                "Center Frequency (MHz)": round(center, 6),
                "Start Frequency (MHz)": round(start_freq, 6),
                "End Frequency (MHz)": round(end_freq, 6),
                "Bandwidth (MHz)": round(use_bw, 6) if use_bw is not None else None,
                "Power (W)": power_w,
                "Tech": tech,
                "Unit": unit,
                "Sponsor": sponsor,
                "Latitude": "",
                "Longitude": "",
                "Location": "NTC Ft Irwin",
                "Antenna Height": altitude,
                "Coverage Radius": "",
                "Site Name": "",
                "Notes": desc,
                "NTC Area": area,
                "Grid / MGRS": "",
                "Allocation Status": allocation_status,
                "Conflict Status": conflict_status,
                "Reuse Group ID": f"{ae_band_label_for(center).replace(' ', '').replace('/', '-')}-{area.replace('/', '-')}-{int(round(center))}",
                "Frequency Locked": locked,
                "Priority": priority,
                "Priority Score": priority_score,
                "Source Sheet": source_sheet_name,
                "Source Row": idx + 2,
                "Requested Frequency": str(freq_raw),
                "Approved Equipment": approved_equipment,
                "Approved Frequency Source": source_note,
                "ISM Notes": ism_notes,
                "Allocation ID": f"ALLOC-{alloc_id:05d}",
                "Request Band": request_band,
                "Channels Requested": seg.get("count", 1),
                "Request Status": status,
                "POC": poc,
            }
            rows.append(row)
            if notes:
                needs_review.append([source_sheet_name, idx + 2, unit, sponsor, area, freq_raw, ism_notes])
            alloc_id += 1
    conflicts = ae_detect_conflicts(rows)
    conflict_ids = set()
    for c in conflicts:
        conflict_ids.add(c["Allocation A"])
        conflict_ids.add(c["Allocation B"])
    for row in rows:
        row["Conflict Status"] = "Conflict" if row["Allocation ID"] in conflict_ids else "No Conflict"
    return ae_export_allocation_workbook(rows, needs_review, conflicts)


def ae_export_allocation_workbook(rows, needs_review, conflicts):
    wb = XLWorkbook()
    default = wb.active
    wb.remove(default)

    base_cols = ["Active"] + [c for c in ALLOCATION_ENGINE_COLUMNS if c != "Active"] + ALLOCATION_ENGINE_APPEND_COLUMNS
    # Keep existing format first. App/helper-only fields stay far right.
    extra_cols = ["Allocation ID", "Request Band", "Channels Requested", "Request Status", "POC"]

    def add_sheet(name, data_rows, fill="1F4E78"):
        ws = wb.create_sheet(title=str(name)[:31])
        headers = base_cols + extra_cols
        ws.append(headers)
        for row in data_rows:
            ws.append([row.get(c, "") for c in headers])
        header_fill = PatternFill("solid", fgColor=fill)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"
        widths = {
            "A": 12, "B": 12, "C": 24, "D": 18, "E": 18, "F": 18, "G": 15, "H": 10,
            "I": 24, "J": 18, "K": 18, "N": 15, "O": 16, "R": 35, "S": 18, "T": 16,
            "U": 18, "V": 18, "W": 18, "X": 14, "Y": 22, "Z": 12, "AA": 18, "AB": 10,
            "AC": 16, "AD": 10, "AE": 24, "AF": 24, "AG": 40, "AH": 16, "AI": 12,
            "AJ": 10, "AK": 16, "AL": 25,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if len(data_rows) > 0:
            end_col = get_column_letter(len(headers))
            tab = Table(displayName=re.sub(r"[^A-Za-z0-9]", "", str(name))[:20] + "Tbl", ref=f"A1:{end_col}{len(data_rows)+1}")
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            try:
                ws.add_table(tab)
            except Exception:
                pass
        return ws

    # Dashboard
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "PCC6 Allocation Engine Dashboard"
    dash["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    dash["A1"].fill = PatternFill("solid", fgColor="0F172A")
    dash.merge_cells("A1:H1")
    metrics = [
        ["Metric", "Value"],
        ["Allocation Rows", len(rows)],
        ["Needs Review Rows", len(needs_review)],
        ["Conflict Rows", len(conflicts)],
        ["Unique Units", len(set(str(r.get("Unit")) for r in rows if str(r.get("Unit")).strip()))],
        ["Unique Sponsors", len(set(str(r.get("Sponsor")) for r in rows if str(r.get("Sponsor")).strip()))],
        ["Location", "NTC Ft Irwin"],
        ["Allocation Philosophy", "Option B + North/Central/South reuse pools"],
    ]
    for r_idx, vals in enumerate(metrics, 3):
        for c_idx, val in enumerate(vals, 1):
            dash.cell(r_idx, c_idx).value = val
    for cell in dash[3]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    dash.column_dimensions["A"].width = 28
    dash.column_dimensions["B"].width = 42

    add_sheet("Master Allocation", rows, "1F4E78")
    for label, _, _ in ALLOCATION_BANDS:
        band_rows = [r for r in rows if ae_band_label_for(r.get("Center Frequency (MHz)")) == label]
        if band_rows:
            add_sheet(label.replace("/", "-"), band_rows, "334155")
    for area in ["North", "Central", "South", "Cantonment", "Not Required", "Needs Review"]:
        if area in ["North", "Central", "South"]:
            area_rows = [r for r in rows if area in str(r.get("NTC Area"))]
        else:
            area_rows = [r for r in rows if str(r.get("NTC Area")) == area]
        if area_rows:
            add_sheet(f"NTC {area}", area_rows, "0F766E")

    # Needs Review
    review_ws = wb.create_sheet("Needs Review")
    review_headers = ["Source Sheet", "Source Row", "Unit", "Sponsor", "NTC Area", "Requested Frequency", "Issue"]
    review_ws.append(review_headers)
    for r in needs_review:
        review_ws.append(r)
    for cell in review_ws[1]:
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.font = Font(bold=True, color="FFFFFF")
    review_ws.freeze_panes = "A2"
    for col in range(1, 8):
        review_ws.column_dimensions[get_column_letter(col)].width = 24

    # Conflict Report
    conf_ws = wb.create_sheet("Conflict Report")
    conf_headers = ["Conflict Type", "Allocation A", "Allocation B", "Unit A", "Unit B", "NTC Area", "Overlap MHz", "Freq A", "Freq B", "Recommendation"]
    conf_ws.append(conf_headers)
    for c in conflicts:
        conf_ws.append([c.get(h, "") for h in conf_headers])
    for cell in conf_ws[1]:
        cell.fill = PatternFill("solid", fgColor="991B1B")
        cell.font = Font(bold=True, color="FFFFFF")
    conf_ws.freeze_panes = "A2"
    for col in range(1, len(conf_headers)+1):
        conf_ws.column_dimensions[get_column_letter(col)].width = 24

    # Frequency Reuse Matrix
    reuse = defaultdict(lambda: {"count": 0, "areas": set(), "units": set(), "sponsors": set()})
    for r in rows:
        key = (r.get("Reuse Group ID"), r.get("Center Frequency (MHz)"), r.get("Bandwidth (MHz)"), r.get("Tech"))
        reuse[key]["count"] += 1
        reuse[key]["areas"].add(str(r.get("NTC Area")))
        reuse[key]["units"].add(str(r.get("Unit")))
        reuse[key]["sponsors"].add(str(r.get("Sponsor")))
    reuse_ws = wb.create_sheet("Frequency Reuse Matrix")
    reuse_headers = ["Reuse Group ID", "Center Frequency (MHz)", "Bandwidth (MHz)", "Tech", "Reuse Count", "NTC Areas", "Units", "Sponsors", "Reuse Risk"]
    reuse_ws.append(reuse_headers)
    for (gid, center, bw, tech), val in sorted(reuse.items(), key=lambda x: str(x[0])):
        if val["count"] > 1:
            risk = "Low" if len(val["areas"]) > 1 else "High"
            reuse_ws.append([gid, center, bw, tech, val["count"], ", ".join(sorted(val["areas"])), ", ".join(sorted(val["units"]))[:250], ", ".join(sorted(val["sponsors"]))[:250], risk])
    for cell in reuse_ws[1]:
        cell.fill = PatternFill("solid", fgColor="7C2D12")
        cell.font = Font(bold=True, color="FFFFFF")
    reuse_ws.freeze_panes = "A2"
    for col in range(1, len(reuse_headers)+1):
        reuse_ws.column_dimensions[get_column_letter(col)].width = 24

    # Assumptions
    assump = wb.create_sheet("Assumptions")
    assump.append(["Item", "Method"])
    for row in [
        ["Workbook format", "Preserves user-provided allocation format; new columns are appended to the right."],
        ["Inputs", "Request Tracker + Approved Frequencies"],
        ["Priority", "Aviation/Safety, Mission Command, Experimental, Tactical Networks, Coalition/Partner"],
        ["Reuse", "North/Central/South reusable spectrum pools"],
        ["Conflict detection", "Bandwidth overlap inside same NTC Area"],
        ["Frequency Locked", "Yes for Aviation/Safety and Mission Command"],
        ["Needs Review", "Missing bandwidth, missing area, no approved frequency match, high channel count"],
    ]:
        assump.append(row)
    for cell in assump[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    assump.column_dimensions["A"].width = 28
    assump.column_dimensions["B"].width = 80

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------- Plotting ----------------
def style_axes(ax, dark=False):
    if dark:
        ax.set_facecolor("black"); ax.figure.patch.set_facecolor("black")
        ax.tick_params(colors="white"); ax.xaxis.label.set_color("white"); ax.yaxis.label.set_color("white"); ax.title.set_color("white")
        ax.grid(True, linestyle="--", linewidth=.4, color="gray", alpha=.5)
        for sp in ax.spines.values(): sp.set_color("white")
    else:
        ax.grid(True, linestyle="--", linewidth=.4, color="gray", alpha=.4)


def add_group_legend(ax, palette, title, dark=False, max_items=30):
    """Add a right-side legend to matplotlib plots."""
    if not palette:
        return
    items = list(palette.items())[:max_items]
    handles = [
        Rectangle((0, 0), 1, 1, facecolor=col, edgecolor=("white" if dark else "black"), alpha=0.75)
        for _, col in items
    ]
    labels = [str(name) for name, _ in items]
    leg = ax.legend(
        handles,
        labels,
        title=title,
        loc="center left",
        bbox_to_anchor=(1.01, 0.5),
        frameon=True,
        fontsize=8,
        title_fontsize=9,
        borderaxespad=0.0,
    )
    if dark:
        leg.get_frame().set_facecolor("black")
        leg.get_frame().set_edgecolor("white")
        leg.get_title().set_color("white")
        for txt in leg.get_texts():
            txt.set_color("white")

def thin_labels_df(df, group_field, min_gap_mhz):
    out = df.copy()
    if "CenterF" not in out.columns or min_gap_mhz is None or min_gap_mhz <= 0:
        out[".__keep_label"] = True; return out
    out[".__keep_label"] = False
    tmp = out.sort_values([group_field, "CenterF"], na_position="last")
    last, keep = {}, []
    for idx, row in tmp.iterrows():
        grp, cf = str(row[group_field]), row["CenterF"]
        if pd.isna(cf): continue
        if grp not in last or (cf - last[grp]) >= min_gap_mhz:
            keep.append(idx); last[grp] = cf
    out.loc[keep, ".__keep_label"] = True
    return out


def color_to_rgba(color, alpha=180):
    """Convert a matplotlib color to pydeck RGBA."""
    try:
        rgba = plt.matplotlib.colors.to_rgba(color)
        return [int(rgba[0] * 255), int(rgba[1] * 255), int(rgba[2] * 255), int(alpha)]
    except Exception:
        return [31, 119, 180, int(alpha)]


def prep_map_df(df, group_field):
    """Prepare valid map rows from decimal-degree Latitude/Longitude."""
    if "Latitude" not in df.columns or "Longitude" not in df.columns:
        return pd.DataFrame()

    m = df.copy()
    m["Latitude"] = pd.to_numeric(m["Latitude"], errors="coerce")
    m["Longitude"] = pd.to_numeric(m["Longitude"], errors="coerce")
    m = m[
        m["Latitude"].between(-90, 90, inclusive="both")
        & m["Longitude"].between(-180, 180, inclusive="both")
    ].copy()

    if m.empty:
        return m

    for col in ["Location", "SiteName"]:
        if col not in m.columns:
            m[col] = ""
    for col in ["CoverageRadius", "AntennaHeight"]:
        if col not in m.columns:
            m[col] = np.nan

    m["Group"] = m[group_field].astype(str) if group_field in m.columns else m["Equipment"].astype(str)
    return m


def build_map_deck(df, group_field, palette, radius_units="miles", show_coverage=True, map_style="light", show_heatmap=False, heatmap_weight_by="Power"):
    """Build an interactive map with black-outlined points and optional coverage circles."""
    m = prep_map_df(df, group_field)
    if m.empty:
        return None, m

    m["PowerW"] = pd.to_numeric(m.get("PowerW", 1), errors="coerce").fillna(1)
    m["CoverageRadius"] = pd.to_numeric(m.get("CoverageRadius", np.nan), errors="coerce")

    if radius_units == "kilometers":
        m["coverage_m"] = m["CoverageRadius"] * 1000.0
    elif radius_units == "nautical miles":
        m["coverage_m"] = m["CoverageRadius"] * 1852.0
    else:
        m["coverage_m"] = m["CoverageRadius"] * 1609.344

    m["coverage_m"] = m["coverage_m"].fillna(0).clip(lower=0)
    m["color"] = m["Group"].map(lambda g: color_to_rgba(palette.get(str(g), "#1f77b4"), 220))
    m["fill_color"] = m["Group"].map(lambda g: color_to_rgba(palette.get(str(g), "#1f77b4"), 55))
    m["point_radius"] = (m["PowerW"].clip(lower=0.1) ** 0.5 * 55).clip(lower=55, upper=350)

    if heatmap_weight_by == "Coverage Radius":
        m["heat_weight"] = m["coverage_m"].replace(0, np.nan).fillna(1).clip(lower=1)
    elif heatmap_weight_by == "Equal":
        m["heat_weight"] = 1
    else:
        m["heat_weight"] = m["PowerW"].clip(lower=0.1).fillna(1)

    layers = []

    if show_heatmap:
        layers.append(
            pdk.Layer(
                "HeatmapLayer",
                data=m,
                get_position="[Longitude, Latitude]",
                get_weight="heat_weight",
                radius_pixels=90,
                intensity=1,
                threshold=0.05,
                pickable=False,
            )
        )

    if show_coverage and (m["coverage_m"] > 0).any():
        layers.append(
            pdk.Layer(
                "ScatterplotLayer",
                data=m[m["coverage_m"] > 0],
                get_position="[Longitude, Latitude]",
                get_radius="coverage_m",
                get_fill_color="fill_color",
                get_line_color=[0, 0, 0, 170],
                line_width_min_pixels=1,
                stroked=True,
                filled=True,
                pickable=True,
            )
        )

    layers.append(
        pdk.Layer(
            "ScatterplotLayer",
            data=m,
            get_position="[Longitude, Latitude]",
            get_radius="point_radius",
            get_fill_color="color",
            get_line_color=[0, 0, 0, 255],
            line_width_min_pixels=1.5,
            stroked=True,
            filled=True,
            pickable=True,
        )
    )

    # Use free CARTO basemap styles so the map works without a Mapbox/Google/Bing token.
    # The previous Mapbox styles can render blank on Streamlit Cloud if no Mapbox token is configured.
    if map_style == "dark":
        style = "https://basemaps.cartocdn.com/gl/dark-matter-gl-style/style.json"
    elif map_style == "satellite":
        # Free no-token fallback. True satellite requires a provider token, such as Mapbox, Google, Bing, or Esri.
        style = "https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"
    else:
        style = "https://basemaps.cartocdn.com/gl/positron-gl-style/style.json"

    deck = pdk.Deck(
        layers=layers,
        initial_view_state=pdk.ViewState(
            latitude=float(m["Latitude"].mean()),
            longitude=float(m["Longitude"].mean()),
            zoom=8 if len(m) > 1 else 11,
            pitch=0,
        ),
        map_style=style,
        tooltip={
            "html": "<b>{Equipment}</b><br/>Group: {Group}<br/>Site: {SiteName}<br/>Location: {Location}<br/>Unit: {Unit}<br/>Tech: {Tech}<br/>Freq: {CenterF} MHz<br/>Power: {PowerW} W<br/>Time: {StartTime} - {EndTime}<br/>Coverage: {CoverageRadius}",
            "style": {"backgroundColor": "white", "color": "black"},
        },
    )
    return deck, m


def build_power_plot(df, group_field, dark, alpha_val, tick_major, tick_minor, label_digits, palette, auto_thin, min_label_gap, high_power_top, power_style, outline_lwd, show_center_labels):
    d = df.copy().sort_values(["PowerW", "StartF", "EndF"], ascending=[True if high_power_top else False, True, True])
    has_center = "CenterF" in d.columns and d["CenterF"].notna().any()
    d = thin_labels_df(d, group_field, min_label_gap) if auto_thin and has_center else d.assign(**{".__keep_label": True})
    x_min, x_max = min(d["StartF"].min(), d["EndF"].min()), max(d["StartF"].max(), d["EndF"].max())
    if tick_minor:
        x_min = math.floor(x_min/tick_minor)*tick_minor; x_max = math.ceil(x_max/tick_minor)*tick_minor
    y_max = float(d["PowerW"].max()); y_pad = max(1.0, .12*y_max)
    fig, ax = plt.subplots(figsize=(14,7), dpi=150)
    for _, row in d.iterrows():
        col = palette.get(str(row[group_field]), "#1f77b4")
        x, w, h = row["StartF"], row["EndF"]-row["StartF"], row["PowerW"]
        if power_style == "outline":
            ax.add_patch(Rectangle((x,0), w,h, fill=False, edgecolor="black", linewidth=max(outline_lwd, 0.8)))
        elif power_style == "outline_fill":
            ax.add_patch(Rectangle((x,0), w,h, facecolor=col, edgecolor="black", alpha=alpha_val, linewidth=max(outline_lwd, 0.8)))
        else:
            ax.add_patch(Rectangle((x,0), w,h, facecolor=col, edgecolor="black", alpha=alpha_val, linewidth=0.8))
            ax.add_line(Line2D([row["StartF"], row["EndF"]], [h,h], color=col, linewidth=.9))
    if has_center and show_center_labels:
        label_color = "white" if dark else "black"
        for _, row in d.loc[d[".__keep_label"] & d["CenterF"].notna()].iterrows():
            ax.text(
                row["CenterF"],
                max(0.02, row["PowerW"] - max(0.12, row["PowerW"] * 0.05)),
                f"{row['CenterF']:.{label_digits}f} MHz",
                rotation=90,
                fontsize=8,
                ha="center",
                va="top",
                color=label_color,
                fontweight="bold",
                clip_on=True,
            )
    ax.set_xlim(x_min, x_max); ax.set_ylim(0, y_max+y_pad)
    ax.set_xlabel("Frequency (MHz)"); ax.set_ylabel("Power (W)")
    ax.set_title(f"Frequency Allocation vs Power — by {group_field}", fontweight="bold")
    if tick_major: ax.set_xticks(np.arange(x_min, x_max+tick_major, tick_major))
    if tick_minor:
        ax.set_xticks(np.arange(x_min, x_max+tick_minor, tick_minor), minor=True); ax.grid(which="minor", linestyle=":", linewidth=.25, alpha=.25)
    style_axes(ax, dark)
    add_group_legend(ax, palette, group_field, dark)
    plt.tight_layout(rect=[0, 0, 0.84, 1])
    return fig

def build_deconflict_plot(d0, grp_field, palette, dark, tick_major, tick_minor, show_box_labels, min_label_height_min, show_shift_label, moved_outline_thickness, conflicts_df):
    d = d0.copy(); d["Group"] = d[grp_field].astype(str)
    d["StartSec"] = d["StartTime"].apply(parse_time_one); d["EndSec"] = d["EndTime"].apply(parse_time_one)
    d = d.loc[d["StartSec"].notna() & d["EndSec"].notna()].copy()
    if d.empty:
        fig, ax = plt.subplots(figsize=(14, 4), dpi=150)
        ax.axis("off")
        ax.text(0.02, 0.70, "No time values could be parsed for this deconfliction plot.", fontsize=14, fontweight="bold", transform=ax.transAxes)
        ax.text(0.02, 0.52, "Check that your uploaded sheet has Start Time and End Time columns, or values like 6am, 08:00, 1300.", fontsize=11, transform=ax.transAxes)
        ax.text(0.02, 0.36, "Tip: re-upload the file after this update; the importer now auto-detects time columns.", fontsize=11, transform=ax.transAxes)
        return fig
    d["EndSecUnwrapped"] = np.where(d["EndSec"] < d["StartSec"], d["EndSec"]+86400, d["EndSec"])
    d["BoxHeightMin"] = (d["EndSecUnwrapped"]-d["StartSec"])/60
    d["Moved"] = d["ShiftSec"].fillna(0).abs() > 0 if "ShiftSec" in d.columns else False
    x_min, x_max = min(d["StartF"].min(), d["EndF"].min()), max(d["StartF"].max(), d["EndF"].max())
    if tick_minor:
        x_min = math.floor(x_min/tick_minor)*tick_minor; x_max = math.ceil(x_max/tick_minor)*tick_minor
    fig, ax = plt.subplots(figsize=(14,7), dpi=150)
    for _, row in d.iterrows():
        col = palette.get(str(row["Group"]), "#1f77b4")
        x0, w = row["StartF"], row["EndF"]-row["StartF"]
        y0, h = row["StartSec"]/3600, (row["EndSecUnwrapped"]-row["StartSec"])/3600
        ax.add_patch(Rectangle((x0,y0), w,h, facecolor=col, edgecolor="black", alpha=.95, linewidth=0.8))
        if moved_outline_thickness > 0 and bool(row["Moved"]):
            ax.add_patch(Rectangle((x0,y0), w,h, facecolor="none", edgecolor="black", linewidth=moved_outline_thickness))
        if show_box_labels and row["BoxHeightMin"] >= float(min_label_height_min):
            # V40: only show frequency/range inside boxes. Equipment/Tech/Unit/Sponsor names stay in the legend.
            center_f = pd.to_numeric(row.get("CenterF"), errors="coerce")
            start_f = pd.to_numeric(row.get("StartF"), errors="coerce")
            end_f = pd.to_numeric(row.get("EndF"), errors="coerce")
            if np.isfinite(center_f):
                label = f"{center_f:.3f} MHz"
            elif np.isfinite(start_f) and np.isfinite(end_f):
                label = f"{start_f:.3f}-{end_f:.3f} MHz"
            else:
                label = ""
            if label:
                ax.text(x0+w/2, y0+h/2, label, ha="center", va="center", fontsize=8, fontweight="bold", color=("white" if dark else "black"), clip_on=True)
        if show_shift_label and "ShiftSec" in d.columns and row["BoxHeightMin"] >= float(min_label_height_min):
            shift = row.get("ShiftSec", np.nan)
            if pd.notna(shift) and abs(shift) > 0:
                ax.text(x0+w/2, min(y0+h-.2, y0+.2), f"Δ {round(shift/60):.0f}m", ha="center", va="center", fontsize=7, fontweight="bold", color=("white" if dark else "black"), clip_on=True)
    if conflicts_df is not None and not conflicts_df.empty:
        for _, row in conflicts_df.iterrows():
            y1, y2 = (row["StartOverlap"]%86400)/3600, (row["EndOverlap"]%86400)/3600
            ax.add_patch(Rectangle((row["FreqLeft"], y1), row["FreqRight"]-row["FreqLeft"], y2-y1, facecolor="red", edgecolor="red", alpha=.15, linewidth=.8))
    ax.set_xlim(x_min, x_max); ax.set_ylim(0,24)
    ax.set_xlabel("Frequency (MHz)"); ax.set_ylabel("Time (hours)")
    ax.set_title(f"Time × Frequency — by {grp_field}", fontweight="bold")
    if tick_major: ax.set_xticks(np.arange(x_min, x_max+tick_major, tick_major))
    if tick_minor:
        ax.set_xticks(np.arange(x_min, x_max+tick_minor, tick_minor), minor=True); ax.grid(which="minor", linestyle=":", linewidth=.25, alpha=.25)
    style_axes(ax, dark)
    add_group_legend(ax, palette, grp_field, dark)
    plt.tight_layout(rect=[0, 0, 0.84, 1])
    return fig


logged_in_user, current_role = require_login()
current_user_id = st.session_state.get("auth_user_id", "")
is_admin = current_role == "admin"
can_edit = current_role in ["admin", "editor"]

# ---------------- Sidebar ----------------
with st.sidebar:
    st.markdown("## Workspace")
    user_name = logged_in_user
    st.success(f"Logged in as: {user_name}")
    st.caption(f"Role: {current_role}")
    if st.button("Log out", use_container_width=True):
        for k in ["auth_user", "auth_user_id", "auth_role", "auth_full_name"]:
            st.session_state.pop(k, None)
        try:
            sb.auth.sign_out()
        except Exception:
            pass
        st.rerun()
    projects = list_projects(); project_names = [p["name"] for p in projects]
    create_new = st.toggle("Create new project", value=False, disabled=not can_edit)
    if create_new:
        new_name = st.text_input("New project name"); new_desc = st.text_area("Description", height=80)
        if st.button("Create project", type="primary", use_container_width=True, disabled=not can_edit):
            if not new_name.strip(): st.error("Project name is required.")
            else:
                proj = create_project(new_name.strip(), new_desc.strip()); st.session_state["project_id"] = proj["id"]; st.rerun()
    else:
        if projects:
            selected_name = st.selectbox("Project", project_names, index=0)
            st.session_state["project_id"] = next(p for p in projects if p["name"] == selected_name)["id"]
        else: st.info("Create your first project.")

    if can_edit:
        with st.expander("Mission templates", expanded=False):
            templates = list_mission_templates()
            template_names = [t.get("name", "Unnamed") for t in templates]

            st.markdown("#### Create project from template")
            if templates:
                selected_template_name = st.selectbox("Template", template_names, key="template_select_create")
                selected_template = templates[template_names.index(selected_template_name)]
                new_template_project_name = st.text_input("New project name from template", key="template_new_project_name")
                new_template_project_desc = st.text_area("New project description", height=70, key="template_new_project_desc")
                if st.button("Create project from selected template", type="primary", use_container_width=True):
                    if not new_template_project_name.strip():
                        st.error("Project name is required.")
                    else:
                        try:
                            proj = create_project_from_template(selected_template, new_template_project_name.strip(), new_template_project_desc.strip(), logged_in_user)
                            st.session_state["project_id"] = proj["id"]
                            st.success("Project created from template.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not create project from template: {e}")
            else:
                st.info("No templates saved yet.")

            st.markdown("#### Save current project as template")
            template_name = st.text_input("Template name", key="template_save_name")
            template_desc = st.text_area("Template description", height=70, key="template_save_desc")
            if st.button("Save current workbook as template", use_container_width=True, disabled=not st.session_state.get("project_id")):
                try:
                    current_project_id_for_template = st.session_state.get("project_id")
                    sheets_for_template = st.session_state.get("workbook_sheets") or load_project_sheets(current_project_id_for_template)
                    active_template_sheet = st.session_state.get("active_sheet_name") or (list(sheets_for_template.keys())[0] if sheets_for_template else "Working")
                    template_id = save_mission_template(template_name, template_desc, sheets_for_template, active_template_sheet, logged_in_user)
                    log_audit_event(current_project_id_for_template, "mission_template_saved", logged_in_user, {"template": template_name})
                    st.success(f"Template saved: {template_name}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not save template: {e}")

            if is_admin and templates:
                st.markdown("#### Delete template")
                delete_template_name = st.selectbox("Template to delete", template_names, key="template_delete_select")
                delete_template = templates[template_names.index(delete_template_name)]
                confirm_delete_template = st.checkbox(f"I understand: delete template '{delete_template_name}'", key="confirm_delete_template")
                if st.button("Delete selected template", type="primary", use_container_width=True, disabled=not confirm_delete_template):
                    try:
                        delete_mission_template(delete_template["id"])
                        st.success("Template deleted.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not delete template: {e}")

    if is_admin and projects:
        with st.expander("Admin: Delete old projects", expanded=False):
            st.warning("Deleting a project removes its allocation rows, workbook tabs, version history, saved files, and the project record.")
            delete_options = [f"{p.get('name','Unnamed')} — {p.get('id','')}" for p in projects]
            delete_label = st.selectbox("Project to delete", delete_options, key="delete_project_select")
            delete_project_obj = projects[delete_options.index(delete_label)]
            confirm_delete_project = st.checkbox(
                f"I understand: permanently delete '{delete_project_obj.get('name','Unnamed')}'",
                key="confirm_delete_project",
            )
            if st.button(
                "Delete selected project",
                type="primary",
                use_container_width=True,
                disabled=not confirm_delete_project,
                key="delete_project_button",
            ):
                try:
                    delete_project(delete_project_obj["id"])
                    if st.session_state.get("project_id") == delete_project_obj["id"]:
                        st.session_state.pop("project_id", None)
                        st.session_state.pop("workbook_sheets", None)
                        st.session_state.pop("active_sheet_name", None)
                        st.session_state.pop("workbook_project_id", None)
                    st.success("Project deleted.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not delete project: {e}")

    st.divider(); st.header("Frequency visibility")
    show_inactive_rows = st.checkbox("Show inactive frequencies", value=False)
    inactive_action_mode = st.radio("Inactive row action", ["Hide from planning", "Show gray reference"], index=0)

    st.divider(); st.header("Plot controls")
    dark = st.checkbox("Dark theme", value=True)
    power_style = st.selectbox("Power plot style", ["outline_fill", "filled", "outline"], index=0, format_func=lambda x: {"outline_fill":"Outline + light fill", "filled":"Filled bands", "outline":"Outline only"}[x])
    alpha_val = st.slider("Fill transparency", 0.0, 1.0, 0.95, 0.05)
    high_power_top = st.checkbox("Draw HIGH power on top", value=False)
    outline_lwd = st.slider("Outline thickness", 0.2, 2.0, 0.6, 0.1)
    auto_thin = st.checkbox("Auto thin center-frequency labels", value=False)
    show_center_labels = st.checkbox("Show center-frequency labels inside boxes", value=True)
    label_gap = st.number_input("Min label gap (MHz)", min_value=0.0, value=2.0, step=0.5)
    label_digits = st.number_input("Label decimals", min_value=0, max_value=6, value=2, step=1)
    tick_major = ticks_value(st.text_input("Major tick (MHz)", value="")); tick_minor = ticks_value(st.text_input("Minor grid (MHz)", value=""))
    st.divider(); st.header("Auto deconflict")
    auto_dc = st.checkbox("Auto deconflict by time", value=False)
    dc_start = st.text_input(
        "Anchor",
        value=st.session_state.get("dc_anchor_time", "6am"),
        key="dc_anchor_time",
        help="Deconfliction will pack moved rows starting at this time. Examples: 4am, 0730, 13:00, 1:30pm.",
    )
    dc_window = st.selectbox("Max shift window (hours)", [2,4,6,8,10,12,16,20,24], index=3)
    dc_pad_min = st.number_input("Min separation (minutes)", min_value=0.0, value=1.0, step=0.5)
    guard_mhz = st.number_input("Frequency guard band (MHz)", min_value=0.0, value=0.0, step=0.1)
    allow_earlier = st.checkbox("Allow moving earlier into open gaps", value=True)
    priority_mode = st.selectbox("Scheduling priority", ["Power + Original Time", "Highest Power First", "Shortest Duration First"], index=0)

    st.divider()
    st.header("Smart planner")
    plan_band_start = st.text_input("Planning band start MHz", value="")
    plan_band_end = st.text_input("Planning band end MHz", value="")
    plan_guard_mhz = st.number_input("Planner guard band MHz", min_value=0.0, value=0.025, step=0.025)
    plan_step_mhz = st.number_input("Planner frequency step MHz", min_value=0.001, value=0.025, step=0.001)
    max_suggestions = st.number_input("Max suggestions per conflict", min_value=1, max_value=10, value=3, step=1)

    st.divider(); st.header("Box labels")
    box_labels = st.checkbox("Show label inside boxes", value=True)
    box_label_min_height_min = st.number_input("Min box height for label (min)", min_value=0.0, value=15.0, step=5.0)
    show_shift_label = st.checkbox("Show Δ minutes for moved boxes", value=False)
    moved_outline = st.slider("Moved outline thickness", 0.0, 2.5, 0.0, 0.1)

    st.divider()
    st.header("Map controls")
    map_group_by = st.selectbox("Map color by", ["Equipment", "Unit", "Tech"], index=0)
    radius_units = st.selectbox("Coverage radius units", ["miles", "kilometers", "nautical miles"], index=0)
    show_coverage = st.checkbox("Show coverage circles", value=True)
    show_heatmap = st.checkbox("Show heatmap / congestion layer", value=False)
    heatmap_weight_by = st.selectbox("Heatmap weight by", ["Power", "Coverage Radius", "Equal"], index=0)
    map_style_choice = st.selectbox("Map style", ["light", "dark", "satellite"], index=0)

    st.divider()
    st.header("Terrain / Range")
    default_ant_height_m = st.number_input("Default antenna height when blank (m)", min_value=1.0, value=10.0, step=1.0)
    show_range_warning = st.checkbox("Show LOS/range warnings", value=True)

    st.divider()
    st.header("Collaboration")
    auto_refresh = st.checkbox("Auto-refresh project", value=False)
    refresh_seconds = st.selectbox("Refresh interval", [10, 20, 30, 60], index=2)

project_id = st.session_state.get("project_id")
if not project_id:
    st.info("Create or select a project to begin."); st.stop()

current_project = next((p for p in projects if p["id"] == project_id), {})
project_name = current_project.get("name", "Selected")
if not user_has_project_access(project_id, current_user_id):
    st.error("You do not have access to this project. Ask an administrator to add you as a project member.")
    st.stop()

status_info = get_project_status(project_id)
project_status = status_info.get("status", "Draft")

# Approved projects are locked for editors/viewers. Admins can still make emergency changes.
project_locked = project_status == "Approved" and not is_admin
if project_locked:
    can_edit = False

# Collaboration presence and optional auto-refresh.
update_presence(project_id, current_user_id, logged_in_user, current_role)

if auto_refresh:
    st.caption(f"Auto-refresh enabled: every {refresh_seconds} seconds.")
    st.markdown(
        f"<meta http-equiv='refresh' content='{int(refresh_seconds)}'>",
        unsafe_allow_html=True,
    )

current_df = get_project_rows(project_id)
current_df_all = current_df.copy()
current_df = ensure_active_first_preserve_order(apply_active_filter(current_df, show_inactive_rows))

# Restore saved workbook sheets. If there are no saved sheets yet, use the shared allocation table as a single Working sheet.
saved_sheets = load_project_sheets(project_id)
if saved_sheets:
    if st.session_state.get("workbook_project_id") != project_id:
        st.session_state["workbook_sheets"] = saved_sheets
        st.session_state["active_sheet_name"] = list(saved_sheets.keys())[0]
        st.session_state["workbook_project_id"] = project_id
else:
    if st.session_state.get("workbook_project_id") != project_id or "workbook_sheets" not in st.session_state:
        st.session_state["workbook_sheets"] = {"Working": normalize_uploaded_df(current_df)}
        st.session_state["active_sheet_name"] = "Working"
        st.session_state["workbook_project_id"] = project_id

st.session_state.setdefault("pending_upload_df", None)
st.session_state.setdefault("pending_upload_sheets_dict", None)
st.session_state.setdefault("pending_upload_sheet", None)
st.session_state.setdefault("pending_upload_name", None)
st.session_state.setdefault("pending_upload_bytes", None)

# Workbook persistence status.
try:
    _saved_sheet_count = len(load_project_sheets(project_id))
except Exception:
    _saved_sheet_count = 0
if _saved_sheet_count == 0 and len(st.session_state.get("workbook_sheets", {})) > 1:
    st.warning("Workbook tabs are visible in this session, but they have not been confirmed saved in Supabase yet. Click Save shared changes, or run the v17 SQL setup if saving fails.")

cols = st.columns(4)
active_count = len(apply_active_filter(current_df_all, False)) if "Active" in current_df_all.columns else len(current_df_all)
inactive_count = len(current_df_all) - active_count if "Active" in current_df_all.columns else 0
cols[0].metric("Active Rows", active_count); cols[1].metric("Inactive", inactive_count); cols[2].metric("Status", project_status); cols[3].metric("User", user_name)
if st.button("Refresh latest project data", use_container_width=True):
    log_audit_event(project_id, "manual_refresh", logged_in_user, {})
    st.rerun()

with st.expander("Project dashboard", expanded=False):
    st.caption("At-a-glance project health, collaboration, and data status.")

    dash_cols = st.columns(5)
    dash_cols[0].metric("Active Rows", active_count)
    dash_cols[1].metric("Workbook tabs", len(st.session_state.get("workbook_sheets", {})))
    dash_cols[2].metric("Status", project_status)
    try:
        dash_cols[3].metric("Versions", len(list_versions(project_id)))
    except Exception:
        dash_cols[3].metric("Versions", "—")
    try:
        dash_cols[4].metric("Members", len(list_project_members(project_id)))
    except Exception:
        dash_cols[4].metric("Members", "—")

    dashboard_activity = list_recent_activity(project_id, limit=10)
    if not dashboard_activity.empty:
        st.markdown("#### Latest activity")
        st.dataframe(dashboard_activity, use_container_width=True)




with st.expander("Workflow: Approval status", expanded=False):
    st.write(f"Current status: **{project_status}**")
    if status_info.get("approved_by"):
        st.caption(f"Approved by {status_info.get('approved_by')} at {status_info.get('approved_at')}")
    if status_info.get("status_note"):
        st.caption(f"Note: {status_info.get('status_note')}")

    status_note = st.text_input("Status note", value="", key=f"status_note_{project_id}")

    c_submit, c_approve, c_reject, c_draft = st.columns(4)

    with c_submit:
        if st.button("Submit for review", use_container_width=True, disabled=not can_edit):
            set_project_status(project_id, "In Review", logged_in_user, status_note)
            st.success("Submitted for review.")
            st.rerun()

    with c_approve:
        if st.button("Approve", type="primary", use_container_width=True, disabled=not is_admin):
            set_project_status(project_id, "Approved", logged_in_user, status_note)
            st.success("Project approved and locked for non-admin editors.")
            st.rerun()

    with c_reject:
        if st.button("Reject", use_container_width=True, disabled=not is_admin):
            set_project_status(project_id, "Rejected", logged_in_user, status_note)
            st.warning("Project rejected.")
            st.rerun()

    with c_draft:
        if st.button("Reopen draft", use_container_width=True, disabled=not is_admin):
            set_project_status(project_id, "Draft", logged_in_user, status_note)
            st.success("Project reopened as draft.")
            st.rerun()




with st.expander("Mission template library", expanded=False):
    templates = list_mission_templates()
    if not templates:
        st.info("No mission templates saved yet.")
    else:
        template_df = pd.DataFrame(templates)
        show_cols = [c for c in ["name", "description", "created_by", "updated_by", "updated_at", "active_sheet"] if c in template_df.columns]
        st.dataframe(template_df[show_cols], use_container_width=True)


with st.expander("Live collaboration dashboard", expanded=False):
    st.caption("Shows recently active users and recent project activity. Turn on Auto-refresh in the sidebar for live updates.")

    c_presence, c_activity = st.columns(2)

    with c_presence:
        st.markdown("#### Online / recent users")
        presence_df = list_presence(project_id)
        if presence_df.empty:
            st.info("No presence records yet.")
        else:
            display_cols = [c for c in ["online", "email", "role", "minutes_ago", "last_seen"] if c in presence_df.columns]
            st.dataframe(presence_df[display_cols], use_container_width=True)

    with c_activity:
        st.markdown("#### Recent activity")
        activity_df = list_recent_activity(project_id, limit=25)
        if activity_df.empty:
            st.info("No recent activity.")
        else:
            st.dataframe(activity_df, use_container_width=True)

    last_update = get_last_project_update(project_id)
    if last_update:
        st.caption(f"Project last updated: {last_update.get('updated_at', 'unknown')}")



if is_admin:
    with st.expander("Admin: User management", expanded=False):
        st.caption("Admins can create users, change roles, disable accounts, or delete auth users. Deleting requires the Supabase service role key.")

        users = list_app_users()
        if users:
            st.dataframe(pd.DataFrame(users), use_container_width=True)
        else:
            st.info("No users are listed yet. The first authenticated user becomes admin automatically.")

        st.markdown("#### Create user")
        with st.form("admin_create_user_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                new_email = st.text_input("New user email")
                new_full_name = st.text_input("Full name")
            with c2:
                new_password = st.text_input("Temporary password", type="password")
                new_role = st.selectbox("Role", ["viewer", "editor", "admin"], index=0)
            submitted_create_user = st.form_submit_button("Create user", type="primary")

        if submitted_create_user:
            try:
                if not new_email.strip() or not new_password:
                    st.error("Email and temporary password are required.")
                elif len(new_password) < 8:
                    st.error("Temporary password must be at least 8 characters.")
                else:
                    uid = admin_create_user(new_email.strip(), new_password, new_role, new_full_name.strip(), logged_in_user)
                    st.success(f"Created user {new_email.strip()} as {new_role}.")
                    st.rerun()
            except Exception as e:
                st.error(f"Could not create user: {e}")

        st.markdown("#### Change role / disable / remove user")
        users = list_app_users()
        if users:
            user_labels = [f"{u.get('email','')} — {u.get('role','viewer')} — {u.get('user_id','')}" for u in users]
            selected_label = st.selectbox("Select user", user_labels)
            selected_user = users[user_labels.index(selected_label)]

            valid_roles = ["viewer", "editor", "admin", "disabled"]
            current_selected_role = selected_user.get("role", "viewer")
            if current_selected_role not in valid_roles:
                current_selected_role = "viewer"

            role_choice = st.selectbox(
                "New role",
                valid_roles,
                index=valid_roles.index(current_selected_role),
            )

            c1, c2 = st.columns(2)
            with c1:
                if st.button("Update selected user role", use_container_width=True):
                    try:
                        update_user_role(
                            selected_user["user_id"],
                            selected_user.get("email", ""),
                            role_choice,
                            selected_user.get("full_name", ""),
                            logged_in_user,
                        )
                        st.success("User role updated.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not update role: {e}")
            with c2:
                if st.button("Delete selected auth user", use_container_width=True):
                    try:
                        update_user_role(
                            selected_user["user_id"],
                            selected_user.get("email", ""),
                            "disabled",
                            selected_user.get("full_name", ""),
                            logged_in_user,
                        )
                        admin_delete_user(selected_user["user_id"])
                        st.success("User deleted from Supabase Auth and disabled in app roles.")
                        st.rerun()
                    except Exception as e:
                        st.warning(f"User was disabled in app roles, but auth delete failed or is unavailable: {e}")
                        st.rerun()




if is_admin:
    with st.expander("Admin: Project access", expanded=False):
        st.caption("Control which users can open the selected project. Admins can always see every project.")

        members = list_project_members(project_id)
        if members:
            st.dataframe(pd.DataFrame(members), use_container_width=True)
        else:
            st.info("No explicit members yet. Add users below.")

        app_users = list_app_users()
        if not app_users:
            st.warning("No app users found yet.")
        else:
            user_options = [
                f"{u.get('email','')} — {u.get('role','viewer')} — {u.get('user_id','')}"
                for u in app_users
                if u.get("role") != "disabled"
            ]

            if user_options:
                selected_user_label = st.selectbox("User to add/update", user_options, key=f"member_user_{project_id}")
                selected_user_obj = app_users[user_options.index(selected_user_label)]

                access_role = st.selectbox(
                    "Project access role",
                    ["viewer", "editor", "owner"],
                    index=1,
                    key=f"member_access_{project_id}",
                )

                c_add, c_remove = st.columns(2)

                with c_add:
                    if st.button("Add / update project member", use_container_width=True, key=f"add_member_{project_id}"):
                        try:
                            upsert_project_member(
                                project_id,
                                selected_user_obj["user_id"],
                                selected_user_obj.get("email", ""),
                                access_role,
                                logged_in_user,
                            )
                            st.success("Project access updated.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not update project access: {e}")

                with c_remove:
                    if st.button("Remove selected user from project", use_container_width=True, key=f"remove_member_{project_id}"):
                        try:
                            remove_project_member(project_id, selected_user_obj["user_id"], logged_in_user)
                            st.success("Project member removed.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not remove project member: {e}")



if is_admin:
    with st.expander("Admin: Project backup / restore", expanded=False):
        st.caption("Export the complete selected project as JSON, or restore a backup into a new project.")

        b1, b2 = st.columns(2)

        with b1:
            st.markdown("#### Backup current project")
            st.download_button(
                "Download full project backup JSON",
                data=export_project_backup(project_id),
                file_name=f"{safe_storage_filename(project_name)}_project_backup.json",
                mime="application/json",
                use_container_width=True,
            )

        with b2:
            st.markdown("#### Restore backup as new project")
            backup_upload = st.file_uploader("Upload project backup JSON", type=["json"], key="restore_backup_json")
            restore_name = st.text_input("Restored project name", value=f"Restored - {project_name}", key="restore_project_name")
            confirm_restore = st.checkbox("I understand: restore backup into a new project", key="confirm_restore_backup")

            if st.button("Restore backup", type="primary", use_container_width=True, disabled=not (backup_upload and confirm_restore)):
                try:
                    restored_project = restore_project_backup(backup_upload.getvalue(), restore_name, logged_in_user)
                    st.session_state["project_id"] = restored_project["id"]
                    st.success("Backup restored into a new project.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not restore backup: {e}")


if is_admin:
    with st.expander("Admin: Delete data", expanded=False):
        st.warning("These actions affect the selected project only. They cannot be undone.")

        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown("#### Delete all version history")
            st.caption("Deletes every saved snapshot for this project. Current workbook/table rows stay untouched.")
            confirm_versions = st.checkbox(
                "I understand: delete all version history",
                key=f"confirm_delete_versions_{project_id}",
            )
            if st.button(
                "Delete all version history",
                type="primary",
                use_container_width=True,
                disabled=not confirm_versions,
                key=f"btn_delete_versions_{project_id}",
            ):
                try:
                    delete_all_versions(project_id)
                    st.success("All version history for this project was deleted.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not delete version history: {e}")

        with col_b:
            st.markdown("#### Clear shared allocation table")
            st.caption("Deletes current allocation rows and saved workbook tabs. Project, users, uploaded files, and versions remain.")
            confirm_table = st.checkbox(
                "I understand: clear shared allocation table",
                key=f"confirm_clear_table_{project_id}",
            )
            if st.button(
                "Clear shared allocation table",
                type="primary",
                use_container_width=True,
                disabled=not confirm_table,
                key=f"btn_clear_table_{project_id}",
            ):
                try:
                    clear_shared_allocation_table(project_id, logged_in_user)
                    st.session_state.pop("editor", None)
                    st.session_state.pop("workbook_sheets", None)
                    st.session_state.pop("active_sheet_name", None)
                    st.session_state.pop("workbook_project_id", None)
                    st.success("Shared allocation table and workbook tabs were cleared.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not clear shared allocation table: {e}")



with st.expander("Audit trail", expanded=False):
    audit_rows = list_audit_events(project_id)
    if not audit_rows:
        st.info("No audit events yet.")
    else:
        audit_df = pd.DataFrame(audit_rows)
        show_cols = [c for c in ["created_at", "event_type", "event_by", "details"] if c in audit_df.columns]
        st.dataframe(audit_df[show_cols], use_container_width=True)



with st.expander("Row-level edit history", expanded=False):
    history_rows = list_row_history(project_id)
    if not history_rows:
        st.info("No row history yet. Save changes to start tracking row-level edits.")
    else:
        hist_df = pd.DataFrame(history_rows)
        show_cols = [c for c in ["created_at", "row_order", "changed_by", "change_source", "changes"] if c in hist_df.columns]
        st.dataframe(hist_df[show_cols], use_container_width=True)

        if can_edit:
            st.markdown("#### Restore one row")
            labels = [
                f"{r.get('created_at','')} | row {r.get('row_order')} | {r.get('changed_by','')}"
                for r in history_rows
            ]
            selected_hist_label = st.selectbox("History record", labels, key=f"row_history_select_{project_id}")
            selected_hist = history_rows[labels.index(selected_hist_label)]
            confirm_restore_row = st.checkbox("I understand: restore this one row from its previous value", key=f"confirm_row_restore_{project_id}")
            if st.button("Restore selected row", type="primary", use_container_width=True, disabled=not confirm_restore_row):
                try:
                    restored_df = restore_row_from_history(project_id, selected_hist, logged_in_user)
                    st.session_state["workbook_sheets"] = {st.session_state.get("active_sheet_name", "Restored"): restored_df}
                    st.success("Selected row restored.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not restore row: {e}")



with st.expander("Active / inactive frequency control", expanded=False):
    st.caption("Turn off frequencies you are not using. Inactive rows remain saved but are excluded from plots, conflicts, smart planner, maps, and dashboards unless 'Show inactive frequencies' is enabled.")

    active_source_df = current_df_all.copy()
    if active_source_df.empty:
        st.info("No rows available.")
    else:
        if "Active" not in active_source_df.columns:
            active_source_df["Active"] = True

        active_source_df["Active"] = active_source_df["Active"].apply(to_active_bool)
        active_source_df["Status"] = active_source_df["Active"].apply(lambda x: "Active" if x else "Inactive")

        display_cols = ["Active"] + [c for c in active_source_df.columns if c != "Active"]
        edited_active_df = st.data_editor(
            active_source_df[display_cols],
            use_container_width=True,
            num_rows="fixed",
            disabled=[c for c in display_cols if c not in ["Active"]],
            key=f"active_editor_{project_id}",
        )

        c_on, c_off, c_save_active = st.columns(3)

        with c_on:
            if st.button("Set all active", use_container_width=True):
                active_source_df["Active"] = True
                replace_project_rows(project_id, active_source_df, logged_in_user)
                save_version(project_id, active_source_df, logged_in_user, "Set all frequencies active")
                log_audit_event(project_id, "set_all_frequencies_active", logged_in_user, {"rows": len(active_source_df)})
                st.success("All rows set active.")
                st.rerun()

        with c_off:
            if st.button("Set all inactive", use_container_width=True):
                active_source_df["Active"] = False
                replace_project_rows(project_id, active_source_df, logged_in_user)
                save_version(project_id, active_source_df, logged_in_user, "Set all frequencies inactive")
                log_audit_event(project_id, "set_all_frequencies_inactive", logged_in_user, {"rows": len(active_source_df)})
                st.warning("All rows set inactive.")
                st.rerun()

        with c_save_active:
            if st.button("Save active/inactive changes", type="primary", use_container_width=True):
                updated = active_source_df.copy()
                if "Active" in edited_active_df.columns:
                    updated["Active"] = edited_active_df["Active"].apply(to_active_bool).values

                replace_project_rows(project_id, updated, logged_in_user)
                save_version(project_id, updated, logged_in_user, "Updated active/inactive frequency status")
                log_audit_event(
                    project_id,
                    "active_inactive_updated",
                    logged_in_user,
                    {
                        "active": int(updated["Active"].apply(to_active_bool).sum()),
                        "inactive": int((~updated["Active"].apply(to_active_bool)).sum()),
                    },
                )
                st.success("Active/inactive frequency status saved.")
                st.rerun()



with st.expander("Import / replace table from file or pasted CSV", expanded=(len(current_df)==0)):
    c1, c2 = st.columns(2)

    with c1:
        uploaded = st.file_uploader("Upload CSV/XLSX", type=["csv", "xlsx", "xls"], key="upload_file_widget")

        if uploaded is not None:
            try:
                file_bytes = uploaded.getvalue()
                file_sig = f"{uploaded.name}-{len(file_bytes)}"

                if st.session_state.get("pending_upload_sig") != file_sig:
                    sheets_dict, active_sheet, file_bytes = read_uploaded_workbook(uploaded)
                    st.session_state["pending_upload_sheets_dict"] = sheets_dict
                    st.session_state["pending_upload_sheet"] = active_sheet
                    st.session_state["pending_upload_df"] = sheets_dict[active_sheet]
                    st.session_state["pending_upload_name"] = uploaded.name
                    st.session_state["pending_upload_bytes"] = file_bytes
                    st.session_state["pending_upload_sig"] = file_sig

            except Exception as e:
                st.error(f"Upload failed: {e}")

        if st.session_state.get("pending_upload_sheets_dict") is not None:
            sheets_dict = st.session_state["pending_upload_sheets_dict"]
            sheet_names = list(sheets_dict.keys())

            if len(sheet_names) > 1:
                current_sheet = st.selectbox(
                    "Workbook sheet preview",
                    sheet_names,
                    index=sheet_names.index(st.session_state.get("pending_upload_sheet", sheet_names[0]))
                    if st.session_state.get("pending_upload_sheet", sheet_names[0]) in sheet_names else 0,
                    key="pending_upload_sheet_select",
                )
                st.session_state["pending_upload_sheet"] = current_sheet
                st.session_state["pending_upload_df"] = sheets_dict[current_sheet]
            else:
                current_sheet = sheet_names[0]

            st.caption(f"Pending upload: {st.session_state.get('pending_upload_name', 'uploaded file')} — {len(sheet_names)} sheet(s)")
            st.dataframe(st.session_state["pending_upload_df"].head(20), use_container_width=True)

            with st.expander("Smart import cleanup / column mapping", expanded=False):
                st.caption("Detected normalized columns. Edit the file headers if anything looks wrong, then re-upload.")
                mapping_preview = pd.DataFrame({
                    "Imported / normalized columns": list(st.session_state["pending_upload_df"].columns)
                })
                st.dataframe(mapping_preview, use_container_width=True)
                if mgrs is None:
                    st.info("MGRS/USNG conversion requires adding `mgrs` to requirements. Without it, Latitude/Longitude must be provided directly.")

            b1, b2 = st.columns(2)
            with b1:
                if st.button("Replace with uploaded workbook/data", type="primary", use_container_width=True, disabled=not can_edit):
                    sheets_dict = st.session_state["pending_upload_sheets_dict"]
                    active_sheet = st.session_state.get("pending_upload_sheet") or list(sheets_dict.keys())[0]
                    tmp = normalize_uploaded_df(sheets_dict[active_sheet].copy())

                    # Save active sheet to the main shared allocation table for plots/deconfliction.
                    replace_project_rows(project_id, tmp, user_name)
                    save_version(project_id, tmp, user_name, f"Imported file — active sheet: {active_sheet}")

                    # Save every sheet so workbook tabs survive logout/login.
                    try:
                        save_project_sheets(project_id, sheets_dict, user_name)
                    except Exception as e:
                        st.warning(f"Main table was saved, but workbook sheets could not be saved: {e}")

                    # Save original file.
                    try:
                        upload_original_file(
                            project_id,
                            st.session_state.get("pending_upload_name", "uploaded_file"),
                            st.session_state.get("pending_upload_bytes"),
                            user_name,
                        )
                    except Exception as e:
                        st.warning(f"Rows were saved, but the original file could not be saved to Storage: {e}")

                    st.session_state["workbook_sheets"] = sheets_dict
                    st.session_state["active_sheet_name"] = active_sheet
                    st.session_state["pending_upload_df"] = None
                    st.session_state["pending_upload_sheets_dict"] = None
                    st.session_state["pending_upload_name"] = None
                    st.session_state["pending_upload_bytes"] = None
                    st.session_state["pending_upload_sig"] = None
                    st.success("Uploaded workbook/data saved.")
                    st.rerun()
            with b2:
                if st.button("Clear pending upload", use_container_width=True):
                    st.session_state["pending_upload_df"] = None
                    st.session_state["pending_upload_sheets_dict"] = None
                    st.session_state["pending_upload_name"] = None
                    st.session_state["pending_upload_bytes"] = None
                    st.session_state["pending_upload_sig"] = None
                    st.session_state["pending_upload_sheet"] = None
                    st.rerun()

    with c2:
        pasted = st.text_area("Paste CSV", height=180, placeholder="Start Time,End Time,Equipment,Center Frequency (MHz),Bandwidth (MHz),Power (W),Tech,Unit")
        if pasted.strip():
            try:
                tmp = normalize_uploaded_df(pd.read_csv(io.StringIO(pasted)))
                st.dataframe(tmp.head(20), use_container_width=True)
                if st.button("Replace with pasted CSV", type="primary", disabled=not can_edit):
                    replace_project_rows(project_id, tmp, user_name)
                    save_version(project_id, tmp, user_name, "Pasted CSV")
                    save_project_sheets(project_id, {"CSV": tmp}, user_name)
                    st.session_state["workbook_sheets"] = {"CSV": tmp}
                    st.session_state["active_sheet_name"] = "CSV"
                    st.success("Pasted CSV saved.")
                    st.rerun()
            except Exception as e:
                st.error(f"Could not parse pasted CSV: {e}")


st.subheader("Shared allocation workbook")

workbook_sheets = st.session_state.get("workbook_sheets") or {"Working": normalize_uploaded_df(current_df)}
sheet_names = list(workbook_sheets.keys())
saved_sheet_count_now = len(load_project_sheets(project_id))
st.caption(f"Workbook tabs in app: {len(sheet_names)} | Workbook tabs saved in Supabase: {saved_sheet_count_now}")

if len(sheet_names) > 1:
    st.caption("Each Excel worksheet is preserved as a tab. Choose the active plotting sheet before saving.")
else:
    st.caption("Single-sheet project. Upload an XLSX with multiple sheets to preserve workbook tabs.")

active_sheet_name = st.selectbox(
    "Active sheet for plots/deconfliction",
    sheet_names,
    index=sheet_names.index(st.session_state.get("active_sheet_name", sheet_names[0]))
    if st.session_state.get("active_sheet_name", sheet_names[0]) in sheet_names else 0,
    key="active_sheet_selector",
)
st.session_state["active_sheet_name"] = active_sheet_name

sheet_tabs = st.tabs(sheet_names)
edited_sheets = {}
for i, sheet_name in enumerate(sheet_names):
    with sheet_tabs[i]:
        edited_sheets[sheet_name] = st.data_editor(
            workbook_sheets[sheet_name],
            use_container_width=True,
            height=330,
            num_rows="dynamic",
            key=f"sheet_editor_{project_id}_{sheet_name}",
            disabled=not can_edit,
        )

edited_df = normalize_uploaded_df(edited_sheets[active_sheet_name])

s1, s2, s3, s4 = st.columns([1, 1, 1.2, 1.2])
with s1:
    if st.button("💾 Save shared changes", type="primary", use_container_width=True, disabled=not can_edit):
        try:
            # Save all workbook tabs and also save active sheet to main allocation table.
            saved_tabs = save_project_sheets(project_id, edited_sheets, user_name)
            replace_project_rows(project_id, edited_df, user_name)
            st.session_state["workbook_sheets"] = edited_sheets
            st.success(f"Saved active sheet and {saved_tabs} workbook tab(s) permanently.")
            st.rerun()
        except Exception as e:
            st.error(f"Could not save workbook tabs. Make sure you ran the v17 SQL setup. Details: {e}")

with s2:
    version_note = st.text_input("Version note", value="Manual save", label_visibility="collapsed")

with s3:
    if st.button("📌 Save version snapshot", use_container_width=True, disabled=not can_edit):
        st.success(f"Saved version {save_version(project_id, edited_df, user_name, version_note)}.")

with s4:
    try:
        st.download_button(
            "Download workbook XLSX",
            data=workbook_to_xlsx_bytes(edited_sheets),
            file_name="spectrum_planner_workbook.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as e:
        st.warning(f"Workbook download unavailable: {e}")

with st.expander("Version history / restore"):
    versions = list_versions(project_id)
    if not versions:
        st.info("No versions saved yet.")
    else:
        st.dataframe(pd.DataFrame(versions)[["version_no", "saved_by", "save_note", "created_at"]], use_container_width=True)
        selected_v = st.selectbox("Restore version", versions, format_func=lambda v: f"v{v['version_no']} — {v.get('save_note') or ''} — {v.get('saved_by') or ''}")
        if st.button("Restore selected version", disabled=not can_edit):
            v = load_version(selected_v["id"])
            restored = normalize_uploaded_df(pd.DataFrame(v["snapshot"]))
            replace_project_rows(project_id, restored, user_name)
            save_project_sheets(project_id, {"Restored": restored}, user_name)
            save_version(project_id, restored, user_name, f"Restored v{selected_v['version_no']}")
            st.session_state["workbook_sheets"] = {"Restored": restored}
            st.session_state["active_sheet_name"] = "Restored"
            st.success("Restored.")
            st.rerun()

try:
    df_ready = prep_df(edited_df)
except Exception as e:
    st.error(f"Could not prepare table for plots: {e}"); st.stop()
if df_ready.empty:
    st.warning("No valid plot rows. Need Start/End Frequency and Power."); st.stop()

grp_ut = unittech_field(df_ready)
pal_equipment = make_palette(sorted(safe_group(df_ready["Equipment"]).unique().tolist()))
pal_unittech = make_palette(sorted(safe_group(df_ready[grp_ut]).unique().tolist()))
map_group_field = map_group_by if map_group_by in df_ready.columns else "Equipment"
pal_map = make_palette(sorted(safe_group(df_ready[map_group_field]).unique().tolist()))
ss, ee = df_ready["StartTime"].apply(parse_time_one), df_ready["EndTime"].apply(parse_time_one)
bad_count = int((ss.isna() | ee.isna()).sum())
base_conf_eq = detect_conflicts_generic(df_ready, "Equipment", guard_mhz)
base_conf_ut = detect_conflicts_generic(df_ready, grp_ut, guard_mhz)
m1, m2, m3 = st.columns(3)
m1.metric("Time parse issues", bad_count); m2.metric("Equipment conflicts", 0 if base_conf_eq is None else len(base_conf_eq)); m3.metric(f"{grp_ut} conflicts", 0 if base_conf_ut is None else len(base_conf_ut))
if bad_count:
    with st.expander("Rows with time parse issues"):
        st.dataframe(df_ready.loc[ss.isna() | ee.isna()], use_container_width=True)

plot_df_conf = df_ready.copy()
if auto_dc:
    anchor = parse_time_one(dc_start)
    if pd.isna(anchor):
        st.error("Deconflict anchor is invalid. Use examples like 4am, 0730, 13:00, or 1:30pm.")
        st.stop()
    st.caption(f"Auto-deconflict anchor applied: {fmt_hhmm(float(anchor))}")
    plot_df_conf = auto_deconflict_smart(
        df_ready,
        float(dc_window) * 3600,
        float(dc_pad_min) * 60,
        float(anchor),
        float(guard_mhz),
        allow_earlier,
        priority_mode,
    )
    if "StartTimeDC" in plot_df_conf.columns:
        plot_df_conf["StartTimeOrig"] = plot_df_conf["StartTime"]; plot_df_conf["EndTimeOrig"] = plot_df_conf["EndTime"]
        plot_df_conf["StartTime"] = np.where(plot_df_conf["StartTimeDC"].notna(), plot_df_conf["StartTimeDC"], plot_df_conf["StartTime"])
        plot_df_conf["EndTime"] = np.where(plot_df_conf["EndTimeDC"].notna(), plot_df_conf["EndTimeDC"], plot_df_conf["EndTime"])
conf_eq = detect_conflicts_generic(plot_df_conf, "Equipment", guard_mhz)
conf_ut = detect_conflicts_generic(plot_df_conf, grp_ut, guard_mhz)


# Safety fallback for older merged code paths.
try:
    guard
except NameError:
    try:
        guard = float(freq_guard)
    except Exception:
        try:
            guard = float(plan_guard_mhz)
        except Exception:
            guard = 0.0



# ---------------- Allocation Builder UI V47 ----------------
with st.expander("Build Allocation Plan from Request Tracker + Approved Frequencies", expanded=False):
    st.markdown(
        "Upload the **Request Tracker** and **Approved Frequencies** file. "
        "The engine will preserve your allocation workbook format and append planning fields to the right."
    )
    build_col1, build_col2 = st.columns(2)
    with build_col1:
        request_tracker_upload = st.file_uploader(
            "Request Tracker (.xlsx)",
            type=["xlsx"],
            key="v47_request_tracker_upload",
        )
    with build_col2:
        approved_freq_upload = st.file_uploader(
            "Approved Frequencies (.xlsx)",
            type=["xlsx"],
            key="v47_approved_freq_upload",
        )

    st.caption("Allocation philosophy: Option B, North/Central/South reusable spectrum pools, geographic reuse first. Output keeps Active as Column A and preserves your spreadsheet column order.")
    if st.button("Build Allocation Plan", type="primary", use_container_width=True, key="v47_build_allocation_plan"):
        if request_tracker_upload is None or approved_freq_upload is None:
            st.warning("Upload both the Request Tracker and Approved Frequencies file first.")
        else:
            try:
                with st.spinner("Building allocation plan..."):
                    output_bytes = build_allocation_workbook_from_request_and_pool(
                        request_tracker_upload,
                        approved_freq_upload,
                    )
                st.success("Allocation plan built successfully.")
                st.download_button(
                    "Download Allocation Plan",
                    data=output_bytes,
                    file_name="PCC6_Built_Allocation_Plan.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Allocation build failed: {e}")


# ---------------- Clean Workspace Header V49 ----------------
st.markdown(
    f"""
    <div class="v49-clean-title">
        <h1>PCC6 Spectrum Planner</h1>
        <p>Project: {project_name} · User: {user_name} · Status: {project_status}</p>
    </div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    """
    <div class="v49-clean-note">
        Use the sidebar for project controls, import/export, and visibility filters. Use the tabs below for allocation views, conflicts, maps, planner tools, and admin functions.
    </div>
    """,
    unsafe_allow_html=True,
)


tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13 = st.tabs([
    "Equipment Power",
    "Equipment Deconfliction",
    "Tech Power",
    "Tech Deconfliction",
    "Unit Deconfliction",
    "Sponsor Deconfliction",
    "Map View",
    "Conflict Tables",
    "Conflict Recommendations",
    "Terrain / Range",
    "Smart Planner",
    "Band Utilization",
    "Allocation Validation",
])

with tab1:
    st.markdown(section_header_html("Equipment Power", "Frequency allocation versus power by equipment."), unsafe_allow_html=True)
    fig = build_power_plot(df_ready, "Equipment", dark, alpha_val, tick_major, tick_minor, int(label_digits), pal_equipment, auto_thin, float(label_gap), high_power_top, power_style, float(outline_lwd), show_center_labels)
    st.pyplot(fig, use_container_width=True)
    st.download_button("Download PNG", fig_to_png_bytes(fig), "equipment_power.png", "image/png")

with tab2:
    st.markdown("""<div class="pcc-panel"><div class="pcc-panel-title">Time × Frequency — By Equipment</div><div class="pcc-panel-caption">Boxes display frequency only; equipment names remain in the legend.</div></div>""", unsafe_allow_html=True)
    fig = build_deconflict_plot(plot_df_conf, "Equipment", pal_equipment, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_eq)
    st.pyplot(fig, use_container_width=True)
    st.download_button("Download PNG", fig_to_png_bytes(fig), "equipment_deconfliction.png", "image/png")

with tab3:
    fig = build_power_plot(df_ready, "Tech", dark, alpha_val, tick_major, tick_minor, int(label_digits), pal_unittech, auto_thin, float(label_gap), high_power_top, power_style, float(outline_lwd), show_center_labels)
    st.pyplot(fig, use_container_width=True)
    st.download_button("Download PNG", fig_to_png_bytes(fig), "tech_power.png", "image/png")

with tab4:
    st.markdown("""<div class="pcc-panel"><div class="pcc-panel-title">Time × Frequency — By Tech</div><div class="pcc-panel-caption">Boxes display frequency only; tech names remain in the legend.</div></div>""", unsafe_allow_html=True)
    fig = build_deconflict_plot(plot_df_conf, "Tech", pal_unittech, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_ut)
    st.pyplot(fig, use_container_width=True)
    st.download_button("Download PNG", fig_to_png_bytes(fig), "tech_deconfliction.png", "image/png")


with tab5:
    st.markdown("""<div class="pcc-panel"><div class="pcc-panel-title">Deconfliction by Unit</div><div class="pcc-panel-caption">Frequency-only labels with unit legend and summary.</div></div>""", unsafe_allow_html=True)
    st.markdown("#### Unit Deconfliction")
    if "Unit" in plot_df_conf.columns:
        unit_series = plot_df_conf["Unit"].fillna("").astype(str).str.strip()
        if unit_series.replace(["", "None", "nan", "(blank)"], pd.NA).isna().all():
            st.warning("Unit column exists, but it is blank. Re-upload the populated allocation workbook or fill Unit values before using this view.")
        plot_df_conf_unit = plot_df_conf.copy()
        plot_df_conf_unit["Unit"] = plot_df_conf_unit["Unit"].fillna("(blank)").replace({"": "(blank)", "None": "(blank)", "nan": "(blank)"})
        pal_unit = make_palette(plot_df_conf_unit["Unit"].astype(str).unique())
        conf_unit = detect_conflicts_generic(plot_df_conf_unit, "Unit", guard)
        fig = build_deconflict_plot(plot_df_conf_unit, "Unit", pal_unit, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_unit)
        st.pyplot(fig, use_container_width=True)
        st.download_button("Download PNG", fig_to_png_bytes(fig), "unit_deconfliction.png", "image/png")
        st.markdown("##### Unit summary")
        st.dataframe(summary_by_group(df_ready, "Unit", conf_unit), use_container_width=True)
    else:
        st.info("No Unit column found.")

with tab6:
    st.markdown("""<div class="pcc-panel"><div class="pcc-panel-title">Deconfliction by Sponsor</div><div class="pcc-panel-caption">Frequency-only labels with sponsor legend and summary.</div></div>""", unsafe_allow_html=True)
    st.markdown("#### Sponsor Deconfliction")
    sponsor_col = "Sponsor" if "Sponsor" in plot_df_conf.columns else ("Sponser" if "Sponser" in plot_df_conf.columns else None)
    if sponsor_col:
        sponsor_series = plot_df_conf[sponsor_col].fillna("").astype(str).str.strip()
        if sponsor_series.replace(["", "None", "nan", "(blank)"], pd.NA).isna().all():
            st.warning("Sponsor column exists, but it is blank. Re-upload the populated allocation workbook or fill Sponsor values before using this view.")
        plot_df_conf_sponsor = plot_df_conf.copy()
        plot_df_conf_sponsor[sponsor_col] = plot_df_conf_sponsor[sponsor_col].fillna("(blank)").replace({"": "(blank)", "None": "(blank)", "nan": "(blank)"})
        pal_sponsor = make_palette(plot_df_conf_sponsor[sponsor_col].astype(str).unique())
        conf_sponsor = detect_conflicts_generic(plot_df_conf_sponsor, sponsor_col, guard)
        fig = build_deconflict_plot(plot_df_conf_sponsor, sponsor_col, pal_sponsor, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_sponsor)
        st.pyplot(fig, use_container_width=True)
        st.download_button("Download PNG", fig_to_png_bytes(fig), "sponsor_deconfliction.png", "image/png")
        st.markdown("##### Sponsor summary")
        st.dataframe(summary_by_group(df_ready, sponsor_col, conf_sponsor), use_container_width=True)
    else:
        st.info("No Sponsor/Sponser column found.")


with tab7:
    st.markdown(section_header_html("Map View", "Mapped allocations, coverage circles, heat map, and export tools."), unsafe_allow_html=True)
    st.markdown("#### Map View")
    st.caption("Uses decimal-degree Latitude and Longitude columns. Coverage Radius draws circles using the selected units. Basemap uses free CARTO tiles; no Mapbox/Google token required.")
    deck, map_df = build_map_deck(
        df_ready,
        map_group_field,
        pal_map,
        radius_units=radius_units,
        show_coverage=show_coverage,
        map_style=map_style_choice,
        show_heatmap=show_heatmap,
        heatmap_weight_by=heatmap_weight_by,
    )

    if deck is None:
        st.info("No valid map rows found. Add Latitude and Longitude columns with decimal-degree values, for example 31.12345 and -97.12345.")
        st.dataframe(
            pd.DataFrame({
                "Required": ["Latitude", "Longitude"],
                "Recommended": ["Location", "Site Name"],
                "Optional": ["Coverage Radius", "Antenna Height"],
            }),
            use_container_width=True,
        )
    else:
        st.pydeck_chart(deck, use_container_width=True)

        d1, d2, d3, d4 = st.columns(4)
        with d1:
            st.download_button(
                "Download map HTML",
                data=deck_to_html_bytes(deck),
                file_name="spectrum_map.html",
                mime="text/html",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "Download map data CSV",
                data=map_df.to_csv(index=False).encode("utf-8"),
                file_name="spectrum_map_data.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with d3:
            st.download_button(
                "Download KML",
                data=map_df_to_kml(map_df, project_name, radius_units, show_coverage),
                file_name=f"{safe_storage_filename(project_name)}_map.kml",
                mime="application/vnd.google-earth.kml+xml",
                use_container_width=True,
            )
        with d4:
            st.download_button(
                "Download GeoJSON",
                data=map_df_to_geojson(map_df, project_name),
                file_name=f"{safe_storage_filename(project_name)}_map.geojson",
                mime="application/geo+json",
                use_container_width=True,
            )

        st.markdown("#### Map rows")
        display_cols = [c for c in ["Active"] + [c for c in ["Equipment", "Tech", "Unit", "Latitude", "Longitude", "MGRS", "USNG", "Location", "SiteName", "CoverageRadius", "AntennaHeight", "CenterF", "PowerW", "StartTime", "EndTime"] if c in map_df.columns] if c in map_df.columns]
        st.dataframe(map_df[display_cols], use_container_width=True)

        with st.expander("Map congestion summary", expanded=False):
            st.caption("Ranks locations/sites by row count, power, coverage, and unique frequencies.")
            st.dataframe(map_congestion_summary(map_df), use_container_width=True)

with tab8:
    st.markdown(section_header_html("Conflict Tables", "Equipment, tech, unit, and sponsor conflict records."), unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### Equipment conflicts")
        st.dataframe(conflict_summary(conf_eq), use_container_width=True)
    with c2:
        st.markdown(f"#### {grp_ut} conflicts")
        st.dataframe(conflict_summary(conf_ut), use_container_width=True)

    if auto_dc and "ShiftSec" in plot_df_conf.columns:
        st.markdown("#### Auto-deconflict moves")
        moves = plot_df_conf.copy()
        moves["ShiftMin"] = moves["ShiftSec"].fillna(0) / 60
        moves = moves.loc[abs(moves["ShiftMin"]) > .001]
        cols = [c for c in ["Equipment", "Tech", "Unit", "StartF", "EndF", "StartTimeOrig", "EndTimeOrig", "StartTimeDC", "EndTimeDC", "ShiftMin", "Placed"] if c in moves.columns]
        st.dataframe(moves[cols] if not moves.empty else pd.DataFrame({"Message": ["No rows were moved."]}), use_container_width=True)



with tab9:
    st.markdown("#### Conflict Severity + Recommended Actions")
    rec_df = combined_conflict_recommendations(conf_eq, conf_ut, grp_ut)

    if "Message" in rec_df.columns:
        st.info(rec_df["Message"].iloc[0])
    else:
        high_count = int((rec_df["Severity"] == "High").sum())
        medium_count = int((rec_df["Severity"] == "Medium").sum())
        low_count = int((rec_df["Severity"] == "Low").sum())

        c_high, c_med, c_low = st.columns(3)
        c_high.metric("High", high_count)
        c_med.metric("Medium", medium_count)
        c_low.metric("Low", low_count)

        st.dataframe(rec_df, use_container_width=True)

        st.download_button(
            "Download conflict recommendations CSV",
            data=rec_df.to_csv(index=False).encode("utf-8"),
            file_name="conflict_recommendations.csv",
            mime="text/csv",
            use_container_width=True,
        )

        log_audit_event(project_id, "conflict_recommendations_viewed", logged_in_user, {"rows": len(rec_df)})




with tab10:
    st.markdown("#### Terrain / Range Planning")
    st.caption("This is an approximate RF-horizon screening tool using antenna height and great-circle distance. It does not use live elevation terrain data yet.")

    deck_for_range, map_df_for_range = build_map_deck(
        df_ready,
        map_group_field,
        pal_map,
        radius_units=radius_units,
        show_coverage=show_coverage,
        map_style=map_style_choice,
        show_heatmap=False,
        heatmap_weight_by=heatmap_weight_by,
    )

    if map_df_for_range is None or map_df_for_range.empty:
        st.info("Add valid Latitude and Longitude values to use terrain/range analysis.")
    else:
        if "AntennaHeight" not in map_df_for_range.columns:
            map_df_for_range["AntennaHeight"] = default_ant_height_m

        range_df = build_range_analysis(map_df_for_range)

        if "Message" in range_df.columns:
            st.info(range_df["Message"].iloc[0])
        else:
            likely_count = int((range_df["LOS Screen"] == "Likely LOS").sum())
            marginal_count = int((range_df["LOS Screen"] == "Marginal LOS").sum())
            blocked_count = int((range_df["LOS Screen"] == "Beyond horizon").sum())

            c1, c2, c3 = st.columns(3)
            c1.metric("Likely LOS", likely_count)
            c2.metric("Marginal", marginal_count)
            c3.metric("Beyond horizon", blocked_count)

            if show_range_warning and blocked_count:
                st.warning("Some paths are beyond the approximate RF horizon. Increase antenna height, add relay sites, or reduce path distance.")

            st.dataframe(range_df, use_container_width=True)

            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    "Download range analysis CSV",
                    data=range_df.to_csv(index=False).encode("utf-8"),
                    file_name="range_analysis.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with d2:
                st.download_button(
                    "Download path GeoJSON",
                    data=path_lines_geojson(range_df, map_df_for_range, project_name),
                    file_name=f"{safe_storage_filename(project_name)}_paths.geojson",
                    mime="application/geo+json",
                    use_container_width=True,
                )

            log_audit_event(project_id, "range_analysis_viewed", logged_in_user, {"paths": len(range_df)})




with tab11:
    st.markdown(section_header_html("Smart Frequency Planner", "Recommended frequency moves and time-shift fallback actions."), unsafe_allow_html=True)
    st.markdown("#### Smart Frequency Planner")
    st.caption("Suggests alternate frequency ranges and time fallback actions using the current conflicts, planning band, guard band, and priority rules.")

    plan_df = combined_smart_plan(
        df_ready,
        conf_eq,
        conf_ut,
        grp_ut,
        plan_band_start,
        plan_band_end,
        float(plan_guard_mhz),
        float(plan_step_mhz),
        int(max_suggestions),
    )

    if "Message" in plan_df.columns:
        st.info(plan_df["Message"].iloc[0])
    else:
        freq_moves = int((plan_df["Plan Type"] == "Frequency move").sum())
        time_moves = int((plan_df["Plan Type"] == "Time move").sum())
        c1, c2 = st.columns(2)
        c1.metric("Frequency move options", freq_moves)
        c2.metric("Time fallback options", time_moves)

        st.dataframe(plan_df, use_container_width=True)

        st.download_button(
            "Download smart plan CSV",
            data=plan_df.to_csv(index=False).encode("utf-8"),
            file_name="smart_frequency_plan.csv",
            mime="text/csv",
            use_container_width=True,
        )

        st.markdown("#### Apply selected recommendation")
        st.warning("Review carefully before applying. This updates the active workbook sheet and saves a new version snapshot.")

        option_labels = [
            f"{i}: {row['Plan Type']} | {row['Move Group']} | {row['Conflict']} | {row['Reason'][:90]}"
            for i, row in plan_df.reset_index(drop=True).iterrows()
        ]

        selected_plan_label = st.selectbox("Recommendation to apply", option_labels, key=f"apply_plan_select_{project_id}")
        selected_plan_idx = int(selected_plan_label.split(":", 1)[0])
        selected_plan = plan_df.reset_index(drop=True).iloc[selected_plan_idx].to_dict()

        st.json({k: json_safe_value(v) for k, v in selected_plan.items()})

        confirm_apply = st.checkbox(
            "I reviewed this recommendation and want to apply it to the active sheet",
            key=f"confirm_apply_smart_plan_{project_id}",
        )

        if st.button(
            "Apply selected smart plan",
            type="primary",
            use_container_width=True,
            disabled=not (can_edit and confirm_apply),
        ):
            try:
                active_sheet = st.session_state.get("active_sheet_name") or list(edited_sheets.keys())[0]
                base_sheet = edited_sheets.get(active_sheet, edited_df)

                updated_sheet, apply_note = apply_smart_plan_to_sheet(base_sheet, selected_plan)

                updated_sheets = dict(edited_sheets)
                updated_sheets[active_sheet] = updated_sheet

                save_project_sheets(project_id, updated_sheets, logged_in_user)
                replace_project_rows(project_id, updated_sheet, logged_in_user)
                version_no = save_version(project_id, updated_sheet, logged_in_user, f"Applied smart plan: {apply_note}")

                st.session_state["workbook_sheets"] = updated_sheets
                st.session_state["active_sheet_name"] = active_sheet

                log_audit_event(
                    project_id,
                    "smart_plan_applied",
                    logged_in_user,
                    {
                        "active_sheet": active_sheet,
                        "version_no": version_no,
                        "note": apply_note,
                        "recommendation": {k: json_safe_value(v) for k, v in selected_plan.items()},
                    },
                )

                st.success(f"{apply_note} Saved version {version_no}.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not apply selected recommendation: {e}")

        log_audit_event(project_id, "smart_frequency_plan_generated", logged_in_user, {"rows": len(plan_df)})



with tab12:
    st.markdown("#### Band Utilization Dashboard")
    band_df = band_utilization_summary(df_ready)
    st.dataframe(band_df, use_container_width=True)
    if "Message" not in band_df.columns:
        st.download_button("Download band utilization CSV", band_df.to_csv(index=False).encode("utf-8"), "band_utilization.csv", "text/csv", use_container_width=True)

with tab13:
    st.markdown("#### Allocation Validation Center")
    val_df = allocation_validation_summary(df_ready)
    if "Message" in val_df.columns:
        st.success(val_df["Message"].iloc[0])
    else:
        st.warning(f"{len(val_df)} allocation row(s) need review.")
        st.dataframe(val_df, use_container_width=True)
        st.download_button("Download validation issues CSV", val_df.to_csv(index=False).encode("utf-8"), "allocation_validation_issues.csv", "text/csv", use_container_width=True)

    st.markdown("#### Geographic Reuse Quick Look")
    reuse_df = geographic_reuse_summary(df_ready)
    st.dataframe(reuse_df, use_container_width=True)
    if "Message" not in reuse_df.columns:
        st.download_button("Download geographic reuse CSV", reuse_df.to_csv(index=False).encode("utf-8"), "geographic_reuse.csv", "text/csv", use_container_width=True)


# ---------------- Briefing export ----------------
with st.expander("Export briefing PDF", expanded=False):
    st.caption("Exports a PDF briefing using the current active plotting sheet and current plot settings.")
    if st.button("Build PDF briefing", use_container_width=True):
        try:
            pdf_figs = [
                build_power_plot(df_ready, "Equipment", dark, alpha_val, tick_major, tick_minor, int(label_digits), pal_equipment, auto_thin, float(label_gap), high_power_top, power_style, float(outline_lwd), show_center_labels),
                build_deconflict_plot(plot_df_conf, "Equipment", pal_equipment, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_eq),
                build_power_plot(df_ready, grp_ut, dark, alpha_val, tick_major, tick_minor, int(label_digits), pal_unittech, auto_thin, float(label_gap), high_power_top, power_style, float(outline_lwd), show_center_labels),
                build_deconflict_plot(plot_df_conf, grp_ut, pal_unittech, dark, tick_major, tick_minor, box_labels, box_label_min_height_min, show_shift_label, moved_outline, conf_ut),
            ]
            pdf_bytes = briefing_pdf_bytes(project_name, status_info, df_ready, conf_eq, conf_ut, pdf_figs)
            for _fig in pdf_figs:
                plt.close(_fig)
            st.download_button(
                "Download briefing PDF",
                data=pdf_bytes,
                file_name=f"{safe_storage_filename(project_name)}_briefing.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
            log_audit_event(project_id, "briefing_pdf_generated", logged_in_user, {"project": project_name})
        except Exception as e:
            st.error(f"Could not build PDF briefing: {e}")
